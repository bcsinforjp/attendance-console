#!/usr/bin/env python3
"""
Attendance PDF to Excel Converter - FastAPI Backend
Converts Japanese attendance PDFs to Excel files
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Query, Body, BackgroundTasks
from fastapi.responses import FileResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from calendar import monthrange
from datetime import datetime, timedelta
import json
import os
from pathlib import Path
import re
import shutil
import threading
import time
from uuid import uuid4
import zipfile

try:
    import psutil  # type: ignore
    HAVE_PSUTIL = True
except ImportError:  # pragma: no cover
    psutil = None  # type: ignore
    HAVE_PSUTIL = False

# Excel and PDF libraries
import unicodedata
import uuid
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pdfplumber
import psycopg2
from psycopg2.extras import RealDictCursor

# Create app
app = FastAPI(title="V3 Attendance Console", version="4.1")

# External company-to-company data API — self-contained separate module
# (deny-by-default; managed from /admin → Partner API).
from partner_api import (
    router as partner_router,
    partner_status as _partner_status,
    partner_set_enabled as _partner_set_enabled,
    partner_add as _partner_add,
    partner_update as _partner_update,
    partner_revoke as _partner_revoke,
    partner_log_tail as _partner_log_tail,
    register_partner_error_handler as _register_partner_errs,
)
app.include_router(partner_router)
_register_partner_errs(app)   # scoped enveloped errors for /partner/v1 only

# Unified AI-Agent API (REST + MCP) — self-contained module (deny-by-default;
# managed from /admin → Agent API). Part A: read/write data. Part B: change-mgmt.
from agent_api import (
    router as agent_router,
    mcp_post as _agent_mcp_post,
    mcp_get as _agent_mcp_get,
    register_agent_error_handler as _register_agent_errs,
    agent_status as _agent_status,
    agent_set_enabled as _agent_set_enabled,
    agent_add as _agent_add,
    agent_update as _agent_update,
    agent_revoke as _agent_revoke,
    agent_log_tail as _agent_log_tail,
)
import change_requests as _change_requests
app.include_router(agent_router)
_register_agent_errs(app)              # scoped enveloped errors for /api/agent/v1

# MCP (Model Context Protocol) front door — JSON-RPC over plain HTTP. Explicit
# routes (not app.mount) so POST /mcp has no trailing-slash redirect.
app.add_api_route("/mcp", _agent_mcp_post, methods=["POST"], include_in_schema=False)
app.add_api_route("/mcp", _agent_mcp_get, methods=["GET"], include_in_schema=False)

# OAuth 2.1 authorization server for the MCP connector (lets Claude's "Add custom
# connector" sign in instead of pasting a token). Tokens map to a managed agent.
import agent_api as _agent_api_mod
from mcp_oauth import router as _mcp_oauth_router, resolve_token as _mcp_oauth_resolve
app.include_router(_mcp_oauth_router)
_agent_api_mod.set_oauth_resolver(_mcp_oauth_resolve)

BASE_DIR = Path(__file__).resolve().parent
EMPLOYEE_ROSTER_PATH = BASE_DIR / "employee_roster.json"
EMPLOYEE_ROSTER = json.loads(EMPLOYEE_ROSTER_PATH.read_text(encoding="utf-8"))
EMPLOYEE_ROSTER_BY_ID = {
    employee["employee_code"]: employee["name"]
    for employee in EMPLOYEE_ROSTER
}
EMPLOYEE_CODES = set(EMPLOYEE_ROSTER_BY_ID)
EXCEL_SECTION_INSERT_AFTER = "00000326"
EXCEL_SECTION_INSERT_BEFORE = "00000401"
EXCEL_SECTION_LABEL = "Section Two 2 Depanment"
SECTIONS_PATH = BASE_DIR / "sections.json"
SECTIONS = json.loads(SECTIONS_PATH.read_text(encoding="utf-8"))["sections"]


def clean_code(raw: str) -> str:
    """Strip the '*' display-only marker from a sections.json code entry."""
    return raw.rstrip("*").strip()


def _build_section_indexes(sections: list[dict]) -> tuple[dict, set, set, list, dict]:
    """Derive every section-related lookup from a sections.json `sections` list.

    Returns (section_of_code, muted_codes, active_codes, muted_placements, position_in_section)
      • section_of_code: code -> sid. Each code's "home" section. First non-muted
        appearance wins; if a code only ever appears as muted, its first appearance wins.
        DB attendance rows are routed here.
      • muted_codes: every clean code that has at least one '*' placement anywhere.
      • active_codes: every clean code that has at least one non-'*' placement.
        DB rows are dropped entirely if a code has only muted placements (no active home).
      • muted_placements: list of (clean_code, sid). One entry per '*' appearance.
        Each becomes a "Disabled" ghost row in that section, independent of where the
        code's active home is. Lets the same employee show as a ghost in section 1 and
        as the real worker in section 2 (or wherever else).
      • position_in_section: {sid: {code: pos}} — first position wins per code per section.
        Drives in-bucket sort order so each section preserves the user's hand-edited order.
    """
    section_of_code: dict[str, int] = {}
    for s in sections:
        for raw in s["codes"]:
            c = clean_code(raw)
            if c not in section_of_code and not raw.endswith("*"):
                section_of_code[c] = s["id"]
    for s in sections:
        for raw in s["codes"]:
            c = clean_code(raw)
            if c not in section_of_code:
                section_of_code[c] = s["id"]

    muted_codes = {
        clean_code(raw) for s in sections for raw in s["codes"] if raw.endswith("*")
    }
    active_codes = {
        clean_code(raw) for s in sections for raw in s["codes"] if not raw.endswith("*")
    }
    muted_placements = [
        (clean_code(raw), s["id"])
        for s in sections
        for raw in s["codes"]
        if raw.endswith("*")
    ]

    pos_by_sec: dict[int, dict[str, int]] = {0: {}}
    for s in sections:
        sid = s["id"]
        pos_by_sec[sid] = {}
        for pos, raw in enumerate(s["codes"]):
            c = clean_code(raw)
            if c not in pos_by_sec[sid]:
                pos_by_sec[sid][c] = pos

    return section_of_code, muted_codes, active_codes, muted_placements, pos_by_sec


SECTION_OF_CODE, MUTED_CODES, ACTIVE_CODES, MUTED_PLACEMENTS, SECTION_POSITIONS = (
    _build_section_indexes(SECTIONS)
)
SECTION_LABEL_BY_ID = {section["id"]: section["label"] for section in SECTIONS}


def is_muted_code(code: str) -> bool:
    return clean_code(code) in MUTED_CODES


def name_for_code(code: str) -> str:
    # "???" surfaces typos in sections.json visually instead of silently dropping the row.
    return EMPLOYEE_ROSTER_BY_ID.get(clean_code(code), "???")


def iter_codes_in_section_order():
    """Yield each unique (code, section_id, section_label) in sections.json order.

    Each code is yielded once at its **effective** section (where SECTION_OF_CODE
    points it). Cross-section ghost placements are skipped here so Excel exports,
    member lists, and the management UI never show the same employee twice.
    Followed by roster codes that aren't in any section (section_id=0 / Unassigned).
    """
    seen: set[str] = set()
    for section in SECTIONS:
        sid = section["id"]
        label = section["label"]
        for raw in section["codes"]:
            code = clean_code(raw)
            if code in seen:
                continue
            # Skip if this isn't the code's effective home section (it's a ghost here).
            if SECTION_OF_CODE.get(code) != sid:
                continue
            seen.add(code)
            yield code, sid, label
    for employee in EMPLOYEE_ROSTER:
        code = employee["employee_code"]
        if code not in seen:
            seen.add(code)
            yield code, 0, "Unassigned"


SECTION_TEXT_PATTERNS = {
    1: [
        re.compile(r"製造\s*[1１一]\s*課"),
        re.compile(r"製造\s*1\s*課"),
    ],
    2: [
        re.compile(r"製造\s*[2２二]\s*課"),
        re.compile(r"製造\s*2\s*課"),
    ],
}
STATIC_DIR = BASE_DIR / "static"
STATIC_DIR.mkdir(exist_ok=True)
DATABASE_HOST = os.getenv("ATTENDANCE_DB_HOST", "127.0.0.1")
DATABASE_PORT = int(os.getenv("ATTENDANCE_DB_PORT", "5432"))
DATABASE_NAME = os.getenv("ATTENDANCE_DB_NAME", "attendance_db")
DATABASE_USER = os.getenv("ATTENDANCE_DB_USER", "attendance")
DATABASE_PASSWORD = os.getenv("ATTENDANCE_DB_PASSWORD", "attendance2026")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create upload directory
UPLOAD_DIR = Path("/tmp/attendance_uploads")
EXPORT_DIR = Path("/tmp/attendance_exports")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

uploaded_files = []
parsed_data = []

# ============================================================================
# Console build (2026-04-27): API keys, access tracking, auto-upload config,
# admin panel. Kept inline in main.py per project convention (no new modules).
# ============================================================================
import hmac
import secrets
import hashlib
from fastapi import Depends, Header, Request, Response, Form
from fastapi.responses import JSONResponse, HTMLResponse
from starlette.middleware.base import BaseHTTPMiddleware

API_KEYS_PATH = BASE_DIR / "api_keys.json"
AUTO_UPLOAD_CONFIG_PATH = BASE_DIR / "auto_upload_config.json"
ADMIN_CONFIG_PATH = BASE_DIR / "admin_config.json"
ACCESS_LOG_DIR = Path(os.environ.get("AI_SERVER_LOG_DIR", "/var/log/ai_server"))
try:
    ACCESS_LOG_DIR.mkdir(parents=True, exist_ok=True)
except PermissionError:
    ACCESS_LOG_DIR = BASE_DIR / "logs"
    ACCESS_LOG_DIR.mkdir(parents=True, exist_ok=True)
ACCESS_LOG_PATH = ACCESS_LOG_DIR / "access.jsonl"
API_DEBUG_LOG_PATH = ACCESS_LOG_DIR / "api_debug.jsonl"
KNOWN_CLIENTS_PATH = BASE_DIR / "known_clients.json"
KNOWN_IPS_PATH = BASE_DIR / "known_ips.json"
ANNOUNCEMENT_PATH = BASE_DIR / "announcement.json"
AGENT_LOG_PATH = ACCESS_LOG_DIR / "agent_requests.jsonl"
APP_UPDATE_DIR = BASE_DIR / "app_updates"
APP_UPDATE_CONFIG_PATH = BASE_DIR / "app_update.json"
DESKTOP_SPEC_PATH = BASE_DIR / "DESKTOP_AGENT_INTEGRATION.md"
PARTNER_SPEC_PATH = BASE_DIR / "PARTNER_API.md"

def _load_json(path: Path, default):
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        print(f"[CONFIG] Failed to read {path}: {exc}")
    return default

def _save_json(path: Path, payload) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)

# ---- API Keys ---------------------------------------------------------------
# Three named keys: TEST (development), APP (mobile/CLI clients), WEB (browser).
# Stored as plain strings in api_keys.json; first run auto-generates safe defaults.
def _bootstrap_api_keys() -> dict:
    if API_KEYS_PATH.exists():
        return _load_json(API_KEYS_PATH, {})
    payload = {
        "TEST": "test-" + secrets.token_urlsafe(12),
        "APP":  "app-"  + secrets.token_urlsafe(20),
        "WEB":  "web-"  + secrets.token_urlsafe(20),
        "_note": "Send via header  X-API-Key: <value>.  Edit values here to rotate.",
    }
    _save_json(API_KEYS_PATH, payload)
    try:
        os.chmod(API_KEYS_PATH, 0o600)
    except Exception:
        pass
    return payload

API_KEYS = _bootstrap_api_keys()
print(f"[API_KEYS] Loaded {sum(1 for k,v in API_KEYS.items() if not k.startswith('_'))} keys from {API_KEYS_PATH}")

def _resolve_key_label(presented: str | None) -> str | None:
    if not presented:
        return None
    for label, value in API_KEYS.items():
        if label.startswith("_"):
            continue
        if isinstance(value, str) and hmac.compare_digest(value, presented):
            return label
    return None

async def require_api_key(x_api_key: str | None = Header(default=None)) -> str:
    label = _resolve_key_label(x_api_key)
    if label is None:
        raise HTTPException(status_code=401, detail="Missing or invalid X-API-Key header.")
    return label

# ---- Admin password ---------------------------------------------------------
def _bootstrap_admin_config() -> dict:
    if ADMIN_CONFIG_PATH.exists():
        return _load_json(ADMIN_CONFIG_PATH, {})
    pw = os.environ.get("ADMIN_PASSWORD", "admin-" + secrets.token_urlsafe(8))
    payload = {
        "password": pw,
        "session_token": secrets.token_urlsafe(24),
        "_note": "Change 'password' to set the /admin login. session_token rotates on logout.",
    }
    _save_json(ADMIN_CONFIG_PATH, payload)
    try:
        os.chmod(ADMIN_CONFIG_PATH, 0o600)
    except Exception:
        pass
    return payload

ADMIN_CONFIG = _bootstrap_admin_config()
print(f"[ADMIN] Config loaded from {ADMIN_CONFIG_PATH}")

def _admin_session_valid(token: str | None) -> bool:
    if not token:
        return False
    return hmac.compare_digest(ADMIN_CONFIG.get("session_token", ""), token)


# ---- Desktop Agent: request log, live state, pause switch ------------------
# Every /api/v1/* call is appended to agent_requests.jsonl and folded into
# AGENT_STATE (used by the /admin Agent tab + the header pill via /api/health).
# INGESTION_PAUSED (admin-toggled, persisted in admin_config.json) makes the
# precheck answer {"action":"paused"} and uploads return HTTP 423 Locked.
AGENT_ONLINE_WINDOW_SEC = 1800  # "online" if a v1 call within this many seconds

def _initial_ingestion_paused() -> bool:
    env = os.environ.get("ATTENDANCE_INGESTION_PAUSED")
    if env is not None:
        return env.strip() in ("1", "true", "TRUE", "yes", "on")
    return bool(ADMIN_CONFIG.get("ingestion_paused", False))

INGESTION_PAUSED = _initial_ingestion_paused()
print(f"[AGENT] ingestion_paused={INGESTION_PAUSED}  log={AGENT_LOG_PATH}")

_agent_lock = threading.Lock()
AGENT_STATE: dict = {
    "last_seen": None, "last_ip": None, "last_key": None,
    "last_precheck": None, "last_upload": None,
    "last_error": None, "last_error_at": None,
    "calls": 0, "errors": 0, "agent_version": None,
}

# Desktop-app update channel: admin uploads a file (any type) into APP_UPDATE_DIR;
# metadata persists in app_update.json; the agent polls /api/v1/app-update.
try:
    APP_UPDATE_DIR.mkdir(parents=True, exist_ok=True)
except Exception as _exc:
    print(f"[APP_UPDATE] mkdir failed: {_exc}")
APP_UPDATE_META: dict = _load_json(APP_UPDATE_CONFIG_PATH, {}) or {}

def _record_agent(path: str, method: str, status: int, ip: str,
                   key_label: str | None, query: str, dur_ms: int,
                   agent_version: str | None = None) -> None:
    """Append one Desktop Agent API call to the log + update live state."""
    ts = datetime.now().isoformat(timespec="seconds")
    entry = {
        "ts": ts, "ip": ip, "key": key_label or "?",
        "method": method, "path": path, "query": (query or "")[:300],
        "status": status, "duration_ms": dur_ms,
    }
    try:
        with AGENT_LOG_PATH.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry, ensure_ascii=False) + "\n")
        if AGENT_LOG_PATH.stat().st_size > 5 * 1024 * 1024:  # ~5 MB cap
            lines = AGENT_LOG_PATH.read_text(encoding="utf-8").splitlines()[-4000:]
            AGENT_LOG_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    except Exception as exc:
        print(f"[AGENT_LOG] write failed: {exc}")
    with _agent_lock:
        AGENT_STATE["last_seen"] = ts
        AGENT_STATE["last_ip"] = ip
        AGENT_STATE["last_key"] = key_label or "?"
        AGENT_STATE["calls"] += 1
        if "/status/precheck" in path:
            AGENT_STATE["last_precheck"] = ts
        if "/upload" in path and method == "POST" and status < 400:
            AGENT_STATE["last_upload"] = {"ts": ts, "path": path}
        if agent_version:
            AGENT_STATE["agent_version"] = agent_version[:60]
        if status >= 400:
            AGENT_STATE["errors"] += 1
            AGENT_STATE["last_error"] = {"ts": ts, "path": path, "status": status}
            AGENT_STATE["last_error_at"] = ts

def _agent_status_summary() -> dict:
    with _agent_lock:
        s = dict(AGENT_STATE)
    online = False
    if s.get("last_seen"):
        try:
            age = (datetime.now() - datetime.fromisoformat(s["last_seen"])).total_seconds()
            online = age <= AGENT_ONLINE_WINDOW_SEC
        except Exception:
            online = False
    s["online"] = online
    s["paused"] = INGESTION_PAUSED
    s["update"] = dict(APP_UPDATE_META)
    s["lan_upload_url"] = f"http://{_detect_lan_ip()}/lan-upload"
    return s


# ============================================================================
# GenbaLink multi-user auth (separate from /admin's single-password scheme).
# Stores users in users.json with PBKDF2-SHA256 hashes (200k iterations, no
# bcrypt dependency). Sessions live in-memory — restart = everyone re-logs-in.
# Auth is host-gated: requests arriving with Host=genbafms.com require a valid
# `gbl_session` cookie; existing /attendance/* access (default vhost) stays
# unauthenticated so office bookmarks keep working.
# ============================================================================
USERS_PATH = BASE_DIR / "users.json"
GBL_HASH_ALGO = "pbkdf2_sha256"
GBL_HASH_ITERS = 200_000
GBL_SESSION_TTL_SECONDS = 12 * 3600
GBL_AUTH_HOSTS = {"genbafms.com", "www.genbafms.com", "link.genbafms.com"}  # host-gated auth


def _gbl_hash_password(password: str, salt: str | None = None) -> str:
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt), GBL_HASH_ITERS).hex()
    return f"{GBL_HASH_ALGO}${GBL_HASH_ITERS}${salt}${h}"


def _gbl_verify_password(password: str, stored: str) -> bool:
    try:
        algo, iters_s, salt, expected = stored.split("$", 3)
        if algo != GBL_HASH_ALGO:
            return False
        candidate = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), bytes.fromhex(salt), int(iters_s)).hex()
        return hmac.compare_digest(candidate, expected)
    except Exception:
        return False


def _gbl_load_users() -> list[dict]:
    if not USERS_PATH.exists():
        return []
    try:
        data = json.loads(USERS_PATH.read_text(encoding="utf-8"))
        return list(data.get("users") or [])
    except Exception as exc:
        print(f"[GBL_AUTH] users.json load error: {exc}")
        return []


def _gbl_save_users(users: list[dict]) -> None:
    payload = {
        "_note": "GenbaLink user store. Passwords are PBKDF2-SHA256, 200k iterations. Edit via Setup UI; do NOT hand-edit hashes.",
        "users": users,
    }
    tmp = USERS_PATH.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    os.replace(tmp, USERS_PATH)
    try:
        os.chmod(USERS_PATH, 0o600)
    except Exception:
        pass


GBL_USERS: list[dict] = _gbl_load_users()
print(f"[GBL_AUTH] Loaded {len(GBL_USERS)} user(s) from {USERS_PATH}")

# Sessions: token -> {"username", "is_admin", "created_at", "expires_at"}
_gbl_sessions: dict[str, dict] = {}
_gbl_sessions_lock = threading.Lock()


def _gbl_session_create(username: str, is_admin: bool) -> str:
    import time
    token = secrets.token_urlsafe(32)
    now = time.time()
    with _gbl_sessions_lock:
        _gbl_sessions[token] = {
            "username": username,
            "is_admin": is_admin,
            "created_at": now,
            "expires_at": now + GBL_SESSION_TTL_SECONDS,
        }
        # Opportunistic cleanup of expired sessions on every create.
        for tok in [t for t, s in _gbl_sessions.items() if s["expires_at"] < now]:
            del _gbl_sessions[tok]
    return token


def _gbl_session_lookup(token: str | None) -> dict | None:
    import time
    if not token:
        return None
    with _gbl_sessions_lock:
        s = _gbl_sessions.get(token)
        if not s:
            return None
        if s["expires_at"] < time.time():
            del _gbl_sessions[token]
            return None
        return dict(s)


def _gbl_session_revoke(token: str | None) -> None:
    if not token:
        return
    with _gbl_sessions_lock:
        _gbl_sessions.pop(token, None)


def _gbl_request_user(request: "Request") -> dict | None:
    return _gbl_session_lookup(request.cookies.get("gbl_session"))


# Short-lived, single-use tokens for browser auto-login. The desktop agent
# exchanges its X-API-Key (sent in the HEADER, never a URL) for one of these
# via POST /api/v1/auth/login-token, then opens /auto-login?token=... in the
# user's default browser. 60s TTL + consumed on first use — this keeps the
# long-lived, high-privilege API key out of browser history, Referer headers
# and access/agent logs. token -> {"key_label", "expires_at"}
_LOGIN_TOKEN_TTL_SECONDS = 60
_login_tokens: dict[str, dict] = {}
_login_tokens_lock = threading.Lock()


def _login_token_create(key_label: str) -> str:
    token = secrets.token_urlsafe(32)
    now = time.time()
    with _login_tokens_lock:
        _login_tokens[token] = {"key_label": key_label,
                                "expires_at": now + _LOGIN_TOKEN_TTL_SECONDS}
        # Opportunistic cleanup of expired tokens on every create.
        for tok in [t for t, s in _login_tokens.items() if s["expires_at"] < now]:
            del _login_tokens[tok]
    return token


def _login_token_consume(token: str | None) -> str | None:
    """Validate AND consume (single use). Returns the key_label, or None."""
    if not token:
        return None
    now = time.time()
    with _login_tokens_lock:
        s = _login_tokens.pop(token, None)   # pop == single-use consume
    if not s or s["expires_at"] < now:
        return None
    return s["key_label"]


def _gbl_host_requires_auth(request: "Request") -> bool:
    """Auth is required only when the request arrives via the new GenbaLink
    domain; existing /attendance/ access via the default vhost stays open."""
    host = (request.headers.get("host") or "").split(":")[0].lower()
    return host in GBL_AUTH_HOSTS


def _gbl_find_user(username: str) -> dict | None:
    for u in GBL_USERS:
        if u.get("username", "").lower() == username.lower():
            return u
    return None

# ---- Access tracking --------------------------------------------------------
# Lightweight ring buffer (last 500) + JSONL file. Separate "known clients" set
# (IP+UA hash) so brand-new logins can be flagged in the admin panel.
_access_lock = threading.Lock()
_recent_access: list[dict] = []
_RECENT_MAX = 500
KNOWN_CLIENTS: set[str] = set(_load_json(KNOWN_CLIENTS_PATH, []))
NEW_LOGIN_ALERTS: list[dict] = []
_ALERT_MAX = 100

def _client_fingerprint(ip: str, ua: str) -> str:
    return hashlib.sha256(f"{ip}|{ua}".encode("utf-8")).hexdigest()[:16]

def _parse_device(ua: str) -> str:
    """Best-effort human-readable device label from User-Agent."""
    if not ua:
        return "unknown"
    s = ua
    os_part = "?"
    for token, label in [
        ("Windows NT 10", "Windows 10/11"),
        ("Windows NT", "Windows"),
        ("Mac OS X", "macOS"),
        ("Android", "Android"),
        ("iPhone", "iPhone"),
        ("iPad", "iPad"),
        ("Linux", "Linux"),
    ]:
        if token in s:
            os_part = label
            break
    br = "?"
    for token, label in [
        ("Edg/", "Edge"),
        ("Chrome/", "Chrome"),
        ("Firefox/", "Firefox"),
        ("Safari/", "Safari"),
        ("curl/", "curl"),
        ("python-requests", "Python"),
    ]:
        if token in s:
            br = label
            break
    return f"{br} on {os_part}"

def _record_access(entry: dict) -> None:
    with _access_lock:
        _recent_access.append(entry)
        if len(_recent_access) > _RECENT_MAX:
            del _recent_access[: len(_recent_access) - _RECENT_MAX]
    try:
        with ACCESS_LOG_PATH.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception as exc:
        print(f"[ACCESS_LOG] write failed: {exc}")

# ---- Verbose API debug log --------------------------------------------------
# A second log file that captures full request / response details for
# troubleshooting auth issues, malformed bodies, etc. Two modes:
#   * Default (toggle OFF): only writes entries for status >= 400 (critical only).
#   * Toggle ON: writes every non-static request with full details.
# Toggle is persisted in admin_config.json under "debug_api_log" and can be
# overridden at process start via env var ATTENDANCE_DEBUG_LOG=0|1. Flippable
# live (no restart) via POST /admin/api/debug-log.
_REDACTED_HEADERS = {"x-api-key", "authorization", "cookie", "set-cookie"}
_SAFE_BODY_TYPE_PREFIXES = ("application/json", "application/x-www-form-urlencoded", "text/")
_MAX_REQ_BODY_BYTES = 1024
_MAX_RESP_BODY_BYTES = 2048
_MAX_REQ_BODY_CAPTURE_BYTES = 32 * 1024  # never read more than this off the wire

def _initial_debug_flag() -> bool:
    env = os.environ.get("ATTENDANCE_DEBUG_LOG")
    if env is not None:
        return env.strip() in ("1", "true", "TRUE", "yes", "on")
    return bool(ADMIN_CONFIG.get("debug_api_log", False))

DEBUG_API_LOG_ENABLED = _initial_debug_flag()
print(f"[API_DEBUG_LOG] enabled={DEBUG_API_LOG_ENABLED}  path={API_DEBUG_LOG_PATH}")

def _redact_headers(headers) -> dict:
    out = {}
    for k, v in headers.items():
        out[k] = "<redacted>" if k.lower() in _REDACTED_HEADERS else v
    return out

def _capture_body_snippet(content_type: str, body_bytes: bytes, limit: int) -> str:
    if not body_bytes:
        return ""
    ct = (content_type or "").lower()
    if not any(ct.startswith(p) for p in _SAFE_BODY_TYPE_PREFIXES):
        return f"<{len(body_bytes)} bytes of {ct or 'unknown'} skipped>"
    snippet = body_bytes[:limit]
    return snippet.decode("utf-8", errors="replace")

def _should_capture_request_body(content_type: str, content_length: str | None) -> bool:
    ct = (content_type or "").lower()
    if not ct or ct.startswith("multipart/"):
        return False
    try:
        cl = int(content_length) if content_length else 0
    except ValueError:
        cl = 0
    if cl > _MAX_REQ_BODY_CAPTURE_BYTES:
        return False
    return any(ct.startswith(p) for p in _SAFE_BODY_TYPE_PREFIXES)

def _record_debug(entry: dict) -> None:
    try:
        with API_DEBUG_LOG_PATH.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception as exc:
        print(f"[API_DEBUG_LOG] write failed: {exc}")

class AccessTrackingMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        start = time.time()
        # Real client IP: CF-Connecting-IP wins (Cloudflare Tunnel "rnd-pi" sets
        # this with the original client's IP); otherwise fall back to X-Real-IP /
        # X-Forwarded-For. The raw request.client.host is the loopback peer of
        # nginx (always ::1 for tunneled traffic) so it's the last resort.
        ip = (request.headers.get("cf-connecting-ip")
              or request.headers.get("x-real-ip")
              or request.headers.get("x-forwarded-for", "").split(",")[0].strip()
              or (request.client.host if request.client else "?"))
        ua = request.headers.get("user-agent", "")
        path = request.url.path
        # Skip noisy static asset hits from the log file (still tracked in-memory? no — too noisy).
        skip = path.startswith("/static/") or path.endswith(".ico") or path.endswith(".css") or path.endswith(".js")

        # Cache request body up-front so we can both log it AND let the handler
        # read it. Only capture for safe text-ish content types and small sizes.
        req_body_bytes = b""
        if not skip and _should_capture_request_body(
            request.headers.get("content-type", ""),
            request.headers.get("content-length"),
        ):
            try:
                req_body_bytes = await request.body()
            except Exception:
                req_body_bytes = b""

            async def _replay_receive() -> dict:
                return {"type": "http.request", "body": req_body_bytes, "more_body": False}

            request._receive = _replay_receive  # type: ignore[attr-defined]

        response = None
        status = 500
        try:
            response = await call_next(request)
            status = response.status_code
        except Exception:
            raise
        finally:
            if not skip:
                resp_body_bytes = b""
                want_resp_body = (status >= 400) or DEBUG_API_LOG_ENABLED
                if want_resp_body and response is not None:
                    try:
                        chunks = []
                        async for chunk in response.body_iterator:
                            chunks.append(chunk)
                        resp_body_bytes = b"".join(chunks)
                        response = Response(
                            content=resp_body_bytes,
                            status_code=status,
                            headers=dict(response.headers),
                            media_type=response.media_type,
                        )
                    except Exception:
                        resp_body_bytes = b""

                fp = _client_fingerprint(ip, ua)
                is_new = fp not in KNOWN_CLIENTS
                entry = {
                    "ts": datetime.now().isoformat(timespec="seconds"),
                    "ip": ip,
                    "device": _parse_device(ua),
                    "ua": ua[:240],
                    "method": request.method,
                    "path": path,
                    "status": status,
                    "duration_ms": int((time.time() - start) * 1000),
                    "fingerprint": fp,
                    "new_client": is_new,
                }
                _record_access(entry)

                # Desktop Agent activity → dedicated log + live state.
                if path.startswith("/api/v1/"):
                    try:
                        _record_agent(
                            path, request.method, status, ip,
                            _resolve_key_label(request.headers.get("x-api-key")),
                            request.url.query or "",
                            entry["duration_ms"],
                            request.headers.get("x-agent-version"),
                        )
                    except Exception as exc:
                        print(f"[AGENT_LOG] hook failed: {exc}")

                # Verbose debug log — every request when toggle ON, otherwise
                # only 4xx/5xx so admins still see the failure detail without
                # paying the disk cost of full logging in steady state.
                if DEBUG_API_LOG_ENABLED or status >= 400:
                    debug_entry = {
                        **entry,
                        "query": request.url.query or "",
                        "headers": _redact_headers(request.headers),
                        "api_key_label": _resolve_key_label(request.headers.get("x-api-key")),
                        "referer": request.headers.get("referer"),
                        "content_type": request.headers.get("content-type"),
                        "content_length": request.headers.get("content-length"),
                        "req_body": _capture_body_snippet(
                            request.headers.get("content-type", ""),
                            req_body_bytes,
                            _MAX_REQ_BODY_BYTES,
                        ),
                    }
                    if status >= 400 and resp_body_bytes:
                        resp_ct = response.headers.get("content-type", "") if response is not None else ""
                        debug_entry["resp_body"] = _capture_body_snippet(
                            resp_ct, resp_body_bytes, _MAX_RESP_BODY_BYTES
                        )
                    _record_debug(debug_entry)

                if is_new:
                    KNOWN_CLIENTS.add(fp)
                    try:
                        _save_json(KNOWN_CLIENTS_PATH, sorted(KNOWN_CLIENTS))
                    except Exception:
                        pass
                    alert = {**entry, "alert": "new client/device first seen"}
                    NEW_LOGIN_ALERTS.append(alert)
                    if len(NEW_LOGIN_ALERTS) > _ALERT_MAX:
                        del NEW_LOGIN_ALERTS[: len(NEW_LOGIN_ALERTS) - _ALERT_MAX]
                    print(f"[ACCESS_NEW] {ip} ({_parse_device(ua)}) → {request.method} {path}")
        return response

app.add_middleware(AccessTrackingMiddleware)


# ----------------------------------------------------------------------------
# GenbaLink host-gated auth middleware. Runs only for requests arriving via
# `genbafms.com` (or `www.genbafms.com`) — existing /attendance/* via the
# default vhost is untouched. Public allow-list (mobile + LINE + assets +
# auth endpoints) is matched BEFORE the gate, so LINE webhooks and the
# mobile report views keep working without a session.
# ----------------------------------------------------------------------------
GBL_PUBLIC_PREFIXES = (
    "/login", "/logout",
    "/auto-login",         # token→cookie bridge for agent deep-links (self-guarded)
    "/api/auth/",          # login/logout/whoami POST endpoints
    "/api/announcement",
    "/api/line/",
    "/api/health",
    "/api/data-status/",                # dashboard + agent verify (browser-facing)
    "/api/daily-packs/auto-extract",    # Auto-update PDF + Excel triggers (browser-facing)
                                        # — startswith match also covers /auto-extract-excel
    "/api/daily-packs/save-excel-batch",# Desktop-agent chains here after auto-extract-excel.
                                        # Handler itself enforces session-or-X-API-Key on wall hosts.
    "/api/daily-packs/items/",          # GET-only — used by /m/report to load pack data
    "/api/gantt/",          # GET-only family: /latest-date, /dates-with-data, /{date} — used by /m/report
    "/api/members/",        # GET-only family: /list, /compare — used by /m/report comparison panel
    "/api/productivity",    # GET — used by /m/summary
    "/api/m/",              # GET — used by /m/summary (rolling 30d summary feed)
    "/m/",                  # mobile report pages (employees access via LINE)
    "/hmi",                 # SCADA/HMI wall-display overview (data already public)
    "/ai-guide",            # public AI integration guide (+ /ai-guide.md)
    "/desktop-guide",       # public desktop-app user guide (+ /desktop-guide.md)
    "/guides",              # public guides hub (AI + Desktop, switchable)
    "/onboard",             # per-user onboarding card (creds via #hash, not logged)
    "/chatgpt-setup",       # personalized ChatGPT setup card (key via ?key= in URL)
    "/download/desktop",    # public desktop-app installer download
    "/dashboard",           # now serves the HMI wall display (public, like /hmi)
    "/static/",
    "/favicon",
    "/admin",               # legacy /admin has its own password
    "/partner/v1",          # external partner API — self-protected (per-partner tokens)
    "/api/v1",              # desktop-agent API — self-protected (X-API-Key, deny-by-default)
    "/api/agent/v1",        # unified AI-agent API — self-protected (agent bearer token)
    "/mcp",                 # MCP front door — self-protected (same agent token)
    "/.well-known/oauth",   # MCP OAuth discovery metadata (public)
    "/.well-known/openid",  # OIDC config alias (public)
    "/oauth/",              # MCP OAuth endpoints — self-protected (admin password / PKCE)
)


@app.middleware("http")
async def genbalink_host_auth(request: Request, call_next):
    if not _gbl_host_requires_auth(request):
        return await call_next(request)
    path = request.url.path
    if any(path == p.rstrip("/") or path.startswith(p) for p in GBL_PUBLIC_PREFIXES):
        return await call_next(request)
    if _gbl_request_user(request) is not None:
        return await call_next(request)
    # HTML navigation → redirect to /login; AJAX/API → 401 JSON
    accepts_html = "text/html" in (request.headers.get("accept") or "")
    if accepts_html:
        from fastapi.responses import RedirectResponse
        nxt = path
        if request.url.query:
            nxt = f"{path}?{request.url.query}"
        from urllib.parse import quote
        return RedirectResponse(url=f"/login?next={quote(nxt)}", status_code=302)
    return JSONResponse({"detail": "authentication required"}, status_code=401)
# ============================================================================


# ----------------------------------------------------------------------------
# Strip the `/attendance` URL prefix when requests arrive directly from the
# Cloudflare tunnel (which bypasses the nginx vhost that would normally strip
# it). Registered AFTER the auth middleware so it runs FIRST in the chain,
# meaning the auth middleware sees the already-cleaned path.
# ----------------------------------------------------------------------------
@app.middleware("http")
async def strip_attendance_prefix(request: Request, call_next):
    path = request.scope.get("path", "")
    if path == "/attendance" or path.startswith("/attendance/"):
        new_path = path[len("/attendance"):] or "/"
        request.scope["path"] = new_path
        raw = request.scope.get("raw_path")
        if isinstance(raw, (bytes, bytearray)):
            try:
                raw_str = raw.decode("latin-1")
                if raw_str.startswith("/attendance"):
                    request.scope["raw_path"] = (raw_str[len("/attendance"):] or "/").encode("latin-1")
            except Exception:
                pass
    return await call_next(request)


def write_json_atomic(path: Path, payload: object) -> None:
    """Write JSON through a temp file so roster saves never leave half-written data."""
    temp_path = path.with_suffix(f"{path.suffix}.tmp")
    temp_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    os.replace(temp_path, path)

def refresh_management_data(roster: list[dict], sections_payload: list[dict]) -> None:
    """Refresh global roster/section caches after management saves."""
    global EMPLOYEE_ROSTER, EMPLOYEE_ROSTER_BY_ID, EMPLOYEE_CODES
    global SECTIONS, SECTION_OF_CODE, SECTION_LABEL_BY_ID
    global MUTED_CODES, ACTIVE_CODES, MUTED_PLACEMENTS, SECTION_POSITIONS

    EMPLOYEE_ROSTER = roster
    EMPLOYEE_ROSTER_BY_ID = {
        employee["employee_code"]: employee["name"]
        for employee in EMPLOYEE_ROSTER
    }
    EMPLOYEE_CODES = set(EMPLOYEE_ROSTER_BY_ID)
    SECTIONS = sections_payload
    SECTION_LABEL_BY_ID = {section["id"]: section["label"] for section in SECTIONS}
    SECTION_OF_CODE, MUTED_CODES, ACTIVE_CODES, MUTED_PLACEMENTS, SECTION_POSITIONS = (
        _build_section_indexes(SECTIONS)
    )

def management_payload_from_files() -> dict:
    """Build the management UI payload — employees emitted in sections.json order."""
    employees = [
        {
            "code": code,
            "name": name_for_code(code),
            "section_id": section_id,
            "section_label": section_label,
        }
        for code, section_id, section_label in iter_codes_in_section_order()
    ]
    sections = [
        {
            "id": section["id"],
            "label": section["label"],
            "codes": list(section["codes"]),
        }
        for section in SECTIONS
    ]
    return {
        "sections": sections,
        "employees": employees,
    }

# Batch job infrastructure (multi-file upload with progress tracking).
BATCH_UPLOAD_DIR = Path("/tmp/attendance_batch_uploads")
BATCH_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
BATCH_MAX_TOTAL_BYTES = 500 * 1024 * 1024  # 500 MB hard cap per batch.
BATCH_JOB_TTL_SECONDS = 60 * 60  # 1 hour retention for completed jobs.
BATCH_CPU_THROTTLE_THRESHOLD = float(os.getenv("ATTENDANCE_BATCH_CPU_MAX", "88"))
BATCH_RAM_THROTTLE_THRESHOLD = float(os.getenv("ATTENDANCE_BATCH_RAM_MAX", "94"))
BATCH_THROTTLE_SLEEP_SECONDS = 0.75
BATCH_MAX_THROTTLE_WAITS = 12  # ~9 s of back-off per file before giving up.
BATCH_JOBS: dict[str, dict] = {}
BATCH_JOBS_LOCK = threading.Lock()

def _sample_system_load() -> dict:
    """Return current CPU/RAM usage so the batch worker can self-throttle."""
    if not HAVE_PSUTIL:
        return {"cpu_percent": None, "ram_percent": None, "ram_available_mb": None}
    try:
        cpu = psutil.cpu_percent(interval=None)
        mem = psutil.virtual_memory()
        return {
            "cpu_percent": round(cpu, 1),
            "ram_percent": round(mem.percent, 1),
            "ram_available_mb": round(mem.available / (1024 * 1024), 1),
        }
    except Exception:  # pragma: no cover
        return {"cpu_percent": None, "ram_percent": None, "ram_available_mb": None}

def _batch_set(job_id: str, **updates) -> None:
    """Thread-safe shallow update of a job record."""
    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if job is None:
            return
        job.update(updates)
        load = _sample_system_load()
        job["cpu_percent"] = load["cpu_percent"]
        job["ram_percent"] = load["ram_percent"]
        job["ram_available_mb"] = load["ram_available_mb"]

def _batch_update_file(job_id: str, index: int, **updates) -> None:
    """Thread-safe update of a single file entry inside a job."""
    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if job is None:
            return
        files = job.get("files") or []
        if 0 <= index < len(files):
            files[index].update(updates)

def _batch_snapshot(job_id: str) -> dict | None:
    """Thread-safe read of a job record (shallow copy; safe for JSON serialization)."""
    with BATCH_JOBS_LOCK:
        job = BATCH_JOBS.get(job_id)
        if job is None:
            return None
        snapshot = {key: value for key, value in job.items() if key != "files"}
        snapshot["files"] = [dict(entry) for entry in job.get("files") or []]
        return snapshot

def _batch_cleanup_expired() -> None:
    """Drop finished jobs older than BATCH_JOB_TTL_SECONDS and remove their temp dirs."""
    cutoff = time.time() - BATCH_JOB_TTL_SECONDS
    expired: list[str] = []
    with BATCH_JOBS_LOCK:
        for job_id, job in list(BATCH_JOBS.items()):
            finished_at = job.get("finished_at_ts")
            if finished_at and finished_at < cutoff:
                expired.append(job_id)
                BATCH_JOBS.pop(job_id, None)
    for job_id in expired:
        job_dir = BATCH_UPLOAD_DIR / job_id
        if job_dir.exists():
            shutil.rmtree(job_dir, ignore_errors=True)

def _wait_for_resources(job_id: str) -> None:
    """Adaptive throttle: pause briefly when CPU/RAM are saturated."""
    if not HAVE_PSUTIL:
        return
    for _ in range(BATCH_MAX_THROTTLE_WAITS):
        load = _sample_system_load()
        cpu = load.get("cpu_percent")
        ram = load.get("ram_percent")
        if cpu is None or ram is None:
            return
        if cpu < BATCH_CPU_THROTTLE_THRESHOLD and ram < BATCH_RAM_THROTTLE_THRESHOLD:
            return
        with BATCH_JOBS_LOCK:
            job = BATCH_JOBS.get(job_id)
            if job is not None:
                job["throttled"] = True
                job["cpu_percent"] = cpu
                job["ram_percent"] = ram
        time.sleep(BATCH_THROTTLE_SLEEP_SECONDS)

EMPTY_MARKERS = {"", "__:_", "__:__", "----", "------", "早退"}
TIME_PATTERN = re.compile(r"^(?:(当|翌)\s*)?(\d{1,2}):(\d{2})$")
MONTH_PATTERN = re.compile(r"(20\d{2})[.\-/](\d{1,2})")

def get_db_connection():
    """Open a PostgreSQL connection for attendance persistence."""
    connection = psycopg2.connect(
        host=DATABASE_HOST,
        port=DATABASE_PORT,
        dbname=DATABASE_NAME,
        user=DATABASE_USER,
        password=DATABASE_PASSWORD,
    )
    return connection

def init_db() -> None:
    """Ensure the PostgreSQL tables needed for uploads and attendance rows exist."""
    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS upload_batches (
                    id SERIAL PRIMARY KEY,
                    file_name VARCHAR(255) NOT NULL,
                    total_records INTEGER NOT NULL DEFAULT 0,
                    upload_date TIMESTAMP NOT NULL DEFAULT NOW()
                )
            """)
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS attendance_records (
                    id SERIAL PRIMARY KEY,
                    upload_date TIMESTAMP NOT NULL DEFAULT NOW(),
                    file_name VARCHAR(255) NOT NULL,
                    personal_code VARCHAR(20) NOT NULL,
                    full_name VARCHAR(100) NOT NULL,
                    commute_time VARCHAR(10) DEFAULT '',
                    time_to_leave VARCHAR(10) DEFAULT '',
                    working_hours VARCHAR(10) DEFAULT '',
                    record_date DATE NOT NULL,
                    month_year VARCHAR(7) NOT NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT NOW()
                )
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_code ON attendance_records(personal_code)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance_records(record_date)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_month ON attendance_records(month_year)")
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS daily_packs (
                    record_date DATE PRIMARY KEY,
                    number_of_packs INTEGER NOT NULL DEFAULT 0,
                    note TEXT,
                    created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                    updated_at TIMESTAMP NOT NULL DEFAULT NOW()
                )
            """)
            # temp_staff stores フルキャスト manual entries per date. Each row is one
            # bucket of N workers starting together and leaving at the same time.
            # total_hours = headcount * hours_per_person and is used by the Summary Report
            # to fold temp-staff labor into the company-wide hour total.
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS temp_staff (
                    id SERIAL PRIMARY KEY,
                    record_date DATE NOT NULL,
                    company VARCHAR(50) NOT NULL DEFAULT 'フルキャスト',
                    headcount INTEGER NOT NULL,
                    start_time VARCHAR(5) NOT NULL,
                    leave_time VARCHAR(5) NOT NULL,
                    leave_next_day BOOLEAN NOT NULL DEFAULT FALSE,
                    hours_per_person NUMERIC(6,2) NOT NULL,
                    total_hours NUMERIC(8,2) NOT NULL,
                    note TEXT,
                    created_at TIMESTAMP NOT NULL DEFAULT NOW()
                )
            """)
            cursor.execute("ALTER TABLE temp_staff ADD COLUMN IF NOT EXISTS leave_next_day BOOLEAN NOT NULL DEFAULT FALSE")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_temp_staff_date ON temp_staff(record_date)")
            # 2026-05-01 — daily_packs gets right-side N合計/Y合計 + section-2 plan start_time.
            cursor.execute("ALTER TABLE daily_packs ADD COLUMN IF NOT EXISTS n_total INTEGER")
            cursor.execute("ALTER TABLE daily_packs ADD COLUMN IF NOT EXISTS y_total INTEGER")
            cursor.execute("ALTER TABLE daily_packs ADD COLUMN IF NOT EXISTS section_start_time VARCHAR(5)")
            # 2026-05-01 — production_plan persists the A/B/C line breakdown
            # parsed off 製造予定表 ＮＹ. One row per (date, line, item).
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS production_plan (
                    id              SERIAL PRIMARY KEY,
                    record_date     DATE    NOT NULL,
                    line_code       CHAR(1) NOT NULL,
                    item_name       TEXT    NOT NULL,
                    planned_qty     INTEGER NOT NULL DEFAULT 0,
                    n_qty           INTEGER NOT NULL DEFAULT 0,
                    y_qty           INTEGER NOT NULL DEFAULT 0,
                    combined_qty    INTEGER NOT NULL DEFAULT 0,
                    start_time      VARCHAR(5),
                    takt_time       VARCHAR(8),
                    pph_target      INTEGER,
                    required_staff  INTEGER,
                    source_filename TEXT,
                    saved_at        TIMESTAMP NOT NULL DEFAULT NOW(),
                    UNIQUE (record_date, line_code, item_name)
                )
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_production_plan_date ON production_plan(record_date)")
            # daily_pack_items stores the per-product Excel breakdown (one row per
            # product per upload). Replaces a same-date batch on re-upload via
            # delete-then-insert, so production_date is unique-per-upload not strict PK.
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS daily_pack_items (
                    id              SERIAL PRIMARY KEY,
                    batch_id        UUID    NOT NULL,
                    production_date DATE    NOT NULL,
                    product_name    TEXT    NOT NULL,
                    product_key     TEXT    NOT NULL,
                    n_yamanashi  INTEGER, n_nagano INTEGER, n_matsumoto INTEGER,
                    y_yamanashi  INTEGER, y_nagano INTEGER, y_matsumoto INTEGER,
                    n_total       INTEGER,
                    y_total       INTEGER,
                    grand_total   INTEGER,
                    packs_per_case INTEGER,
                    rate_per_hour    INTEGER,
                    est_seconds      INTEGER,
                    source_filename  TEXT,
                    weather          VARCHAR(20),
                    temperature      NUMERIC(5,1),
                    input_by         VARCHAR(60),
                    start_time       VARCHAR(5),
                    end_time         VARCHAR(5),
                    uploaded_at      TIMESTAMP NOT NULL DEFAULT NOW()
                )
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_daily_pack_items_date ON daily_pack_items(production_date)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_daily_pack_items_batch ON daily_pack_items(batch_id)")
            # 2026-05-05 — uploaded_file_registry: dedupe by SHA-256, track lifecycle
            # ('received' on upload, 'loaded' once rows reach the DB, 'replaced' if a
            # later file with the same target_date overwrites it). Lets the desktop
            # agent ask "have you seen this exact file?" before re-sending bytes.
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS uploaded_file_registry (
                    id              SERIAL PRIMARY KEY,
                    sha256          CHAR(64) NOT NULL UNIQUE,
                    file_type       VARCHAR(20) NOT NULL,
                    file_name       TEXT NOT NULL,
                    file_size       BIGINT NOT NULL,
                    target_date     DATE,
                    record_count    INTEGER,
                    received_at     TIMESTAMP NOT NULL DEFAULT NOW(),
                    loaded_at       TIMESTAMP,
                    moved_path      TEXT,
                    source_key      VARCHAR(40),
                    status          VARCHAR(20) NOT NULL DEFAULT 'received'
                )
            """)
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_ufr_type_date ON uploaded_file_registry(file_type, target_date)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_ufr_status ON uploaded_file_registry(status)")
        connection.commit()

init_db()
try:
    _change_requests.ensure_tables()   # Part B: agent change-request ledger
except Exception as _cr_exc:
    print(f"[CHANGE_REQUESTS] ensure_tables failed: {_cr_exc}")

def clean_cell(value: object) -> str:
    """Normalize PDF table cell values for matching and display."""
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()

def is_employee_code(value: str) -> bool:
    """Attendance rows use 8-digit employee codes; subtotal rows do not."""
    return value.isdigit() and len(value) == 8

def find_column_index(header_row: list[str], keyword: str) -> int | None:
    """Find the first header cell containing the requested keyword."""
    for index, cell in enumerate(header_row):
        if keyword in cell:
            return index
    return None

def get_row_value(row: list[str], index: int | None) -> str:
    """Safely read a column value from an extracted table row."""
    if index is None or index >= len(row):
        return ""
    return row[index]

def normalize_blank_value(value: str) -> str:
    """Convert placeholder markers to blank strings."""
    return "" if value in EMPTY_MARKERS else value

def detect_section_from_text(text: str | None) -> int | None:
    """Infer a department from PDF text when the file prints the section label."""
    if not text:
        return None
    normalized = text.replace(" ", "")
    for section_id, patterns in SECTION_TEXT_PATTERNS.items():
        for pattern in patterns:
            if pattern.search(normalized) or pattern.search(text):
                return section_id
    return None

def normalize_time_value(value: str, *, is_leave: bool = False) -> str:
    """
    Remove attendance prefixes and normalize late-night leave times.

    Leave times at or after midnight are represented in 24+ hour format,
    e.g. 1:00 AM becomes 25:00.
    """
    raw = clean_cell(value)
    if raw in EMPTY_MARKERS:
        return ""

    match = TIME_PATTERN.match(raw)
    if not match:
        return raw

    prefix, hour_str, minute_str = match.groups()
    hour = int(hour_str)
    minute = int(minute_str)

    if is_leave and (prefix == "翌" or hour <= 6):
        hour += 24

    return f"{hour}:{minute:02d}"

def time_to_minutes(value: str) -> int | None:
    """Convert a normalized H:MM string into absolute minutes."""
    normalized = normalize_blank_value(clean_cell(value))
    if not normalized:
        return None

    match = re.match(r"^(\d{1,2}):(\d{2})$", normalized)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2))
    return hour * 60 + minute

def calculate_working_hours(start_time: str, leave_time: str) -> str:
    """Calculate one working-hours value like 8:00 hr from start/end times."""
    start_minutes = time_to_minutes(start_time)
    leave_minutes = time_to_minutes(leave_time)
    if start_minutes is None or leave_minutes is None or leave_minutes < start_minutes:
        return ""

    diff_minutes = leave_minutes - start_minutes
    hours = diff_minutes // 60
    minutes = diff_minutes % 60
    return f"{hours}:{minutes:02d} hr"

def calculate_temp_staff_hours(start_time: str, leave_time: str,
                                leave_next_day: bool = False) -> float:
    """
    Hours for one フルキャスト row. The shift slot is 19:00 → next-day 10:00.
    When leave_next_day is True, 24h is added to the leave time so the
    computation spans midnight correctly. When False, both times are on
    the same calendar day. Returns 0.0 if inputs are unparseable or the
    resulting duration is non-positive.
    """
    raw_start = (start_time or "").strip()
    raw_leave = (leave_time or "").strip()
    match_start = re.match(r"^(\d{1,2}):(\d{2})$", raw_start)
    match_leave = re.match(r"^(\d{1,2}):(\d{2})$", raw_leave)
    if not match_start or not match_leave:
        return 0.0
    start_total = int(match_start.group(1)) * 60 + int(match_start.group(2))
    leave_total = int(match_leave.group(1)) * 60 + int(match_leave.group(2))
    if leave_next_day:
        leave_total += 24 * 60
    if leave_total <= start_total:
        return 0.0
    return round((leave_total - start_total) / 60, 2)

def parse_working_hours_minutes(value: str) -> int | None:
    """Convert values like 8:01 hr into minutes for dashboard calculations."""
    normalized = normalize_blank_value(clean_cell(value)).replace(" hr", "")
    if not normalized:
        return None

    match = re.match(r"^(\d+):(\d{2})$", normalized)
    if not match:
        return None

    return int(match.group(1)) * 60 + int(match.group(2))

def format_average_hours(minutes_values: list[int]) -> float:
    """Return an hours float rounded to 2 decimals for API summaries."""
    if not minutes_values:
        return 0.0
    return round((sum(minutes_values) / len(minutes_values)) / 60, 2)

def record_has_data(record: dict[str, str]) -> bool:
    """Check whether a roster row contains at least one non-blank value."""
    return any(record.get(field) for field in ("commute_time", "leave_time", "working_hours"))

def count_records_with_data(records: list[dict[str, str]]) -> int:
    """Count output rows that contain actual attendance data."""
    return sum(1 for record in records if record_has_data(record))

def extract_month_year(filename: str) -> str:
    """Derive a YYYY-MM value from the uploaded filename when available."""
    match = MONTH_PATTERN.search(filename)
    if not match:
        return datetime.now().strftime("%Y-%m")

    year, month = match.groups()
    return f"{year}-{int(month):02d}"

def resolve_month(month: str | None = None) -> str:
    """Choose the requested month or fall back to the latest saved batch."""
    if month:
        return month

    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT month_year
                FROM attendance_records
                ORDER BY upload_date DESC
                LIMIT 1
                """
            )
            row = cursor.fetchone()
    return row["month_year"] if row else datetime.now().strftime("%Y-%m")

def parse_requested_codes(raw_values: list[str] | None) -> list[str]:
    """Split repeated or comma/newline separated code inputs into clean values."""
    if not raw_values:
        return []

    codes: list[str] = []
    for raw in raw_values:
        if raw is None:
            continue
        for code in re.split(r"[\s,]+", raw.strip()):
            if code:
                codes.append(code)
    return codes

def create_export_filename(original_name: str) -> str:
    """Build a safe Excel filename for an uploaded PDF."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(original_name).stem).strip("._")
    if not safe_stem:
        safe_stem = "attendance"
    return f"{safe_stem}_{timestamp}_{uuid4().hex[:8]}.xlsx"

def create_bundle_filename() -> str:
    """Build a zip filename for multi-file conversion."""
    return f"attendance_bundle_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex[:8]}.zip"

def extract_pdf_metadata(file_path: Path) -> dict[str, object]:
    """Pull production date and pack count from the PDF text content.

    Dates are looked up in this order (first hit wins):
      1. 処理日：YYYY/MM/DD   — the attendance working-day / processing-day line
      2. YYYY年MM月DD日         — generic Japanese long-form date
    """
    metadata: dict[str, object] = {"record_date": None, "number_of_packs": None}
    try:
        with pdfplumber.open(file_path) as pdf:
            full_text = "\n".join((page.extract_text() or "") for page in pdf.pages)
    except Exception:
        return metadata

    # 1) Prefer 処理日 (fullwidth or ASCII colon, slash separators)
    proc_match = re.search(
        r"処理日\s*[:：]\s*(\d{4})\s*[/／年\-\.]\s*(\d{1,2})\s*[/／月\-\.]\s*(\d{1,2})",
        full_text,
    )
    if proc_match:
        try:
            year, month, day = (int(part) for part in proc_match.groups())
            metadata["record_date"] = datetime(year, month, day).date()
        except ValueError:
            pass

    # 2) Fallback to generic Japanese long-form date
    if metadata["record_date"] is None:
        date_match = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", full_text)
        if date_match:
            try:
                year, month, day = (int(part) for part in date_match.groups())
                metadata["record_date"] = datetime(year, month, day).date()
            except ValueError:
                pass

    # Anchor to the 製造パック数 label so productivity numbers like 29.1P/h are ignored.
    pack_match = re.search(r"製造パック数[^\d]*([\d,]+)\s*[PＰ]", full_text)
    if pack_match:
        try:
            metadata["number_of_packs"] = int(pack_match.group(1).replace(",", ""))
        except ValueError:
            pass

    metadata["fullcast_rows"] = extract_fullcast_rows(full_text)

    return metadata


def extract_fullcast_rows(full_text: str) -> list[dict[str, object]]:
    """Find 'フルキャスト N 名 HH:MM HH:MM HH:MM' rows in PDF text.

    The third time column (working_hours) is ignored because the source PDF
    sometimes prints a visibly wrong sum (e.g. 47:15 for a 6h45m shift).
    Hours are always recomputed from start+leave so the stored value is
    authoritative.
    """
    rows: list[dict[str, object]] = []
    # Full-width spaces come through as 　; some exports use ASCII spaces.
    # Time column is optional break between start/leave/hours.
    pattern = re.compile(
        r"フルキャスト[\s　]*(\d+)[\s　]*名"
        r"[\s　]+(\d{1,2}:\d{2})"
        r"[\s　]+(\d{1,2}:\d{2})"
        r"[\s　]+(\d{1,2}:\d{2})"
    )
    for match in pattern.finditer(full_text):
        try:
            headcount = int(match.group(1))
        except ValueError:
            continue
        if headcount <= 0 or headcount > 999:
            continue
        start_raw = match.group(2)
        leave_raw = match.group(3)
        start_norm, start_next = normalize_plus24_time(start_raw)
        leave_norm, leave_next = normalize_plus24_time(leave_raw)
        if start_norm is None or leave_norm is None:
            continue
        # Explicit +24 on leave wins. Otherwise infer overnight by leave <= start.
        leave_next_day = leave_next or (not start_next and leave_norm <= start_norm)
        hours = calculate_temp_staff_hours(start_norm, leave_norm, leave_next_day)
        if hours <= 0:
            continue
        rows.append({
            "company": "フルキャスト",
            "headcount": headcount,
            "start_time": start_norm,
            "leave_time": leave_norm,
            "leave_next_day": leave_next_day,
            "hours_per_person": round(hours, 2),
            "total_hours": round(hours * headcount, 2),
        })
    return rows


def normalize_plus24_time(raw: str) -> tuple[str | None, bool]:
    """Convert '26:40' to ('02:40', True); leave '19:00' as ('19:00', False)."""
    match = re.match(r"^(\d{1,2}):(\d{2})$", raw or "")
    if not match:
        return None, False
    hour = int(match.group(1))
    minute = int(match.group(2))
    if minute >= 60:
        return None, False
    next_day = False
    if hour >= 24:
        hour -= 24
        next_day = True
    if hour >= 24 or hour < 0:
        return None, False
    return f"{hour:02d}:{minute:02d}", next_day

def find_attendance_mismatches(
    record_date,
    records: list[dict[str, str]],
) -> list[dict[str, object]]:
    """Compare freshly parsed PDF records with whatever the DB already holds for that date."""
    if record_date is None:
        return []

    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT personal_code, full_name, commute_time, time_to_leave, working_hours
                FROM attendance_records
                WHERE record_date = %s
                """,
                (record_date,),
            )
            existing_by_code = {row["personal_code"]: row for row in cursor.fetchall()}

    mismatches: list[dict[str, object]] = []
    for record in records:
        if not record_has_data(record):
            continue
        code = record["employee_code"]
        prior = existing_by_code.get(code)
        if prior is None:
            continue
        diffs: dict[str, dict[str, str]] = {}
        if (prior["commute_time"] or "") != (record.get("commute_time") or ""):
            diffs["commute_time"] = {
                "db": prior["commute_time"] or "",
                "pdf": record.get("commute_time", ""),
            }
        if (prior["time_to_leave"] or "") != (record.get("leave_time") or ""):
            diffs["leave_time"] = {
                "db": prior["time_to_leave"] or "",
                "pdf": record.get("leave_time", ""),
            }
        if (prior["working_hours"] or "") != (record.get("working_hours") or ""):
            diffs["working_hours"] = {
                "db": prior["working_hours"] or "",
                "pdf": record.get("working_hours", ""),
            }
        if diffs:
            mismatches.append({
                "employee_code": code,
                "name": record.get("name", "") or prior["full_name"],
                "differences": diffs,
            })
    return mismatches

def upsert_pack_count_for_date(record_date, pack_count: int, source: str | None = None) -> None:
    """Save the pack count extracted from a PDF for a given production date."""
    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO daily_packs (record_date, number_of_packs, note)
                VALUES (%s, %s, %s)
                ON CONFLICT (record_date) DO UPDATE
                    SET number_of_packs = EXCLUDED.number_of_packs,
                        note = COALESCE(EXCLUDED.note, daily_packs.note),
                        updated_at = NOW()
                """,
                (record_date, pack_count, f"auto from {source}" if source else None),
            )
        connection.commit()

def save_batch_to_db(
    source_filename: str,
    export_filename: str,
    records: list[dict[str, str]],
    converted_at: str,
    record_date=None,
) -> tuple[int, int]:
    """Persist one converted upload batch and its attendance rows."""
    month_year = extract_month_year(source_filename)
    records_processed = count_records_with_data(records)
    if record_date is None:
        record_date = datetime.strptime(f"{month_year}-01", "%Y-%m-%d").date()

    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                DELETE FROM attendance_records
                WHERE file_name = %s AND month_year = %s
                """,
                (source_filename, month_year),
            )
            cursor.execute(
                """
                DELETE FROM upload_batches
                WHERE file_name = %s
                """,
                (export_filename,),
            )
            cursor.execute(
                """
                INSERT INTO upload_batches (file_name, total_records, upload_date)
                VALUES (%s, %s, %s)
                RETURNING id
                """,
                (export_filename, records_processed, converted_at),
            )
            batch_id = cursor.fetchone()[0]
            cursor.executemany(
                """
                INSERT INTO attendance_records (
                    upload_date,
                    file_name,
                    personal_code,
                    full_name,
                    commute_time,
                    time_to_leave,
                    working_hours,
                    record_date,
                    month_year,
                    created_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                [
                    (
                        converted_at,
                        source_filename,
                        record["employee_code"],
                        record["name"],
                        record.get("commute_time", ""),
                        record.get("leave_time", ""),
                        record.get("working_hours", ""),
                        record_date,
                        month_year,
                        converted_at,
                    )
                    for record in records
                ],
            )
        connection.commit()

    return batch_id, records_processed

def list_attendance_rows(month: str):
    """Fetch saved attendance rows for a given month."""
    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT personal_code AS employee_code,
                       full_name,
                       commute_time,
                       time_to_leave AS leave_time,
                       working_hours,
                       CASE
                           WHEN commute_time <> '' OR time_to_leave <> '' OR working_hours <> '' THEN 1
                           ELSE 0
                       END AS has_data,
                       upload_date AS converted_at,
                       file_name AS source_filename,
                       record_date
                FROM attendance_records
                WHERE month_year = %s
                ORDER BY personal_code, upload_date
                """,
                (month,),
            )
            return cursor.fetchall()

def build_preview_payload(
    filename: str,
    file_size: int,
    records: list[dict[str, str]],
    *,
    file_count: int = 1,
    preview_filename: str | None = None,
) -> dict[str, object]:
    """Create a consistent preview payload for one or many files."""
    return {
        "filename": filename,
        "preview_filename": preview_filename or filename,
        "file_size": file_size,
        "file_count": file_count,
        "status": "success",
        # Return the full aligned roster so the frontend can paginate/filter locally.
        "records": records,
        "total_records": len(records),
        "records_with_data": count_records_with_data(records),
        "extracted_at": datetime.now().isoformat(),
    }

async def parse_uploaded_pdf(
    file: UploadFile,
) -> tuple[str, int, list[dict[str, str]], dict[str, object]]:
    """Save, parse, and roster-align an uploaded PDF, returning metadata too."""
    original_name, file_path, content = await save_uploaded_pdf(file)
    records = apply_employee_roster(parse_pdf_data(file_path))
    metadata = extract_pdf_metadata(file_path)
    return original_name, len(content), records, metadata

def export_records_to_excel(
    source_filename: str,
    records: list[dict[str, str]],
    record_date=None,
) -> tuple[str, Path, int, int]:
    """Create one Excel file from roster-aligned records and save it to the DB."""
    excel_filename = create_export_filename(source_filename)
    excel_path = EXPORT_DIR / excel_filename
    create_excel_file(records, excel_path)

    converted_at = datetime.now().isoformat()
    batch_id, records_processed = save_batch_to_db(
        source_filename,
        excel_filename,
        records,
        converted_at,
        record_date=record_date,
    )

    uploaded_files.append({
        "source_filename": source_filename,
        "export_filename": excel_filename,
        "records_processed": records_processed,
        "converted_at": converted_at,
        "batch_id": batch_id,
    })

    return excel_filename, excel_path, records_processed, batch_id

def apply_employee_roster(records: list[dict[str, str]]) -> list[dict[str, str]]:
    """Return rows in sections.json order (then any roster-only codes), blanks for missing."""
    records_by_id = {
        record["employee_code"]: record
        for record in records
        if record["employee_code"] in EMPLOYEE_ROSTER_BY_ID
    }

    rostered_records = []
    for code, _sid, _label in iter_codes_in_section_order():
        parsed_record = records_by_id.get(code, {})
        muted = code in MUTED_CODES
        rostered_records.append({
            "employee_code": code,
            "name": name_for_code(code),
            "commute_time": "" if muted else parsed_record.get("commute_time", ""),
            "leave_time": "" if muted else parsed_record.get("leave_time", ""),
            "working_hours": "" if muted else parsed_record.get("working_hours", ""),
        })

    return rostered_records

def build_record_from_row(
    row: list[str],
    columns: dict[str, int | None],
    *,
    section_id: int | None = None,
) -> dict[str, str] | None:
    """Convert one extracted PDF row into the API's record shape."""
    code = get_row_value(row, columns["employee_code"])
    if not is_employee_code(code):
        return None

    commute_time = normalize_time_value(get_row_value(row, columns["commute_time"]))
    leave_time = normalize_time_value(get_row_value(row, columns["leave_time"]), is_leave=True)

    return {
        "employee_code": code,
        "name": get_row_value(row, columns["name"]),
        "commute_time": commute_time,
        "leave_time": leave_time,
        "working_hours": calculate_working_hours(commute_time, leave_time),
        "section_id": section_id,
        "section_label": SECTION_LABEL_BY_ID.get(section_id, "未配属"),
    }

def parse_pdf_data(file_path: Path) -> list[dict[str, str]]:
    """Parse attendance PDF and extract data"""
    try:
        records = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                page_section_id = detect_section_from_text(page_text)
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        normalized_rows = [[clean_cell(cell) for cell in row] for row in table if row]
                        header_row = next((row for row in normalized_rows if any("個人" in cell for cell in row)), None)
                        if not header_row:
                            continue

                        columns = {
                            "employee_code": find_column_index(header_row, "個人"),
                            "name": find_column_index(header_row, "氏名"),
                            "commute_time": find_column_index(header_row, "出勤時刻"),
                            "leave_time": find_column_index(header_row, "退勤時刻"),
                        }

                        for row in normalized_rows:
                            record = build_record_from_row(row, columns, section_id=page_section_id)
                            if record:
                                records.append(record)

        if not records:
            raise ValueError("No attendance records were found in the PDF.")

        return records
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Unable to extract attendance data from the PDF: {exc}") from exc

def validate_pdf_content(content: bytes) -> None:
    """Reject empty uploads and obvious non-PDF files before parsing."""
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    if b"%PDF" not in content[:1024]:
        raise HTTPException(status_code=400, detail="Uploaded file is not a valid PDF.")

async def save_uploaded_pdf(file: UploadFile) -> tuple[str, Path, bytes]:
    """Persist the uploaded PDF using a safe, unique filename."""
    original_name = Path(file.filename or "").name
    if not original_name:
        raise HTTPException(status_code=400, detail="Uploaded file must include a filename.")

    if Path(original_name).suffix.lower() != ".pdf":
        raise HTTPException(status_code=400, detail="Only PDF files are supported.")

    content = await file.read()
    validate_pdf_content(content)

    stored_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex}.pdf"
    file_path = UPLOAD_DIR / stored_name
    file_path.write_bytes(content)

    return original_name, file_path, content

def build_download_url(filename: str) -> str:
    """Return an app-local API path; the proxy prefix is added by the frontend."""
    return f"/api/download/{filename}"

def get_export_file_path(filename: str) -> Path:
    """Resolve a requested export filename inside the export directory only."""
    safe_name = Path(filename).name
    if safe_name != filename or Path(safe_name).suffix.lower() not in {".xlsx", ".zip"}:
        raise HTTPException(status_code=400, detail="Invalid filename.")

    file_path = (EXPORT_DIR / safe_name).resolve()
    if file_path.parent != EXPORT_DIR.resolve():
        raise HTTPException(status_code=400, detail="Invalid filename.")

    return file_path

def create_excel_file(records: list[dict[str, str]], filename: Path) -> Path:
    """Create readable Excel file with Japanese headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = "勤務表"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='MS ゴシック', size=11, bold=True, color="FFFFFF")
    data_font = Font(name='MS ゴシック', size=10)
    section_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    section_font = Font(name='MS ゴシック', size=10, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add title
    ws['A1'] = f"勤務表　{datetime.now().strftime('%Y年%m月%d日')}"
    ws['A1'].font = Font(name='MS ゴシック', size=14, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 25
    
    # Add headers
    headers = ['個人コード', '氏名', '出勤時間', '退勤時間', '労働時間']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    ws.row_dimensions[3].height = 20
    
    # Add data rows
    excel_rows: list[dict[str, str] | None] = []
    for index, record in enumerate(records):
        excel_rows.append(record)
        if (
            record.get("employee_code") == EXCEL_SECTION_INSERT_AFTER
            and index + 1 < len(records)
            and records[index + 1].get("employee_code") == EXCEL_SECTION_INSERT_BEFORE
        ):
            excel_rows.append(None)

    current_row = 4
    for record in excel_rows:
        if record is None:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            section_cell = ws.cell(row=current_row, column=1)
            section_cell.value = EXCEL_SECTION_LABEL
            section_cell.font = section_font
            section_cell.fill = section_fill
            section_cell.alignment = center_align
            section_cell.border = border
            for col in range(2, 6):
                ws.cell(row=current_row, column=col).border = border
                ws.cell(row=current_row, column=col).fill = section_fill
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            continue

        ws.cell(row=current_row, column=1).value = record.get('employee_code', '')
        ws.cell(row=current_row, column=2).value = record.get('name', '')
        ws.cell(row=current_row, column=3).value = record.get('commute_time', '')
        ws.cell(row=current_row, column=4).value = record.get('leave_time', '')
        ws.cell(row=current_row, column=5).value = record.get('working_hours', '')

        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = border

        ws.row_dimensions[current_row].height = 18
        current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    
    # Add summary row
    summary_row = current_row + 1
    ws.cell(row=summary_row, column=1).value = "合計"
    ws.cell(row=summary_row, column=1).font = Font(name='MS ゴシック', size=10, bold=True)
    ws.cell(row=summary_row, column=2).value = f"従業員数: {len(records)}"
    ws.cell(row=summary_row, column=2).font = Font(name='MS ゴシック', size=10, bold=True)
    
    # Save file
    wb.save(filename)
    return filename

@app.get("/")
async def root():
    """Serve web interface"""
    return FileResponse(BASE_DIR / "index.html",
                        headers={"Cache-Control": "no-store"})

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


# ----------------------------------------------------------------------------
# Mount the Upload Hub as a sub-app at /upload so that
# https://rnd.asiakawaii.com/upload/ works while Cloudflare bypasses nginx.
# `upload.service` still runs on port 8003 for direct/LAN access; this is an
# in-process duplicate that shares the same on-disk UPLOAD_DIR.
# ----------------------------------------------------------------------------
try:
    import importlib.util as _ilu
    _spec = _ilu.spec_from_file_location("upload_hub_main", "/var/www/upload_app/main.py")
    _upload_hub = _ilu.module_from_spec(_spec)
    _spec.loader.exec_module(_upload_hub)
    app.mount("/upload", _upload_hub.app, name="upload_hub")
except Exception as _e:
    print(f"[upload mount] failed to mount Upload Hub at /upload: {_e}")


@app.get("/gantt")
async def gantt_page():
    """Standalone attendance Gantt chart (A4 portrait, printable).

    `no-store` header ensures the browser never serves a stale template after a
    redesign, so *every* date sees the newest layout on reload.
    """
    return FileResponse(
        STATIC_DIR / "gantt.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )

@app.get("/lens")
async def lens_page():
    """Lens — analytical views over attendance data. Hosts the Member Hours
    comparison (moved out of the Gantt page). `no-store` so reloads always pick
    up the newest template.
    """
    return FileResponse(
        STATIC_DIR / "lens.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )

@app.get("/download/desktop")
async def download_desktop_app():
    """Public download of the latest desktop-app installer (no secrets inside;
    the API key is configured by the user after install)."""
    meta = _load_json(APP_UPDATE_CONFIG_PATH, {}) or {}
    fn = meta.get("filename")
    if fn:
        p = APP_UPDATE_DIR / fn
        if p.exists():
            return FileResponse(p, filename=fn, media_type="application/octet-stream")
    raise HTTPException(status_code=404, detail="No desktop app build available.")

@app.get("/chatgpt-setup")
async def chatgpt_setup_page():
    """Personalized ChatGPT setup card. Send the link with ?key=at_… to one
    person; the page pre-fills their key + copy buttons. No key is baked in."""
    return FileResponse(STATIC_DIR / "chatgpt_setup.html",
                        headers={"Cache-Control": "no-store"})

@app.get("/onboard")
async def onboard_page():
    """Per-user onboarding card. Credentials are passed in the URL #hash (client-
    side only — never sent to the server/logs). The page pre-fills them."""
    return FileResponse(STATIC_DIR / "onboard.html",
                        headers={"Cache-Control": "no-store"})

@app.get("/guides")
async def guides_hub_page():
    """Public guides hub — switch between the AI and Desktop guides in one page."""
    return FileResponse(STATIC_DIR / "guides.html",
                        headers={"Cache-Control": "no-store"})

@app.get("/desktop-guide")
async def desktop_guide_page():
    """Public Desktop App user guide (rendered)."""
    return FileResponse(STATIC_DIR / "desktop_guide.html",
                        headers={"Cache-Control": "no-store"})

@app.get("/desktop-guide.md")
async def desktop_guide_md():
    """Raw markdown for the desktop app guide."""
    return FileResponse(BASE_DIR / "DESKTOP_APP_GUIDE.md",
                        media_type="text/markdown; charset=utf-8",
                        headers={"Cache-Control": "no-store"})

@app.get("/ai-guide")
async def ai_guide_page():
    """Public AI Integration Guide (rendered)."""
    return FileResponse(STATIC_DIR / "ai_guide.html",
                        headers={"Cache-Control": "no-store"})

@app.get("/ai-guide.md")
async def ai_guide_md():
    """Raw markdown for the AI guide (also what ChatGPT can read)."""
    return FileResponse(BASE_DIR / "AI_INTEGRATION_GUIDE.md",
                        media_type="text/markdown; charset=utf-8",
                        headers={"Cache-Control": "no-store"})

@app.get("/hmi")
async def hmi_page():
    """SCADA/HMI all-lines overview (wall display). Live KPIs/headcount/worker
    names from real data; process-sensor tiles are simulated demo values."""
    return FileResponse(
        STATIC_DIR / "hmi.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )

@app.get("/reports")
async def reports_page():
    """Reports page (separate from main console tabs). `no-store` so edits show
    on reload without a hard refresh."""
    return FileResponse(
        STATIC_DIR / "reports.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


@app.get("/m/report")
@app.get("/m/gantt")
async def mobile_report_page():
    """Mobile attendance viewer — serves the regular gantt page so the graphics
    are 100% identical to /gantt; the page itself detects the /m/ path and
    hides toolbar admin/print controls. `/m/gantt` is an alias for `/m/report`
    because operators sometimes type the latter when the desktop URL is /gantt."""
    return FileResponse(
        STATIC_DIR / "gantt.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


@app.get("/m/summary")
async def mobile_summary_page():
    """Mobile summary viewer — serves the regular summary page; mobile-mode JS
    hides admin/print/reload controls."""
    return FileResponse(
        STATIC_DIR / "summary.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


@app.get("/console")
async def console_page():
    """V3 Attendance Console — four-tab entry workflow."""
    return FileResponse(
        STATIC_DIR / "console.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )

@app.get("/dashboard")
async def dashboard_page():
    """Live operations dashboard → now the SCADA/HMI all-lines overview (bilingual
    JP/EN). Was a placeholder; replaced 2026-06-10. The old placeholder is still on
    disk at static/dashboard.html if it needs restoring."""
    return FileResponse(
        STATIC_DIR / "hmi.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


@app.get("/management")
async def management_page():
    """Mockup user-management GUI for roster reassignment approval."""
    return FileResponse(
        STATIC_DIR / "management.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


@app.get("/pdf/gantt")
async def pdf_gantt_page():
    """Standalone A4 PDF view for the daily Gantt — Print + one-click download.
    Embeds the existing /gantt page chrome-less (report=1); works for any date."""
    return FileResponse(
        STATIC_DIR / "pdf_view.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )


@app.get("/pdf/summary")
async def pdf_summary_page():
    """Standalone A4 PDF view for the productivity Summary — Print + one-click
    download. Embeds the existing /summary page chrome-less; any date/report."""
    return FileResponse(
        STATIC_DIR / "pdf_view.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )

CHANGELOG_PATH = BASE_DIR / "CHANGELOG.md"

# Map the raw area tags in CHANGELOG.md ("[ui]", "[api]"...) to plain-language
# labels a non-technical reader can scan. Anything unmapped stays visible as
# a lowercase tag so nothing is silently dropped.
_LOG_AREA_LABELS = {
    "ui": "What you see",
    "api": "Behind the scenes",
    "db": "Data storage",
    "ops": "Operations",
    "docs": "Documentation",
    "feature": "New feature",
    "bug": "Bug fix",
    "fix": "Bug fix",
    "security": "Security fix",
    "print": "Printing",
    "tooling": "Tooling",
    "ui/logic": "What you see",
    "ui/api": "What you see",
    "db/api": "Data storage",
}

_LOG_AREA_LABELS_JA = {
    "ui": "画面の表示",
    "api": "裏側の処理",
    "db": "データ保存",
    "ops": "運用",
    "docs": "ドキュメント",
    "feature": "新機能",
    "bug": "不具合修正",
    "fix": "不具合修正",
    "security": "セキュリティ対応",
    "print": "印刷",
    "tooling": "ツール",
    "ui/logic": "画面の表示",
    "ui/api": "画面の表示",
    "db/api": "データ保存",
}

# Hand-written Japanese summaries for each CHANGELOG entry so non-technical JA
# readers can scan the update log on /logs. Key is the stripped EN title exactly
# as produced by _strip_markdown on the heading's right side. Unlisted titles
# fall back to the EN title + "日本語訳は準備中です" so nothing disappears.
_LOG_ENTRIES_JA: dict[str, dict[str, str]] = {
    "BETA / test-mode banner + bilingual progress report": {
        "title": "BETAバナーの追加と進捗報告書の作成（英日バイリンガル）",
        "summary": "画面上部に「テスト中・データ検証中」のお知らせバーを追加しました。英語と日本語の両方で表示されます。あわせて、技術的な背景を知らない方でも分かる進捗・不具合の報告書を作成しました。",
    },
    "Daily Packs PDF: bulk upload speed-up + 504 timeout fix": {
        "title": "Daily Packs PDF：一括アップロードの高速化・タイムアウト解消",
        "summary": "タブ3のPDF一括アップロードをグループ分けして送信するように変更し、処理が速くなりました。100枚以上のPDFでもゲートウェイのタイムアウト（504エラー）が起きにくくなっています。",
    },
    "Daily Packs PDF: フルキャスト auto-extract + shift→prod date correction": {
        "title": "Daily Packs PDF：フルキャスト行の自動抽出・製造日付のずれを修正",
        "summary": "生産サマリーPDFをアップロードするだけで、パック数に加えてフルキャストの行（会社名・人数・出退勤時刻）も自動で読み取れるようになりました。PDFに印字された日付とシステム上の製造日付のずれも自動で修正します。",
    },
    "Database reset for clean verification": {
        "title": "お客様検証に向けたデータベース初期化",
        "summary": "お客様テストを始めるにあたり、勤怠データ・アップロード履歴・パック数・派遣スタッフ記録を初期化しました。社員名簿（employee_roster.json）は保持しています。",
    },
    "Security & data-integrity fixes (XSS + negative numbers + shift-window limits)": {
        "title": "セキュリティとデータ整合性の強化",
        "summary": "管理画面に氏名を通じた不正スクリプト挿入（XSS）が起きないように対策しました。フルキャスト人数やパック数にマイナス値を入れられない制限、シフトの開始・退勤時刻の範囲チェック、深夜勤務（翌日退勤）の正式対応も追加しています。",
    },
    "Management GUI mockup on dev branch": {
        "title": "管理画面（Management）のモックアップを追加",
        "summary": "パスワードでロックされた社員管理画面のモックアップを追加しました。ドラッグ＆ドロップで所属課を変更したり、並び替えたりできます。正式版ではログイン機能とサーバー側の権限管理に置き換えます。",
    },
    "V3 attendance console naming refresh": {
        "title": "アプリ名を「V3 Attendance Console」に統一",
        "summary": "アプリの表記を「V3 Attendance Console」に統一しました。ページタイトル・上部ブランド表示・ドキュメントのすべてで同じ名称に揃えています。",
    },
    "Summary PDF → B3 landscape · 3-month demo-data seeder": {
        "title": "サマリーPDFをB3横向きに変更・3か月分のデモデータ生成ツール",
        "summary": "サマリー画面のPDF出力をA4からB3横向きに変更し、グラフ・比較ブロック・14日間の表が1枚に収まるようにしました。あわせて、3か月分のリアルなデモデータを自動生成するツールを追加し、ダッシュボードの紹介がしやすくなりました。",
    },
    "Summary page (/attendance/summary) with target-line KPIs": {
        "title": "サマリーダッシュボードを新規追加（目標値付きKPI表示）",
        "summary": "全体と課別の人時生産性を一画面で見られる新しいサマリー画面を追加しました。日次・週次・月次・3か月の切替、前期間との比較、目標値（S1=85, S2=35, 合計=25 P/h）とのグラフ比較、14日間の詳細表を備えています。",
    },
    "Summary page ( /attendance/summary ) with target-line KPIs": {
        "title": "サマリーダッシュボードを新規追加（目標値付きKPI表示）",
        "summary": "全体と課別の人時生産性を一画面で見られる新しいサマリー画面を追加しました。日次・週次・月次・3か月の切替、前期間との比較、目標値（S1=85, S2=35, 合計=25 P/h）とのグラフ比較、14日間の詳細表を備えています。",
    },
    "Shift/Prod formula fix · Daily Packs DB sync · フルキャスト in Section 2": {
        "title": "シフト/製造日付の計算式を修正・Daily Packsの表示を修正・フルキャストを第2課に表示",
        "summary": "シフト日と製造日の判定式を「10時〜翌朝8時半」の実際のサイクルに合わせて書き直しました。Daily Packsの保存済みパック数が画面に表示されない不具合も修正しています。ガント画面では派遣スタッフ（フルキャスト）を第2課の末尾に表示します。",
    },
    "Dual time labels + responsive layout + previous-day delta arrows": {
        "title": "時刻表示の二重併記・レスポンシブ対応・前日比の矢印追加",
        "summary": "翌日の退勤時刻を「28:00 / 04:00」のように24時間連続軸と実時刻の両方で表示するようにしました。画面サイズに応じて自動で調整され、生産性パネルには前日比の増減矢印（▲/▼）が表示されます。",
    },
    "Gantt report works for every date (cache + graceful fallback)": {
        "title": "ガントレポートがすべての日付で正常表示",
        "summary": "ガントレポートが一部の日付でしか新デザインで表示されなかった問題を修正しました。キャッシュ制御を強化し、パック数が未保存の日付でもレイアウトが崩れないようにしました。",
    },
    "Attendance Report redesign (window, in-bar labels, productivity panel)": {
        "title": "勤怠レポートの刷新（時間軸・バー内ラベル・生産性パネル）",
        "summary": "ガントの時間軸を「10:00〜翌8:30」の22.5時間に変更し、開始・退勤時刻をバーの中に表示するデザインに刷新しました。上部にはパック数・第1課LP・第2課LP・合計LPを並べた生産性パネルを配置しています。",
    },
    "Attendance Gantt \"no data\" bugfix": {
        "title": "ガント表示の「データ無し」不具合を修正",
        "summary": "ある日付でデータが揃っているのに全員「出勤なし」と表示されていた不具合を修正しました。勤務時間の末尾に付く「 hr」の文字列で数値解釈に失敗していたのが原因です。",
    },
    "auto-upload button removed from UI": {
        "title": "自動取込ボタンを画面から一旦削除",
        "summary": "監視フォルダにPiからアクセスできないため「folder not found」のまま操作できない状態だった自動取込ボタンを画面から一旦外しました。APIは残しているため、フォルダ共有が整い次第復活できます。",
    },
    "console cleanup + auto-upload hardening": {
        "title": "コンソール画面の整理・自動取込機能の安定化",
        "summary": "コンソール画面のレイアウト不整合を整理し、時計ウィジェットを常に右上に表示するよう調整しました。自動取込の状態（準備中・N件・フォルダなし 等）を常に表示するようにしています。",
    },
    "v3.0 console hero/clock/auto-upload/date": {
        "title": "v3.0コンソールに時計・自動取込・日付ルールを追加",
        "summary": "コンソール画面にリアルタイム時計ウィジェットを追加し、現在時刻・シフト日・製造日を常に表示します。タブ1に自動取込ボタン、PDFからの日付（処理日）自動読み取りも追加しました。",
    },
    "Tab 2 backend + tab-sync rules": {
        "title": "フルキャスト入力のバックエンド実装・タブ間連携ルールの整備",
        "summary": "タブ2（フルキャスト）の保存先データベースと API を新設し、日付を変えると自動で保存済みデータを読み込むようにしました。タブ2・3・4の日付の連動ルールも整備しています。",
    },
    "v3.0 console first cut": {
        "title": "v3.0コンソール画面の初版を公開",
        "summary": "4つのタブ（勤怠PDF・フルキャスト・Daily Packs・レポート）で構成される新しいコンソール画面の初版を公開しました。既存のダッシュボードと同じデザインに揃えています。",
    },
}

def _strip_markdown(text: str, is_heading: bool = False) -> str:
    """Turn CHANGELOG line fragments into plain prose for the logs page.

    Headings keep inline code and → arrows (they're part of the title, not
    a file-reference pointer). Only body bullets get the full scrub.
    """
    if is_heading:
        # Unwrap backticks so inline code still reads naturally in the title.
        text = re.sub(r"`([^`]*)`", r"\1", text)
    else:
        # Body bullets: inline code is almost always a file path or flag the
        # non-technical reader shouldn't see, so drop it entirely.
        text = re.sub(r"`[^`]*`", "", text)
        # Also drop the trailing "→ file references" pointer the CHANGELOG
        # uses at the end of each bullet.
        text = re.split(r"\s*→\s*", text, maxsplit=1)[0]
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"\*(.+?)\*", r"\1", text)
    return re.sub(r"\s+", " ", text).strip()

def _parse_changelog(raw: str) -> list[dict]:
    """Parse CHANGELOG.md into a list of {date, title, items[]} cards."""
    entries: list[dict] = []
    current: dict | None = None
    date_heading_re = re.compile(r"^##\s+(\d{4}-\d{2}-\d{2})\s+—\s+(.+?)\s*$")
    bullet_re = re.compile(r"^-\s+(?:\*\*\[([^\]]+)\]\*\*|\[([^\]]+)\])\s*(.*)$")
    plain_bullet_re = re.compile(r"^-\s+(.*)$")

    for raw_line in raw.splitlines():
        line = raw_line.rstrip()
        # Stop once we hit the trailing "Notes for future updates" section.
        if line.startswith("_Notes for future updates"):
            break
        heading = date_heading_re.match(line)
        if heading:
            if current:
                entries.append(current)
            current = {
                "date": heading.group(1),
                "title": _strip_markdown(heading.group(2), is_heading=True),
                "items": [],
            }
            continue
        if not current:
            continue
        match = bullet_re.match(line)
        if match:
            tag = (match.group(1) or match.group(2) or "").lower().strip()
            body = _strip_markdown(match.group(3) or "")
            if body:
                current["items"].append({
                    "area": _LOG_AREA_LABELS.get(tag, tag or "Update"),
                    "area_ja": _LOG_AREA_LABELS_JA.get(tag, tag or "更新"),
                    "tag": tag,
                    "text": body,
                })
            continue
        plain = plain_bullet_re.match(line)
        if plain and current["items"]:
            # Continuation sub-bullet — attach to the previous item as context.
            extra = _strip_markdown(plain.group(1))
            if extra:
                current["items"][-1]["text"] += " " + extra

    if current:
        entries.append(current)

    for entry in entries:
        # Prune empty items; keep entries that still have a title even if
        # bullets were all sub-technical so the user sees something per date.
        entry["items"] = [item for item in entry["items"] if item["text"]]
        # Attach hand-written JA title + summary so the /logs page can render
        # a genuinely non-technical Japanese view instead of raw EN prose.
        ja_override = _LOG_ENTRIES_JA.get(entry["title"])
        if ja_override:
            entry["title_ja"] = ja_override["title"]
            entry["summary_ja"] = ja_override["summary"]
        else:
            entry["title_ja"] = entry["title"]
            entry["summary_ja"] = ""
    return entries

@app.get("/logs")
async def logs_page():
    """Stand-alone update-log page, not linked from the main nav."""
    return FileResponse(
        STATIC_DIR / "logs.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )

@app.get("/api/logs")
async def logs_api():
    """Serve CHANGELOG.md parsed into non-technical timeline entries."""
    try:
        raw = CHANGELOG_PATH.read_text(encoding="utf-8")
    except FileNotFoundError:
        return {"entries": []}
    return {"entries": _parse_changelog(raw)}

# ---------------------------------------------------------------------------
# Day-off Schedule — operator-managed planned absences (vacation, day off, etc.)
# Stored in dayoff_schedule.json; consumed by the Management → Day-off tab and
# (in the future) overlaid on the gantt to mark expected non-working days.
# ---------------------------------------------------------------------------
DAYOFF_PATH = BASE_DIR / "dayoff_schedule.json"
_DAYOFF_LOCK = threading.Lock()
_DAYOFF_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _dayoff_load() -> dict:
    if not DAYOFF_PATH.exists():
        return {"schedule": {}, "updated_at": None, "highlight_unauthorized": False}
    try:
        d = json.loads(DAYOFF_PATH.read_text(encoding="utf-8"))
        if not isinstance(d, dict):
            return {"schedule": {}, "updated_at": None, "highlight_unauthorized": False}
        d.setdefault("schedule", {})
        d.setdefault("updated_at", None)
        d.setdefault("highlight_unauthorized", False)
        return d
    except Exception:
        return {"schedule": {}, "updated_at": None, "highlight_unauthorized": False}


def _dayoff_save(payload: dict) -> None:
    tmp = DAYOFF_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(DAYOFF_PATH)


NICKNAME_MAP_PATH = BASE_DIR / "nickname_map.json"
_NICKNAME_LOCK = threading.Lock()


def _nickname_load() -> dict:
    """Returns {nickname_lower: employee_code}. File shape: {nickname: code}."""
    if not NICKNAME_MAP_PATH.exists():
        return {}
    try:
        d = json.loads(NICKNAME_MAP_PATH.read_text(encoding="utf-8"))
        return {str(k).strip(): str(v).strip() for k, v in d.items() if k and v}
    except Exception:
        return {}


def _nickname_save(mapping: dict) -> None:
    tmp = NICKNAME_MAP_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(mapping, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    tmp.replace(NICKNAME_MAP_PATH)


def _norm_name(s: str) -> str:
    """Normalize a Japanese display name for matching:
    - strip whitespace, ideographic space, half/full-width spaces
    - drop the inner space between family / given name in roster names
      ("青山 京子" → "青山京子")
    """
    if not isinstance(s, str):
        return ""
    return s.replace("　", "").replace(" ", "").strip()


def _build_roster_index() -> tuple[dict, dict]:
    """Returns (by_name, by_code) for the current roster.
    by_name: normalized name → {code, name, section_id, section_label}
    by_code: code → same dict (for reverse lookups)"""
    by_name: dict[str, dict] = {}
    by_code: dict[str, dict] = {}
    section_lookup = {section["id"]: section["label"] for section in SECTIONS}
    for emp in EMPLOYEE_ROSTER:
        info = {
            "code": emp["employee_code"],
            "name": emp["name"],
            "section_id": SECTION_OF_CODE.get(emp["employee_code"]),
        }
        info["section_label"] = section_lookup.get(info["section_id"])
        by_name[_norm_name(emp["name"])] = info
        by_code[emp["employee_code"]] = info
    return by_name, by_code


# Strings that appear as column headers / role labels in the 定休表 Excel and
# must NOT be treated as employee names when scanning data rows.
_DAYOFF_HEADER_BLOCKLIST = {
    "PB昼間製造", "工場長", "昼社員", "２課昼製造", "２課　昼製造", "2課昼製造", "2課　昼製造",
    "製造１課", "役職", "加熱・加社員", "加熱・加 社員", "１課製造", "1課製造", "1F製麺社員", "２課班長", "2課班長",
    "製造３課", "製麺社員", "1F\n製麺社員", "2F\n製麺社員", "1F製麺社員", "2F製麺社員", "3課昼", "3課 昼",
    "営業", "パート", "事務・資材", "資材技能実習", "開発", "品質管理", "行事・備考",
    "夜2課", "2課役職", "N・N社員", "N・Y社員",
    "祝日",
}
_DAYOFF_HEADER_BLOCKLIST_NORM = {_norm_name(s) for s in _DAYOFF_HEADER_BLOCKLIST}


@app.post("/api/dayoff/import-excel")
async def dayoff_import_excel(file: UploadFile = File(...)):
    """Parse an uploaded 定休表 .xlsx, match each off-day cell against the
    employee roster (by full name + saved nickname map), and return a preview
    payload the UI can use to:
      • see how many entries matched / didn't match
      • collect manual mappings for unmatched names
      • apply the matched entries to the dayoff schedule

    The actual save to dayoff_schedule.json happens via /api/dayoff/apply-import
    so the UI can preview + edit mappings first.
    """
    raw = await file.read()
    if len(raw) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="file too large (>25 MB)")
    if not raw[:4] == b"PK\x03\x04":
        raise HTTPException(status_code=400, detail="upload is not a .xlsx file")
    tmp_path = Path("/tmp") / f"dayoff_import_{uuid4().hex}.xlsx"
    tmp_path.write_bytes(raw)
    try:
        wb = load_workbook(tmp_path, data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"openpyxl could not parse: {e}")

    by_name, _ = _build_roster_index()
    nick_map = _nickname_load()  # nickname (raw) → employee_code

    # entries collected from the Excel: (employee_match_key, date_iso, sheet, raw_excel_name)
    raw_entries: list[dict] = []
    unmatched_counts: dict[str, int] = {}
    sheet_stats: dict[str, dict] = {}
    date_min = None
    date_max = None

    for sh_name in wb.sheetnames:
        if sh_name.strip() == "初期設定":
            continue
        ws = wb[sh_name]
        n_rows = 0
        n_cells = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx < 5:
                continue
            if not row:
                continue
            date_v = row[0]
            if not isinstance(date_v, datetime):
                continue
            n_rows += 1
            iso = date_v.strftime("%Y-%m-%d")
            if date_min is None or date_v < date_min:
                date_min = date_v
            if date_max is None or date_v > date_max:
                date_max = date_v
            for v in row[10:]:
                if not isinstance(v, str):
                    continue
                norm = _norm_name(v)
                if not norm or len(norm) < 2:
                    continue
                if norm in _DAYOFF_HEADER_BLOCKLIST_NORM:
                    continue
                # Also skip pure-numeric or weekday-only cells
                if norm.isdigit() or norm in {"日", "月", "火", "水", "木", "金", "土"}:
                    continue
                n_cells += 1
                raw_entries.append({"raw": v, "norm": norm, "date": iso, "sheet": sh_name.strip()})
        sheet_stats[sh_name.strip()] = {"date_rows": n_rows, "off_cells": n_cells}

    # Resolve each entry → employee_code
    matched: list[dict] = []
    unmatched_by_name: dict[str, dict] = {}  # norm_name → {raw, count, dates: set}
    for ent in raw_entries:
        norm = ent["norm"]
        # 1. saved nickname map (case-sensitive on the raw value first, then the normalized one)
        code = nick_map.get(ent["raw"]) or nick_map.get(norm)
        info = None
        if code:
            for emp in EMPLOYEE_ROSTER:
                if emp["employee_code"] == code:
                    info = {"code": emp["employee_code"], "name": emp["name"]}
                    break
        # 2. exact match against roster full name (normalized)
        if not info and norm in by_name:
            info = {"code": by_name[norm]["code"], "name": by_name[norm]["name"]}
        if info:
            matched.append({
                "code": info["code"], "name": info["name"],
                "raw": ent["raw"], "date": ent["date"], "sheet": ent["sheet"],
            })
        else:
            slot = unmatched_by_name.setdefault(norm, {"raw": ent["raw"], "count": 0, "dates": set()})
            slot["count"] += 1
            slot["dates"].add(ent["date"])
            unmatched_counts[norm] = unmatched_counts.get(norm, 0) + 1

    # For each unmatched normalized name, suggest roster candidates whose
    # normalized full-name CONTAINS the unmatched string (surname-only or
    # given-name-only forms typical in 定休表). One candidate → auto-suggest.
    def _suggest(norm_excel: str) -> list[dict]:
        if not norm_excel:
            return []
        cands = []
        for nm, info in by_name.items():
            if norm_excel and (norm_excel in nm or nm.startswith(norm_excel) or nm.endswith(norm_excel)):
                cands.append({
                    "code": info["code"],
                    "name": info["name"],
                    "section_label": info["section_label"] or "",
                })
        return cands[:8]  # cap so the UI stays compact

    unmatched_list = sorted(
        [
            {
                "raw": s["raw"],
                "norm": k,
                "count": s["count"],
                "first_date": min(s["dates"]) if s["dates"] else "",
                "suggestions": _suggest(k),
            }
            for k, s in unmatched_by_name.items()
        ],
        key=lambda r: (-r["count"], r["raw"]),
    )

    return {
        "ok": True,
        "date_min": date_min.strftime("%Y-%m-%d") if date_min else None,
        "date_max": date_max.strftime("%Y-%m-%d") if date_max else None,
        "sheets": sheet_stats,
        "total_off_cells": len(raw_entries),
        "matched_count": len(matched),
        "unmatched_count": sum(s["count"] for s in unmatched_by_name.values()),
        "matched": matched,
        "unmatched": unmatched_list,
        "saved_nickname_map_size": len(nick_map),
    }


@app.post("/api/dayoff/apply-import")
async def dayoff_apply_import(payload: dict = Body(...)):
    """Take a list of resolved {code, date} entries from the import preview,
    fold them into the existing dayoff schedule, optionally save any user-
    supplied nickname mappings ({nickname: employee_code}) to nickname_map.json.

    Body:
      entries:    [ {code, date}, … ]      – matched entries to fold in
      mappings:   { nickname: code, … }    – user-confirmed nickname → code (saved)
      mode:       "replace_range" | "merge"
                  - replace_range: clear all dayoff entries for any code that
                    appears in `entries` over [date_min..date_max], then add.
                    Use when re-importing the same date window.
                  - merge: just union the new entries on top of existing.
    """
    entries = payload.get("entries") or []
    mappings = payload.get("mappings") or {}
    mode = payload.get("mode", "merge")
    if not isinstance(entries, list):
        raise HTTPException(status_code=400, detail="entries must be a list")
    if not isinstance(mappings, dict):
        raise HTTPException(status_code=400, detail="mappings must be an object")

    valid_codes = {e["employee_code"] for e in EMPLOYEE_ROSTER}

    # 1. persist nickname mappings (skip empty/invalid)
    if mappings:
        with _NICKNAME_LOCK:
            cur = _nickname_load()
            for nick, code in mappings.items():
                if isinstance(nick, str) and isinstance(code, str) and code.strip() in valid_codes:
                    cur[nick.strip()] = code.strip()
            _nickname_save(cur)

    # 2. validate entries + collect by code
    new_dates_by_code: dict[str, set] = {}
    date_min, date_max = None, None
    for ent in entries:
        code = str(ent.get("code", "")).strip()
        date = str(ent.get("date", "")).strip()
        if code not in valid_codes:
            continue
        if not _DAYOFF_DATE_RE.match(date):
            raise HTTPException(status_code=400, detail=f"bad date '{date}'")
        new_dates_by_code.setdefault(code, set()).add(date)
        if date_min is None or date < date_min:
            date_min = date
        if date_max is None or date > date_max:
            date_max = date

    # 3. fold into existing schedule
    with _DAYOFF_LOCK:
        cur = _dayoff_load()
        sched = cur.get("schedule") or {}
        if mode == "replace_range" and date_min and date_max:
            # Drop existing entries that fall in the imported window for affected codes
            for code in new_dates_by_code:
                old = set(sched.get(code, []))
                kept = {d for d in old if not (date_min <= d <= date_max)}
                sched[code] = sorted(kept)
        for code, dates in new_dates_by_code.items():
            existing = set(sched.get(code, []))
            existing.update(dates)
            sched[code] = sorted(existing)
        # prune any code that ended up empty
        sched = {c: ds for c, ds in sched.items() if ds}
        out = {"schedule": sched, "updated_at": datetime.utcnow().isoformat() + "Z"}
        _dayoff_save(out)

    return {
        "ok": True,
        "applied_codes": len(new_dates_by_code),
        "applied_dates_total": sum(len(v) for v in new_dates_by_code.values()),
        "date_min": date_min,
        "date_max": date_max,
        "updated_at": out["updated_at"],
        "saved_mappings": len(mappings),
    }


@app.get("/api/dayoff/schedule")
async def dayoff_get():
    """Return the full saved day-off schedule.
    Shape: { schedule: {employee_code: ["YYYY-MM-DD", ...]}, updated_at }"""
    return _dayoff_load()


@app.put("/api/dayoff/schedule")
async def dayoff_save(payload: dict = Body(...)):
    """Replace the saved schedule. Validates each date is YYYY-MM-DD and that
    employee codes are 8-digit strings present in EMPLOYEE_ROSTER. Preserves
    the `highlight_unauthorized` flag if not explicitly set in the payload."""
    raw = payload.get("schedule")
    if not isinstance(raw, dict):
        raise HTTPException(status_code=400, detail="schedule must be an object {code: [dates]}")
    valid_codes = {e["employee_code"] for e in EMPLOYEE_ROSTER}
    cleaned: dict[str, list[str]] = {}
    for code, dates in raw.items():
        code_s = str(code).strip()
        if code_s not in valid_codes:
            continue
        if not isinstance(dates, list):
            raise HTTPException(status_code=400, detail=f"dates for {code_s} must be a list")
        seen = []
        for d in dates:
            ds = str(d).strip()
            if not _DAYOFF_DATE_RE.match(ds):
                raise HTTPException(status_code=400, detail=f"bad date '{ds}' (expected YYYY-MM-DD)")
            if ds not in seen:
                seen.append(ds)
        if seen:
            cleaned[code_s] = sorted(seen)
    with _DAYOFF_LOCK:
        prev = _dayoff_load()
        flag = bool(payload.get("highlight_unauthorized", prev.get("highlight_unauthorized", False)))
        out = {
            "schedule": cleaned,
            "updated_at": datetime.utcnow().isoformat() + "Z",
            "highlight_unauthorized": flag,
        }
        _dayoff_save(out)
    return {"ok": True, **out}


@app.post("/api/dayoff/highlight-unauthorized")
async def dayoff_set_highlight(payload: dict = Body(...)):
    """Toggle the global 'highlight unauthorized absence' flag without touching
    the schedule. When ON, the gantt page renders absent employees in red when
    their absence is NOT in the saved day-off list."""
    flag = bool(payload.get("enabled"))
    with _DAYOFF_LOCK:
        cur = _dayoff_load()
        cur["highlight_unauthorized"] = flag
        cur["updated_at"] = datetime.utcnow().isoformat() + "Z"
        _dayoff_save(cur)
    return {"ok": True, "enabled": flag, "updated_at": cur["updated_at"]}


@app.get("/api/management/bootstrap")
async def management_bootstrap():
    """Return roster and section metadata for the management UI."""
    return management_payload_from_files()

@app.put("/api/management/roster")
async def management_save_roster(payload: dict = Body(...)):
    """Persist the management board.

    sections.json: rebuilt from incoming order — drag-reorder in UI becomes the
                   authoritative display order across the whole app.
    employee_roster.json: name-only registry. Existing roster ORDER is preserved
                   (never reordered from UI). Names update for known codes; new
                   codes are appended at the end; codes removed from UI are dropped.
    """
    raw_employees = payload.get("employees")
    if not isinstance(raw_employees, list):
        raise HTTPException(status_code=400, detail="employees must be a list.")

    section_by_id = {int(section["id"]): section for section in SECTIONS}
    next_codes_by_section = {section_id: [] for section_id in section_by_id}
    incoming_pairs: list[tuple[str, str]] = []  # (code, name) in UI order
    seen_codes: set[str] = set()

    for index, raw_employee in enumerate(raw_employees, start=1):
        if not isinstance(raw_employee, dict):
            raise HTTPException(status_code=400, detail=f"Employee row {index} is invalid.")

        code = str(raw_employee.get("code") or raw_employee.get("employee_code") or "").strip()
        name = str(raw_employee.get("name") or "").strip()
        try:
            section_id = int(raw_employee.get("section_id"))
        except (TypeError, ValueError):
            section_id = 0

        if not is_employee_code(code):
            raise HTTPException(status_code=400, detail=f"Employee row {index} has an invalid code.")
        if not name:
            raise HTTPException(status_code=400, detail=f"Employee {code} is missing a name.")
        # Defence-in-depth: even though the frontend escapes on render, the roster
        # JSON is consumed by other tools (Excel export, Gantt PDF, Grafana).
        # Reject control chars and HTML/SQL metacharacters at the boundary so a
        # poisoned name can never land in employee_roster.json in the first place.
        if len(name) > 100:
            raise HTTPException(status_code=400, detail=f"Employee {code}: name is too long (max 100 chars).")
        if any(ch in name for ch in "<>\"'`;\\") or any(ord(ch) < 0x20 for ch in name):
            raise HTTPException(
                status_code=400,
                detail=f"Employee {code}: name contains disallowed characters.",
            )
        if code in seen_codes:
            raise HTTPException(status_code=400, detail=f"Employee code {code} is duplicated.")
        if section_id not in section_by_id:
            raise HTTPException(status_code=400, detail=f"Employee {code} has an invalid section.")

        seen_codes.add(code)
        incoming_pairs.append((code, name))
        next_codes_by_section[section_id].append(code)

    # Re-apply '*' mute markers from the existing sections.json so a UI save
    # never silently strips them (the management UI doesn't surface mutes yet).
    def _reapply_mute(code: str) -> str:
        return f"{code}*" if code in MUTED_CODES else code

    next_sections = [
        {
            "id": int(section["id"]),
            "label": section["label"],
            "codes": [_reapply_mute(c) for c in next_codes_by_section[int(section["id"])]],
        }
        for section in SECTIONS
    ]

    incoming_by_code = dict(incoming_pairs)
    next_roster: list[dict[str, str]] = []
    appended: set[str] = set()
    # Phase 1: keep current roster order for codes still present (name updates allowed).
    for emp in EMPLOYEE_ROSTER:
        code = emp["employee_code"]
        if code in incoming_by_code:
            next_roster.append({"employee_code": code, "name": incoming_by_code[code]})
            appended.add(code)
    # Phase 2: brand-new codes added via UI go to the end of the roster registry.
    for code, name in incoming_pairs:
        if code not in appended:
            next_roster.append({"employee_code": code, "name": name})
            appended.add(code)

    try:
        write_json_atomic(EMPLOYEE_ROSTER_PATH, next_roster)
        write_json_atomic(SECTIONS_PATH, {"sections": next_sections})
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"Unable to save roster files: {exc}") from exc

    refresh_management_data(next_roster, next_sections)
    response = management_payload_from_files()
    response["saved_at"] = datetime.now().isoformat()
    return response

@app.post("/api/management/import-pdf")
async def management_import_pdf(file: UploadFile = File(...)):
    """Parse an attendance PDF and return raw employee rows for import selection."""
    original_name, file_path, _content = await save_uploaded_pdf(file)
    try:
        raw_records = parse_pdf_data(file_path)
    except ValueError as exc:
        return {
            "filename": original_name,
            "record_count": 0,
            "rows": [],
            "message": str(exc),
        }
    seen_codes: set[str] = set()
    rows: list[dict[str, str]] = []
    section_lookup = {section["id"]: section["label"] for section in SECTIONS}

    for record in raw_records:
        code = (record.get("employee_code") or "").strip()
        name = (record.get("name") or "").strip()
        if not is_employee_code(code) or not name or code in seen_codes:
            continue
        seen_codes.add(code)
        section_id = record.get("section_id") or SECTION_OF_CODE.get(code)
        try:
            section_id = int(section_id) if section_id is not None else None
        except (TypeError, ValueError):
            section_id = None
        rows.append({
            "employee_code": code,
            "name": name,
            "section_id": section_id,
            "section_label": record.get("section_label") or section_lookup.get(section_id, "未配属"),
        })

    return {
        "filename": original_name,
        "record_count": len(rows),
        "rows": rows,
    }

@app.get("/summary")
async def summary_page():
    """Attendance Summary dashboard — productivity KPIs with target lines,
    day/week/month range selector, and period-over-period comparisons.
    Targets: Section1=85 P/h, Section2=35 P/h, Combined=25 P/h.
    """
    return FileResponse(
        STATIC_DIR / "summary.html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate"},
    )

@app.get("/api/gantt/latest-date")
async def gantt_latest_date():
    """Most recent REPORT date for which any data exists (attendance or packs).
    Per DATE_RULES.md, report = shift + 1 = production. We take the latest
    shift date in attendance_records and add 1 day; if daily_packs has a
    later report date than that, use it instead."""
    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT GREATEST(
                    (SELECT MAX(record_date) + 1 FROM attendance_records),
                    (SELECT MAX(record_date)     FROM daily_packs)
                )
                """
            )
            row = cursor.fetchone()
            latest = row[0].isoformat() if row and row[0] else None
    return {"date": latest}

@app.get("/api/gantt/dates-with-data")
async def gantt_dates_with_data():
    """REPORT dates that have at least one non-empty attendance row OR a
    pack count. Each attendance shift_date is returned as shift + 1 so the
    UI picker navigates straight to the report URL."""
    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT DISTINCT report_date FROM (
                    SELECT (record_date + 1) AS report_date
                      FROM attendance_records
                     WHERE commute_time <> '' OR time_to_leave <> '' OR working_hours <> ''
                    UNION
                    SELECT record_date AS report_date
                      FROM daily_packs
                     WHERE number_of_packs > 0
                ) u
                ORDER BY report_date DESC
                LIMIT 30
                """
            )
            dates = [r[0].isoformat() for r in cursor.fetchall() if r[0]]
    return {"dates": dates}

def _gantt_clean_wh(value: str | None) -> str:
    """Normalize working_hours to plain H:MM (strip trailing 'hr'/'h')."""
    s = (value or "").strip()
    s = re.sub(r"\s*hrs?\.?\s*$", "", s, flags=re.IGNORECASE).strip()
    return s or "0:00"


def _gantt_wh_to_hours(clean_wh: str) -> float:
    if not clean_wh or clean_wh == "0:00":
        return 0.0
    try:
        h_str, m_str = clean_wh.split(":")
        return int(h_str) + int(m_str) / 60
    except (ValueError, AttributeError):
        return 0.0


def _gantt_hours_to_hhmm(hours: float) -> str:
    if hours <= 0:
        return "0:00"
    total_minutes = round(hours * 60)
    h, m = divmod(total_minutes, 60)
    return f"{h}:{m:02d}"


def _gantt_compute_for_date(cursor, record_date: str) -> tuple[list[dict], dict]:
    """Pull attendance + packs for one day's REPORT, bucket by section, compute productivity.

    `record_date` is the **report / production date** (= the date in the URL).
    Per DATE_RULES.md:
      attendance_records.record_date = shift date  = report − 1
      temp_staff.record_date          = shift date  = report − 1
      daily_packs.record_date         = report date = report
    So the SQL pulls attendance/temp_staff at (record_date − 1) and packs at record_date.

    Pure function of the cursor — reusable for current date and previous day,
    so both appear in the API response and the frontend can draw up/down deltas
    without a second round-trip.
    """
    cursor.execute(
        """
        SELECT DISTINCT ON (personal_code)
               personal_code AS code,
               full_name AS name,
               commute_time,
               time_to_leave,
               working_hours
        FROM attendance_records
        WHERE record_date = (%s::date - 1)
        ORDER BY personal_code, upload_date DESC
        """,
        (record_date,),
    )
    rows = cursor.fetchall()
    cursor.execute(
        "SELECT number_of_packs FROM daily_packs WHERE record_date = %s",
        (record_date,),
    )
    pack_row = cursor.fetchone()

    # Fetch フルキャスト / temp-staff entries for the matching SHIFT date
    # (= report − 1). Shown as synthetic rows at the END of 製造２課 and folded
    # into Section 2's productivity totals (hours + headcount). Each saved
    # temp_staff row becomes one synthetic gantt row — the group worked one
    # shift together.
    cursor.execute(
        """
        SELECT id, company, headcount, start_time, leave_time, leave_next_day,
               hours_per_person, total_hours, note
        FROM temp_staff
        WHERE record_date = (%s::date - 1)
        ORDER BY start_time, id
        """,
        (record_date,),
    )
    temp_rows = cursor.fetchall() or []

    buckets: dict[int, list[dict]] = {s["id"]: [] for s in SECTIONS}
    buckets[0] = []  # Unassigned
    for r in rows:
        code = r["code"]
        # If a code only ever appears as muted ('*') in sections.json, it has no
        # active home — drop the DB row entirely; the ghost injection below
        # represents it. Codes with at least one non-'*' placement route to
        # SECTION_OF_CODE (their effective active section).
        if code not in ACTIVE_CODES and code in MUTED_CODES:
            continue
        sid = SECTION_OF_CODE.get(code, 0)
        buckets[sid].append({
            "code": code,
            "name": r["name"] or "",
            "in": (r["commute_time"] or "").strip() or None,
            "out": (r["time_to_leave"] or "").strip() or None,
            "wh": _gantt_clean_wh(r["working_hours"]),
            "is_temp": False,
        })

    # Inject one Disabled ghost per '*' placement at the section the user wrote
    # it into (independent of where the code's active home is). This lets the
    # same employee appear as a ghost in section 1 AND as the real worker in
    # section 2 if both placements exist in sections.json.
    for muted_code, muted_sid in MUTED_PLACEMENTS:
        buckets[muted_sid].append({
            "code": muted_code,
            "name": name_for_code(muted_code),
            "in": None,
            "out": None,
            "wh": None,
            "is_temp": False,
            "muted": True,
        })

    # Sort each bucket by **per-section** position so the order in sections.json
    # is preserved within each section, even when the same code is also placed
    # elsewhere. BEFORE appending temp-staff rows so they stay pinned at the end.
    for sid in buckets:
        pos_map = SECTION_POSITIONS.get(sid, {})
        buckets[sid].sort(
            key=lambda row: (pos_map.get(row["code"], 10**9), bool(row.get("muted")), row["code"])
        )

    # Append temp_staff as synthetic rows at the end of Section 2 (製造２課).
    # Each row carries its group metadata (headcount, total_hours) so the
    # productivity math downstream can fold the group into Section 2 totals
    # without double-counting.
    temp_staff_total_hours = 0.0
    temp_staff_headcount = 0
    for t in temp_rows:
        headcount = int(t["headcount"] or 0)
        if headcount <= 0:
            continue
        hours_per_person = float(t["hours_per_person"] or 0)
        group_total_hours = float(t["total_hours"] or (hours_per_person * headcount))
        company = (t["company"] or "フルキャスト").strip() or "フルキャスト"
        leave_next_day = bool(t.get("leave_next_day"))
        # When overnight, re-encode leave as +24 notation (e.g. "02:45" -> "26:45")
        # to match the attendance_records.time_to_leave convention so the Gantt
        # renderer can draw the bar crossing midnight with a single hour parser.
        raw_leave = (t["leave_time"] or "").strip()
        if leave_next_day and re.match(r"^\d{1,2}:\d{2}$", raw_leave):
            lh, lm = raw_leave.split(":")
            leave_out = f"{int(lh) + 24}:{lm}"
        else:
            leave_out = raw_leave or None
        synthetic = {
            "code": f"TEMP-{t['id']}",
            "name": f"{company} × {headcount}名",
            "in": (t["start_time"] or None),
            "out": leave_out,
            # wh is per-person so the gantt bar length matches one person's shift
            # (the whole group shares identical start/leave).
            "wh": _gantt_hours_to_hhmm(hours_per_person),
            "is_temp": True,
            "leave_next_day": leave_next_day,
            "headcount": headcount,
            "total_hours": round(group_total_hours, 4),
        }
        buckets[2].append(synthetic)
        temp_staff_total_hours += group_total_hours
        temp_staff_headcount += headcount

    sections_out = [
        {"id": s["id"], "label": s["label"], "rows": buckets[s["id"]]}
        for s in SECTIONS
    ]
    if buckets[0]:
        sections_out.append({"id": 0, "label": "Unassigned", "rows": buckets[0]})

    total_packs = 0
    if pack_row and pack_row.get("number_of_packs") is not None:
        try:
            total_packs = int(pack_row["number_of_packs"])
        except (TypeError, ValueError):
            total_packs = 0

    prod_sections: list[dict] = []
    combined_hours = 0.0
    combined_present = 0
    for s in SECTIONS:
        sect_rows = buckets[s["id"]]
        # Muted rows ('*' marked codes) are display-only — exclude from every
        # productivity metric so they neither inflate denominators nor pollute totals.
        active_rows = [r for r in sect_rows if not r.get("muted")]
        # For temp-staff rows we use the group total_hours (= headcount × hours);
        # for regular rows we parse "H:MM" working_hours to a decimal.
        sect_hours = sum(
            (float(r["total_hours"]) if r.get("is_temp") else _gantt_wh_to_hours(r["wh"]))
            for r in active_rows
        )
        # staff_present counts: temp rows contribute their full headcount;
        # regular rows count as 1 if they actually clocked in.
        sect_present = sum(
            (int(r.get("headcount") or 0) if r.get("is_temp")
             else (1 if _gantt_wh_to_hours(r["wh"]) > 0 else 0))
            for r in active_rows
        )
        # staff_total: temp rows also contribute headcount (a 5-person
        # フルキャスト group is 5 slots), regular rows contribute 1 each.
        sect_total = sum(
            (int(r.get("headcount") or 0) if r.get("is_temp") else 1)
            for r in active_rows
        )
        sect_lp = (total_packs / sect_hours) if sect_hours > 0 else 0.0
        prod_sections.append({
            "id": s["id"],
            "label": s["label"],
            "staff_present": sect_present,
            "staff_total": sect_total,
            "total_hours": round(sect_hours, 4),
            "total_hours_hhmm": _gantt_hours_to_hhmm(sect_hours),
            "lp": round(sect_lp, 2),
        })
        combined_hours += sect_hours
        combined_present += sect_present

    lp_combined = (total_packs / combined_hours) if combined_hours > 0 else 0.0
    productivity = {
        "date": record_date,
        "total_packs": total_packs,
        "sections": prod_sections,
        "combined": {
            "staff_present": combined_present,
            "total_hours": round(combined_hours, 4),
            "total_hours_hhmm": _gantt_hours_to_hhmm(combined_hours),
            "lp": round(lp_combined, 2),
        },
        # Separate visibility into temp-staff contribution so the UI can show
        # "includes X フルキャスト hours" without re-summing the data.
        "temp_staff": {
            "headcount": temp_staff_headcount,
            "total_hours": round(temp_staff_total_hours, 4),
            "total_hours_hhmm": _gantt_hours_to_hhmm(temp_staff_total_hours),
            "row_count": len(temp_rows),
        },
    }
    return sections_out, productivity


@app.get("/api/gantt/{record_date}")
async def gantt_for_date(record_date: str):
    """Attendance rows for one date, grouped by section, with productivity summary
    and a *previous-day* productivity block so the frontend can draw delta arrows.

    Row order within each section is driven by the position of each code inside
    sections.json (single source of truth). When that file is updated and the
    service is restarted, every downstream output — Gantt display, reports,
    Excel exports — reflects the new ordering automatically. employee_roster.json
    is now a name-only registry; its order does not affect display.
    """
    try:
        current_dt = datetime.strptime(record_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="record_date must be YYYY-MM-DD")
    prev_date_str = (current_dt - timedelta(days=1)).isoformat()

    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            sections_out, productivity = _gantt_compute_for_date(cursor, record_date)
            _, prev_productivity = _gantt_compute_for_date(cursor, prev_date_str)

    productivity["previous_date"] = prev_date_str
    productivity["previous"] = {
        "total_packs": prev_productivity["total_packs"],
        "sections": prev_productivity["sections"],
        "combined": prev_productivity["combined"],
    }

    return {
        "date": record_date,
        "sections": sections_out,
        "productivity": productivity,
    }

# ============================================================================
# Dashboard — single-call snapshot for the Status page (production plan +
# today's actuals + 7-day trend + per-section productivity) plus a cached
# wttr.in proxy so the browser doesn't make a cross-origin request itself.
# ============================================================================
_WEATHER_CACHE: dict = {"data": None, "fetched_at": 0.0}
_WEATHER_TTL_SECONDS = 30 * 60  # wttr.in is generous but unmetered — 30 min is plenty.
_WEATHER_LOCATION = "Yamanashi"  # Pi sits in Yamanashi prefecture; override via env if needed.


def _wttr_fetch() -> dict | None:
    """Fetch a 3-day forecast from wttr.in. Returns None on any failure so the
    dashboard degrades gracefully (the card just shows '—' instead of crashing)."""
    import urllib.request
    import urllib.parse
    loc = urllib.parse.quote(os.environ.get("DASHBOARD_WEATHER_LOC", _WEATHER_LOCATION))
    url = f"https://wttr.in/{loc}?format=j1"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "IMMS-Dashboard/1.0"})
        with urllib.request.urlopen(req, timeout=6) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except Exception as exc:  # network error, JSON error, timeout — all degrade silently
        print(f"[DASHBOARD_WEATHER] fetch failed: {exc}")
        return None

    days_out = []
    for d in (payload.get("weather") or [])[:3]:
        try:
            day_iso = d.get("date")
            avg_c = int(d.get("avgtempC") or 0)
            max_c = int(d.get("maxtempC") or 0)
            min_c = int(d.get("mintempC") or 0)
            hourly = d.get("hourly") or []
            # Pick the noon entry (or the middle slot) for an at-a-glance condition.
            noon = next((h for h in hourly if h.get("time") in ("1200", "1100", "1300")), None) or (hourly[len(hourly)//2] if hourly else {})
            desc = ""
            try:
                desc = (noon.get("weatherDesc") or [{}])[0].get("value", "").strip()
            except Exception:
                desc = ""
            days_out.append({
                "date": day_iso,
                "avg_c": avg_c, "max_c": max_c, "min_c": min_c,
                "desc": desc or "—",
                "wind_kph": int(noon.get("windspeedKmph") or 0),
                "humidity": int(noon.get("humidity") or 0),
                "rain_chance": int(noon.get("chanceofrain") or 0),
            })
        except Exception:
            continue
    return {"location": loc, "days": days_out}


@app.get("/api/dashboard/weather")
async def dashboard_weather():
    """Cached wttr.in proxy. Returns at most _WEATHER_TTL_SECONDS-stale data."""
    import time
    now = time.time()
    cached = _WEATHER_CACHE.get("data")
    if cached and (now - _WEATHER_CACHE.get("fetched_at", 0)) < _WEATHER_TTL_SECONDS:
        return {**cached, "cached": True, "age_seconds": int(now - _WEATHER_CACHE["fetched_at"])}
    fresh = _wttr_fetch()
    if fresh is None:
        # If we have ANY stale data, prefer it over an empty card.
        if cached:
            return {**cached, "cached": True, "age_seconds": int(now - _WEATHER_CACHE["fetched_at"]), "stale": True}
        return {"location": _WEATHER_LOCATION, "days": [], "error": "fetch_failed"}
    _WEATHER_CACHE["data"] = fresh
    _WEATHER_CACHE["fetched_at"] = now
    return {**fresh, "cached": False, "age_seconds": 0}


@app.get("/api/dashboard/snapshot")
async def dashboard_snapshot(date: str | None = None):
    """Aggregate data the Status dashboard needs in one round-trip.

    Returns:
      • report_date  — the dashboard's anchor date (defaults to latest daily_packs row)
      • plan         — production_plan grouped by line (A/B/C) with item rows
      • today        — total + per-section actual packs for `report_date`
      • last_30_days — daily totals (date, packs, n, y) for the trailing 30 days
      • last_30_days_hours — per-section labor hours per day for productivity trend
      • productivity — today's per-section p/h with delta vs previous date
    """
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Anchor: explicit date if given, else latest daily_packs row, else today.
            if date:
                try:
                    anchor = datetime.strptime(date, "%Y-%m-%d").date()
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc
            else:
                cur.execute("SELECT MAX(record_date) AS d FROM daily_packs")
                row = cur.fetchone() or {}
                anchor = row.get("d") or datetime.now().date()

            anchor_iso = anchor.isoformat()
            prev_iso = (anchor - timedelta(days=1)).isoformat()
            month_start = anchor - timedelta(days=29)

            # Production plan for anchor date, grouped by line.
            cur.execute(
                """
                SELECT line_code, item_name, planned_qty, n_qty, y_qty, combined_qty,
                       start_time, takt_time, pph_target, required_staff
                FROM production_plan
                WHERE record_date = %s
                ORDER BY line_code, id
                """,
                (anchor,),
            )
            plan_rows = cur.fetchall() or []

            # Today's actual: total packs + per-section split (n_total = section 1 marker
            # in 製造予定表 ＮＹ; y_total = section 2 marker — naming is historical).
            cur.execute(
                "SELECT number_of_packs, n_total, y_total FROM daily_packs WHERE record_date = %s",
                (anchor,),
            )
            today_pack = cur.fetchone() or {}

            # 30-day trend: pad missing dates with 0 so the chart x-axis stays even.
            cur.execute(
                "SELECT record_date, number_of_packs, n_total, y_total FROM daily_packs "
                "WHERE record_date BETWEEN %s AND %s ORDER BY record_date",
                (month_start, anchor),
            )
            trend_by_date: dict[str, dict] = {}
            for r in (cur.fetchall() or []):
                ds = r["record_date"].isoformat() if hasattr(r["record_date"], "isoformat") else str(r["record_date"])
                trend_by_date[ds] = {
                    "packs": int(r["number_of_packs"] or 0),
                    "n": int(r["n_total"] or 0) if r.get("n_total") is not None else None,
                    "y": int(r["y_total"] or 0) if r.get("y_total") is not None else None,
                }

            # Per-section labor hours per day, last 30 days (shift_date = report_date - 1
            # so the hours align with the report dates above). Aggregated server-side
            # so the dashboard renders a productivity trend without N round-trips.
            cur.execute(
                """
                SELECT DISTINCT ON (personal_code, record_date)
                       personal_code, record_date, working_hours
                FROM attendance_records
                WHERE record_date BETWEEN (%s::date - 1) AND (%s::date - 1)
                ORDER BY personal_code, record_date, upload_date DESC
                """,
                (month_start, anchor),
            )
            hours_rows = cur.fetchall() or []
            cur.execute(
                "SELECT record_date, headcount, hours_per_person, total_hours "
                "FROM temp_staff WHERE record_date BETWEEN (%s::date - 1) AND (%s::date - 1)",
                (month_start, anchor),
            )
            temp_rows = cur.fetchall() or []

    # Plan grouping (group AFTER closing the cursor so the connection releases sooner).
    plan_by_line: dict[str, dict] = {}
    for r in plan_rows:
        ln = (r.get("line_code") or "").upper()
        bucket = plan_by_line.setdefault(ln, {"line_code": ln, "items": [], "total_planned": 0, "item_count": 0})
        bucket["items"].append(r)
        bucket["total_planned"] += int(r.get("planned_qty") or 0)
        bucket["item_count"] += 1
    plan_lines = [plan_by_line[k] for k in sorted(plan_by_line)]
    total_planned_all = sum(b["total_planned"] for b in plan_lines)

    # Reuse the gantt productivity computer for today + previous so the dashboard's
    # per-section p/h, hours, and headcount match the gantt page exactly.
    productivity_today = None
    productivity_prev = None
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                _, productivity_today = _gantt_compute_for_date(cur, anchor_iso)
                _, productivity_prev = _gantt_compute_for_date(cur, prev_iso)
    except Exception as exc:
        # Productivity is a nice-to-have — never fail the whole snapshot for it.
        print(f"[DASHBOARD_SNAPSHOT] productivity compute skipped: {exc}")

    # Per-section delta vs previous day.
    sections_perf = []
    today_secs = {s["id"]: s for s in (productivity_today or {}).get("sections", [])}
    prev_secs = {s["id"]: s for s in (productivity_prev or {}).get("sections", [])}
    for sid in sorted(today_secs):
        t = today_secs[sid]
        p = prev_secs.get(sid, {})
        today_lp = float(t.get("lp") or 0)
        prev_lp = float(p.get("lp") or 0)
        delta_pct = ((today_lp - prev_lp) / prev_lp * 100.0) if prev_lp > 0 else None
        sections_perf.append({
            "id": sid,
            "label": t.get("label"),
            "lp": today_lp,
            "lp_prev": prev_lp,
            "delta_pct": round(delta_pct, 1) if delta_pct is not None else None,
            "staff_present": t.get("staff_present"),
            "staff_total": t.get("staff_total"),
            "total_hours_hhmm": t.get("total_hours_hhmm"),
        })

    # Aggregate per-section hours per shift_date (then expose them keyed by REPORT
    # date = shift_date + 1, matching the daily_packs/report-date convention).
    hours_by_report_date: dict[str, dict[int, float]] = {}
    for r in hours_rows:
        code = r["personal_code"]
        sid = SECTION_OF_CODE.get(code)
        if sid is None or code in MUTED_CODES and code not in ACTIVE_CODES:
            continue
        shift_d = r["record_date"]
        report_d = (shift_d + timedelta(days=1)).isoformat() if hasattr(shift_d, "isoformat") else None
        if not report_d:
            continue
        hrs = _gantt_wh_to_hours(_gantt_clean_wh(r["working_hours"]))
        if hrs <= 0:
            continue
        hours_by_report_date.setdefault(report_d, {})
        hours_by_report_date[report_d][sid] = hours_by_report_date[report_d].get(sid, 0.0) + hrs
    for r in temp_rows:
        shift_d = r["record_date"]
        report_d = (shift_d + timedelta(days=1)).isoformat() if hasattr(shift_d, "isoformat") else None
        if not report_d:
            continue
        hpp = float(r["hours_per_person"] or 0)
        hc = int(r["headcount"] or 0)
        group_total = float(r["total_hours"] if r["total_hours"] is not None else (hpp * hc))
        if group_total <= 0:
            continue
        # フルキャスト attaches to section 2 (製造２課) per the gantt convention.
        hours_by_report_date.setdefault(report_d, {})
        hours_by_report_date[report_d][2] = hours_by_report_date[report_d].get(2, 0.0) + group_total

    # Build padded 30-day series (packs + per-section hours + computed p/h).
    days_series = []
    for i in range(30):
        d = (month_start + timedelta(days=i)).isoformat()
        rec = trend_by_date.get(d) or {}
        hours = hours_by_report_date.get(d) or {}
        sec1_hours = round(hours.get(1, 0.0), 2)
        sec2_hours = round(hours.get(2, 0.0), 2)
        total_hours = sec1_hours + sec2_hours
        packs = rec.get("packs") if rec else 0
        days_series.append({
            "date": d,
            "packs": packs or 0,
            "n_packs": rec.get("n"),
            "y_packs": rec.get("y"),
            "sec1_hours": sec1_hours,
            "sec2_hours": sec2_hours,
            "total_hours": round(total_hours, 2),
            "lp": round(packs / total_hours, 2) if total_hours > 0 and packs else None,
            "lp_sec1": round((rec.get("n") or 0) / sec1_hours, 2) if sec1_hours > 0 and rec.get("n") else None,
            "lp_sec2": round((rec.get("y") or 0) / sec2_hours, 2) if sec2_hours > 0 and rec.get("y") else None,
        })

    today_total = int(today_pack.get("number_of_packs") or 0)
    today_progress_pct = (today_total / total_planned_all * 100.0) if total_planned_all > 0 else None

    return {
        "report_date": anchor_iso,
        "previous_date": prev_iso,
        "plan": {
            "lines": plan_lines,
            "total_planned": total_planned_all,
        },
        "today": {
            "total_packs": today_total,
            "n_total": int(today_pack.get("n_total") or 0) if today_pack.get("n_total") is not None else None,
            "y_total": int(today_pack.get("y_total") or 0) if today_pack.get("y_total") is not None else None,
            "progress_pct": round(today_progress_pct, 1) if today_progress_pct is not None else None,
        },
        "last_30_days": days_series,
        "productivity": {
            "combined_today": (productivity_today or {}).get("combined") or {},
            "combined_prev": (productivity_prev or {}).get("combined") or {},
            "sections": sections_perf,
        },
    }


@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM upload_batches")
            batch_count = cursor.fetchone()[0]
    _ag = _agent_status_summary()
    return {
        "status": "healthy",
        "version": "2.0",
        "storage": "postgresql",
        "database": "postgresql",
        "saved_batches": batch_count,
        "agent": {
            "online": _ag.get("online", False),
            "paused": _ag.get("paused", False),
            "deactivated": _ag.get("paused", False),
            "last_seen": _ag.get("last_seen"),
            "last_upload": (_ag.get("last_upload") or {}).get("ts"),
        },
        "timestamp": datetime.now().isoformat()
    }

@app.get("/api/roster/codes", response_class=PlainTextResponse)
async def roster_codes_text(code: list[str] | None = Query(default=None)):
    """
    Return only the requested employee codes as plain text, one code per line.
    Example: /api/roster/codes?code=00000326&code=00000401
    """
    requested_codes = parse_requested_codes(code)
    matched_codes = [employee_code for employee_code in requested_codes if employee_code in EMPLOYEE_CODES]
    return "\n".join(matched_codes)


# ============================================================================
# Member Hours — section-filtered roster + per-day attendance over a range
# (powers the new gantt "Member Hours" tab that replaces the Productivity one)
# ============================================================================
def _hhmm_to_hours(s: str | None) -> float | None:
    if not s:
        return None
    m = re.match(r"^\s*(\d{1,2}):(\d{2})", str(s))
    if not m:
        return None
    return int(m.group(1)) + int(m.group(2)) / 60.0

@app.get("/api/members/list")
async def members_list(section: str = "all"):
    """List employees by section in sections.json order. `section` = "all" | "1" | "2"."""
    sec = section.strip().lower()
    items = []
    for code, sid, label in iter_codes_in_section_order():
        if sec in ("1", "2") and str(sid) != sec:
            continue
        items.append({
            "code": code,
            "name": name_for_code(code),
            "section_id": sid,
            "section_label": label,
        })
    return {"section": sec, "count": len(items), "items": items}


@app.get("/api/members/compare")
async def members_compare(
    from_: str = Query(..., alias="from"),
    to:    str = Query(...),
    codes: str = Query(""),
):
    """Per-day attendance for the requested employee codes between `from` and
    `to` (inclusive). Returns one entry per member with a `days[]` array
    aligned to the requested date range so the frontend can plot directly."""
    try:
        d_from = datetime.strptime(from_, "%Y-%m-%d").date()
        d_to   = datetime.strptime(to,    "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="from/to must be YYYY-MM-DD") from exc
    if d_from > d_to:
        d_from, d_to = d_to, d_from
    code_list = [c.strip() for c in codes.split(",") if c.strip()]
    if not code_list:
        raise HTTPException(status_code=400, detail="codes is required (comma-separated)")
    code_list = code_list[:30]  # safety cap

    # Cap the requested window to 3 months (~92 days) so the daily-strip
    # render stays responsive and the per-section/per-day p/h aggregation
    # below remains a single bounded query.
    if (d_to - d_from).days > 92:
        raise HTTPException(status_code=400, detail="Range cannot exceed 3 months (92 days)")

    # Build the date axis once — every output member gets the same shape so the
    # frontend can iterate one ts-by-index loop without per-member alignment.
    span_days = (d_to - d_from).days + 1
    date_axis = [(d_from + timedelta(days=i)).isoformat() for i in range(span_days)]

    rows: list[tuple] = []
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT personal_code, record_date, commute_time, time_to_leave, working_hours, full_name
                FROM attendance_records
                WHERE personal_code = ANY(%s) AND record_date BETWEEN %s AND %s
                """,
                (code_list, d_from, d_to),
            )
            rows = cur.fetchall()

    by_code: dict[str, dict] = {}
    for code in code_list:
        # EMPLOYEE_ROSTER_BY_ID maps code → name (string), not a dict.
        sid = SECTION_OF_CODE.get(code)
        by_code[code] = {
            "code": code,
            "name": EMPLOYEE_ROSTER_BY_ID.get(code) or "",
            "section_id":    sid,
            "section_label": SECTION_LABEL_BY_ID.get(sid),
            "days": [{"date": d, "in": None, "out": None, "work_hours": None} for d in date_axis],
        }
    date_index = {d: i for i, d in enumerate(date_axis)}
    for code, rdate, c_in, c_out, wh, fname in rows:
        if code not in by_code:
            continue
        if hasattr(rdate, "isoformat"):
            rdate_s = rdate.isoformat()
        else:
            rdate_s = str(rdate)
        idx = date_index.get(rdate_s)
        if idx is None:
            continue
        cell = by_code[code]["days"][idx]
        cell["in"]  = (c_in or "").strip() or None
        cell["out"] = (c_out or "").strip() or None
        cell["work_hours"] = _hhmm_to_hours(wh)
        if fname and not by_code[code]["name"]:
            by_code[code]["name"] = fname

    members = list(by_code.values())

    # ----- Per-day section p/h -----
    # p/h(day) = packs(day) / total_section_hours(day). Compute once for the
    # whole range then attach to each member's days[].
    section_hours_by_date: dict[str, dict[int, float]] = {}
    packs_by_date: dict[str, int] = {}
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT DISTINCT ON (personal_code, record_date)
                       personal_code, record_date, working_hours
                FROM attendance_records
                WHERE record_date BETWEEN %s AND %s
                ORDER BY personal_code, record_date, upload_date DESC
                """,
                (d_from, d_to),
            )
            for code, rdate, wh in cur.fetchall():
                sid = SECTION_OF_CODE.get(code, 0)
                ds = rdate.isoformat() if hasattr(rdate, "isoformat") else str(rdate)
                section_hours_by_date.setdefault(ds, {})
                section_hours_by_date[ds][sid] = section_hours_by_date[ds].get(sid, 0.0) + _gantt_wh_to_hours(_gantt_clean_wh(wh))
            cur.execute(
                "SELECT record_date, headcount, hours_per_person, total_hours "
                "FROM temp_staff WHERE record_date BETWEEN %s AND %s",
                (d_from, d_to),
            )
            for rdate, hc, hpp, tot in cur.fetchall():
                ds = rdate.isoformat() if hasattr(rdate, "isoformat") else str(rdate)
                group_total = float(tot if tot is not None else (float(hpp or 0) * int(hc or 0)))
                section_hours_by_date.setdefault(ds, {})
                section_hours_by_date[ds][2] = section_hours_by_date[ds].get(2, 0.0) + group_total
            cur.execute(
                "SELECT record_date, number_of_packs FROM daily_packs "
                "WHERE record_date BETWEEN %s AND %s",
                (d_from, d_to),
            )
            for rdate, np_ in cur.fetchall():
                ds = rdate.isoformat() if hasattr(rdate, "isoformat") else str(rdate)
                try:
                    packs_by_date[ds] = int(np_ or 0)
                except (TypeError, ValueError):
                    packs_by_date[ds] = 0

    for m in members:
        sid = m.get("section_id")
        for cell in m["days"]:
            packs = packs_by_date.get(cell["date"], 0)
            sh = section_hours_by_date.get(cell["date"], {}).get(sid, 0.0) if sid is not None else 0.0
            cell["pph"] = round(packs / sh, 2) if sh > 0 and packs > 0 else None

    # Per-member summary (days_worked, total_hours, earliest in, latest out, max p/h)
    for m in members:
        days = m["days"]
        worked = [d for d in days if d["work_hours"] is not None and d["work_hours"] > 0]
        total_h = round(sum(d["work_hours"] or 0 for d in worked), 2)
        in_mins  = [int(d["in"][:2]) * 60 + int(d["in"][3:5]) for d in worked if d["in"] and len(d["in"]) >= 5]
        out_mins = []
        for d in worked:
            o = d["out"]
            if not o or len(o) < 5:
                continue
            try:
                hh, mm = int(o[:2]), int(o[3:5])
                if hh < 5: hh += 24
                out_mins.append(hh * 60 + mm)
            except Exception:
                pass
        def _hhmm(x):
            if x is None: return None
            return f"{(x // 60) % 24:02d}:{x % 60:02d}"
        def _avg_hhmm(xs):
            if not xs: return None
            return _hhmm(sum(xs) // len(xs))
        max_pph = max((d["pph"] for d in worked if d.get("pph") is not None), default=None)
        m["summary"] = {
            "days_worked":  len(worked),
            "total_hours":  total_h,
            "avg_in":       _avg_hhmm(in_mins),
            "avg_out":      _avg_hhmm(out_mins),
            "earliest_in":  _hhmm(min(in_mins)) if in_mins else None,
            "latest_in":    _hhmm(max(in_mins)) if in_mins else None,
            "earliest_out": _hhmm(min(out_mins)) if out_mins else None,
            "latest_out":   _hhmm(max(out_mins)) if out_mins else None,
            "longest_day":  round(max((d["work_hours"] or 0 for d in worked), default=0), 2),
            "max_pph":      max_pph,
        }

    return {
        "from":   d_from.isoformat(),
        "to":     d_to.isoformat(),
        "dates":  date_axis,
        "count":  len(members),
        "members": members,
    }

@app.post("/api/preview")
async def preview_pdf(file: UploadFile = File(...)):
    """
    Upload PDF and preview extracted data
    """
    try:
        original_name, file_size, records, metadata = await parse_uploaded_pdf(file)
        preview_data = build_preview_payload(original_name, file_size, records)
        record_date = metadata.get("record_date")
        preview_data["record_date"] = record_date.isoformat() if record_date else None
        preview_data["number_of_packs"] = metadata.get("number_of_packs")

        parsed_data.append(preview_data)
        return preview_data
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error while previewing PDF: {exc}") from exc

@app.post("/api/preview-multiple")
async def preview_multiple_pdfs(files: list[UploadFile] = File(...)):
    """Preview a multi-file upload by showing the first file and total counts."""
    if not files:
        raise HTTPException(status_code=400, detail="Please upload at least one PDF file.")

    try:
        parsed_files = [await parse_uploaded_pdf(file) for file in files]
        preview_name, preview_size, preview_records, preview_metadata = parsed_files[0]
        total_records = sum(len(records) for _, _, records, _ in parsed_files)
        records_with_data = sum(count_records_with_data(records) for _, _, records, _ in parsed_files)

        preview_data = build_preview_payload(
            preview_name,
            preview_size,
            preview_records,
            file_count=len(parsed_files),
            preview_filename=preview_name,
        )
        preview_data["total_records"] = total_records
        preview_data["records_with_data"] = records_with_data
        preview_record_date = preview_metadata.get("record_date")
        preview_data["record_date"] = preview_record_date.isoformat() if preview_record_date else None
        preview_data["number_of_packs"] = preview_metadata.get("number_of_packs")
        preview_data["files"] = [
            {
                "filename": filename,
                "total_records": len(records),
                "records_with_data": count_records_with_data(records),
                "record_date": metadata.get("record_date").isoformat() if metadata.get("record_date") else None,
                "number_of_packs": metadata.get("number_of_packs"),
            }
            for filename, _file_size, records, metadata in parsed_files
        ]

        parsed_data.append(preview_data)
        return preview_data
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error while previewing PDFs: {exc}") from exc

# ==============================================================
# Auto-upload: watched folder → latest PDF → preview (and optional save)
# ==============================================================
# Hardcoded source. On Windows the user's folder is:
#     C:\Users\Buddhika\Desktop\人時生産性　PDF
# On the Raspberry Pi running this service, that folder must be reachable
# from the filesystem — either via a CIFS/Samba mount of the Windows share
# or by rsyncing the PDFs to a local path. Override via env var
# ATTENDANCE_AUTO_UPLOAD_DIR if needed.
# Two independent watched folders: one for attendance PDFs (就業日報),
# one for Daily Packs PDFs (人時生産性). Config schema:
#   { "attendance": {path,updated_at,source,history}, "daily_packs": {...} }
# Legacy flat schema { "path": ... } is auto-migrated to attendance.path.
DEFAULT_ATTENDANCE_DIR = "/var/www/attendance_app/auto_uploads/attendance"
DEFAULT_DAILY_PACKS_DIR = "/var/www/attendance_app/auto_uploads/daily_packs"

def _load_split_config() -> dict:
    cfg = _load_json(AUTO_UPLOAD_CONFIG_PATH, {}) or {}
    if not isinstance(cfg, dict):
        cfg = {}
    if "attendance" not in cfg and cfg.get("path"):
        cfg = {
            "attendance": {
                "path": cfg.get("path"),
                "updated_at": cfg.get("updated_at"),
                "source": cfg.get("source"),
                "history": cfg.get("history", []),
            },
            "daily_packs": {"path": DEFAULT_DAILY_PACKS_DIR, "history": []},
        }
    cfg.setdefault("attendance", {"path": DEFAULT_ATTENDANCE_DIR, "history": []})
    cfg.setdefault("daily_packs", {"path": DEFAULT_DAILY_PACKS_DIR, "history": []})
    return cfg

def _load_auto_upload_path(kind: str = "attendance") -> Path:
    cfg = _load_split_config()
    sub = cfg.get(kind, {}) if isinstance(cfg.get(kind), dict) else {}
    saved = sub.get("path")
    if saved:
        return Path(saved)
    if kind == "daily_packs":
        return Path(os.environ.get("DAILY_PACKS_AUTO_UPLOAD_DIR", DEFAULT_DAILY_PACKS_DIR))
    return Path(os.environ.get("ATTENDANCE_AUTO_UPLOAD_DIR", DEFAULT_ATTENDANCE_DIR))

AUTO_UPLOAD_DIR = _load_auto_upload_path("attendance")
DAILY_PACKS_AUTO_UPLOAD_DIR = _load_auto_upload_path("daily_packs")

def _set_auto_upload_path(new_path: str, source: str = "ui", kind: str = "attendance") -> Path:
    """Persist a new watched-folder path for the given kind (attendance|daily_packs)."""
    global AUTO_UPLOAD_DIR, DAILY_PACKS_AUTO_UPLOAD_DIR
    p = Path(new_path).expanduser()
    cfg = _load_split_config()
    sub = cfg.get(kind, {}) if isinstance(cfg.get(kind), dict) else {}
    history = sub.get("history", []) or []
    current = AUTO_UPLOAD_DIR if kind == "attendance" else DAILY_PACKS_AUTO_UPLOAD_DIR
    if str(current) and str(current) != str(p):
        history.append({"path": str(current), "replaced_at": datetime.now().isoformat(timespec="seconds")})
        history = history[-20:]
    cfg[kind] = {
        "path": str(p),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "history": history,
    }
    _save_json(AUTO_UPLOAD_CONFIG_PATH, cfg)
    if kind == "attendance":
        AUTO_UPLOAD_DIR = p
    else:
        DAILY_PACKS_AUTO_UPLOAD_DIR = p
    print(f"[AUTO_UPLOAD:{kind}] path updated → {p} (source={source})")
    return p

print(f"[AUTO_UPLOAD] attendance folder = {AUTO_UPLOAD_DIR} (exists={AUTO_UPLOAD_DIR.exists()})")
print(f"[AUTO_UPLOAD] daily_packs folder = {DAILY_PACKS_AUTO_UPLOAD_DIR} (exists={DAILY_PACKS_AUTO_UPLOAD_DIR.exists()})")

def _looks_like_windows_path(s: str) -> bool:
    """Detect Windows-only paths like 'E:\\foo' or 'C:/bar' that the Pi can't reach."""
    if not s:
        return False
    s = s.strip()
    if len(s) >= 2 and s[1] == ":" and s[0].isalpha():
        return True
    if s.startswith("\\\\"):  # UNC
        return True
    return False

def _resolve_watched_target(p: Path) -> tuple[Path | None, Path | None]:
    """Resolve the configured path into (folder, single_file).
    - If `p` is a directory → (p, None)
    - If `p` is a .pdf file → (p.parent, p)
    - Otherwise → (None, None)
    """
    if p.exists() and p.is_dir():
        return (p, None)
    if p.exists() and p.is_file() and p.suffix.lower() == ".pdf":
        return (p.parent, p)
    return (None, None)

def _pick_latest_pdf_in(folder: Path) -> Path | None:
    """Pick the newest attendance PDF in the configured watched location.

    The configured path may be either a directory of PDFs OR a single .pdf
    file (in which case that file is the answer).
    Filenames embed the date as the last date-like token; sorting the
    filenames descending gives us the latest automatically. As a tie-
    breaker we fall back to mtime.
    """
    if folder.exists() and folder.is_file() and folder.suffix.lower() == ".pdf":
        return folder
    if not folder.exists() or not folder.is_dir():
        return None
    pdfs = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"]
    if not pdfs:
        return None
    pdfs.sort(key=lambda p: (p.name, p.stat().st_mtime), reverse=True)
    return pdfs[0]


def _pick_pdf_for_date(folder: Path, target_date) -> Path | None:
    """Find the PDF in `folder` whose filename encodes `target_date`.
    Returns None if the folder has no matching file (caller turns this
    into a 'file_not_available' 404)."""
    if folder is None or not folder.exists() or not folder.is_dir() or target_date is None:
        return None
    matches = []
    for p in folder.iterdir():
        if not (p.is_file() and p.suffix.lower() == ".pdf"):
            continue
        d = _extract_date_from_filename(p.name)
        if d == target_date:
            matches.append(p)
    if not matches:
        return None
    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0]


# ============================================================================
# Done folders — files that have been successfully ingested into the DB.
# Layout: <watched-folder>/done/<original-filename>. After a successful save
# the source file is moved here so the watched folder only ever shows pending
# work; the management UI exposes Restore + Delete for files in Done.
# ============================================================================
_DONE_SUBDIR = "done"


def _done_dir_for(kind: str) -> Path:
    """Resolve the Done subfolder for a kind. Creates it on demand."""
    parent = AUTO_UPLOAD_DIR if kind == "attendance" else DAILY_PACKS_AUTO_UPLOAD_DIR
    done = parent / _DONE_SUBDIR
    done.mkdir(parents=True, exist_ok=True)
    return done


def _parent_dir_for(kind: str) -> Path:
    """Resolve the watched (pending) folder for a kind."""
    return AUTO_UPLOAD_DIR if kind == "attendance" else DAILY_PACKS_AUTO_UPLOAD_DIR


_DONE_WARNINGS_LOG = BASE_DIR / "logs" / "done_warnings.log"


def _log_done_warning(kind: str, src: Path | None, exc: Exception) -> None:
    """Append a one-line JSON warning to logs/done_warnings.log so the
    desktop client can later retrieve it via GET /api/v1/done-files/warnings.
    Best-effort: failure to log is itself swallowed (we never want logging
    to roll back a successful DB save)."""
    try:
        _DONE_WARNINGS_LOG.parent.mkdir(parents=True, exist_ok=True)
        entry = {
            "ts":       datetime.now().isoformat(timespec="seconds"),
            "kind":     kind,
            "filename": src.name if src is not None else None,
            "src_path": str(src) if src is not None else None,
            "error":    f"{type(exc).__name__}: {exc}",
        }
        with _DONE_WARNINGS_LOG.open("a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception as log_exc:
        print(f"[DONE] failed to write done_warnings.log: {log_exc}")


def _move_to_done(src: Path, kind: str) -> Path | None:
    """Move `src` into the Done subfolder for `kind`. Best-effort: any error
    is logged (stdout + logs/done_warnings.log) and swallowed so a successful
    DB save is never rolled back by a filesystem hiccup. Overwrites a same-
    named file in Done if one exists. Returns the destination Path on
    success, None on failure."""
    try:
        if src is None or not src.exists() or not src.is_file():
            return None
        done = _done_dir_for(kind)
        dest = done / src.name
        if dest.exists():
            try:
                dest.unlink()
            except Exception:
                pass
        try:
            src.rename(dest)
        except OSError:
            # Cross-device rename → copy + unlink
            shutil.copy2(src, dest)
            src.unlink()
        return dest
    except Exception as exc:
        print(f"[DONE] move_to_done failed for {src} ({kind}): {exc}")
        _log_done_warning(kind, src, exc)
        return None


def _list_done_files(kind: str) -> list[dict]:
    """Enumerate files in the Done folder for `kind`, newest filename first."""
    done = _done_dir_for(kind)
    rows: list[dict] = []
    if not done.exists():
        return rows
    suffixes = (".pdf",) if kind == "attendance" else (".xlsx", ".xlsm")
    for p in sorted(done.iterdir(), key=lambda x: x.name, reverse=True):
        if not (p.is_file() and p.suffix.lower() in suffixes):
            continue
        if p.name.startswith("~$"):
            continue
        st = p.stat()
        d = _extract_date_from_filename(p.name)
        rows.append({
            "filename": p.name,
            "size": st.st_size,
            "modified": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
            "extracted_date": d.isoformat() if d else None,
        })
    return rows


@app.get("/api/attendance/auto-upload/info")
async def attendance_auto_upload_info():
    """Return the configured watched folder + contents summary (for display in UI)."""
    path_str = str(AUTO_UPLOAD_DIR)
    folder, single_file = _resolve_watched_target(AUTO_UPLOAD_DIR)
    reachable = folder is not None
    latest_name: str | None = None
    pdf_count = 0
    if reachable:
        try:
            if single_file is not None:
                pdf_count = 1
                latest_name = single_file.name
            else:
                pdfs = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"]
                pdf_count = len(pdfs)
                picked = _pick_latest_pdf_in(folder)
                latest_name = picked.name if picked else None
        except Exception:
            pass
    return {
        "path": path_str,
        "exists": reachable,
        "kind": "file" if single_file else ("folder" if reachable else "missing"),
        "pdf_count": pdf_count,
        "latest_filename": latest_name,
        "is_windows_path": _looks_like_windows_path(path_str),
        "env_override": "ATTENDANCE_AUTO_UPLOAD_DIR",
    }


@app.get("/api/daily-packs/auto-upload/info")
async def daily_packs_auto_upload_info():
    """Same shape as /api/attendance/auto-upload/info but for the packs folder."""
    path_str = str(DAILY_PACKS_AUTO_UPLOAD_DIR)
    folder, single_file = _resolve_watched_target(DAILY_PACKS_AUTO_UPLOAD_DIR)
    reachable = folder is not None
    latest_name = None
    pdf_count = 0
    if reachable:
        try:
            if single_file is not None:
                pdf_count = 1
                latest_name = single_file.name
            else:
                pdfs = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"]
                pdf_count = len(pdfs)
                picked = _pick_latest_pdf_in(folder)
                latest_name = picked.name if picked else None
        except Exception:
            pass
    return {
        "path": path_str,
        "exists": reachable,
        "kind": "file" if single_file else ("folder" if reachable else "missing"),
        "pdf_count": pdf_count,
        "latest_filename": latest_name,
        "is_windows_path": _looks_like_windows_path(path_str),
        "env_override": "DAILY_PACKS_AUTO_UPLOAD_DIR",
    }


@app.post("/api/attendance/auto-upload")
async def attendance_auto_upload(save: bool = False, date: str | None = None):
    """Preview (and optionally save) a PDF from the watched folder.

    Without `date`: picks the most-recent PDF (legacy behavior).
    With `date=YYYY-MM-DD`: picks the PDF whose filename encodes that date.
    If no matching file exists the response is a structured 404 carrying
    `error="file_not_available"`, the missing date, and the watched folder
    so the desktop client can show a clear "file not on server" message.

    Response mirrors /api/preview-multiple for a single file when save=False,
    and additionally saves the records to the DB when save=True.
    """
    folder, single_file = _resolve_watched_target(AUTO_UPLOAD_DIR)
    if folder is None:
        path_str = str(AUTO_UPLOAD_DIR)
        if _looks_like_windows_path(path_str):
            raise HTTPException(
                status_code=404,
                detail=(
                    f"Configured path is a Windows-only path the Pi cannot read: {path_str}. "
                    f"Either (a) push PDFs to the Pi via POST /api/v1/pdf/upload (X-API-Key required) — "
                    f"recommended for an E:\\ drive on a personal PC; or "
                    f"(b) mount the Windows folder on the Pi via CIFS/SMB and set the watched folder "
                    f"to the resulting Linux path like /mnt/companydata/AttendancePDF/."
                ),
            )
        raise HTTPException(
            status_code=404,
            detail=f"Watched path not reachable: {path_str}. "
                   f"Set a valid folder (or single .pdf path) under the Pi's filesystem in /attendance/console.",
        )

    target_date = None
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc

    if target_date is not None:
        picked = _pick_pdf_for_date(folder, target_date) if single_file is None else (
            single_file if _extract_date_from_filename(single_file.name) == target_date else None
        )
        if picked is None:
            raise HTTPException(
                status_code=404,
                detail={
                    "error": "file_not_available",
                    "kind": "attendance_pdf",
                    "requested_date": target_date.isoformat(),
                    "watched_folder": str(folder),
                    "message": f"No attendance PDF for {target_date.isoformat()} in {folder}. "
                               f"Upload the PDF first via POST /api/v1/pdf/upload.",
                },
            )
    else:
        picked = single_file if single_file is not None else _pick_latest_pdf_in(folder)
        if picked is None:
            raise HTTPException(status_code=404, detail=f"No PDF files found in {folder}.")

    # Parse from the filesystem path directly — no re-upload needed.
    try:
        records = apply_employee_roster(parse_pdf_data(picked))
        metadata = extract_pdf_metadata(picked)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to parse {picked.name}: {exc}") from exc

    record_date = metadata.get("record_date")
    pack_count = metadata.get("number_of_packs")
    file_size = picked.stat().st_size

    preview_data = build_preview_payload(
        picked.name,
        file_size,
        records,
        file_count=1,
        preview_filename=picked.name,
    )
    preview_data["record_date"] = record_date.isoformat() if record_date else None
    preview_data["number_of_packs"] = pack_count
    preview_data["picked_filename"] = picked.name
    preview_data["picked_size"] = file_size
    preview_data["files"] = [
        {
            "filename": picked.name,
            "total_records": len(records),
            "records_with_data": count_records_with_data(records),
            "record_date": record_date.isoformat() if record_date else None,
            "number_of_packs": pack_count,
        }
    ]

    if save:
        mismatches = find_attendance_mismatches(record_date, records)
        excel_filename, _excel_path, records_processed, batch_id = export_records_to_excel(
            picked.name,
            records,
            record_date=record_date,
        )
        if record_date is not None and pack_count is not None:
            upsert_pack_count_for_date(record_date, pack_count, source=picked.name)
        preview_data["saved"] = True
        preview_data["records_processed"] = records_processed
        preview_data["batch_id"] = batch_id
        preview_data["excel_filename"] = excel_filename
        preview_data["download_url"] = build_download_url(excel_filename)
        preview_data["mismatches"] = mismatches
        preview_data["mismatch_count"] = len(mismatches)
        # File now lives in DB → move source PDF to <watched>/done/ so the
        # next Auto-update doesn't re-pick it. Best-effort: failures here are
        # logged but never roll back a successful DB save.
        moved = _move_to_done(picked, "attendance")
        preview_data["moved_to_done"] = bool(moved)
        if moved is not None:
            preview_data["done_path"] = str(moved)
    else:
        preview_data["saved"] = False

    return preview_data


# ============================================================================
# Auto-upload config (settable from UI/CLI; persisted in auto_upload_config.json)
# ============================================================================
def _auto_upload_kind_payload(kind: str) -> dict:
    cfg = _load_split_config()
    sub = cfg.get(kind, {}) if isinstance(cfg.get(kind), dict) else {}
    current = AUTO_UPLOAD_DIR if kind == "attendance" else DAILY_PACKS_AUTO_UPLOAD_DIR
    return {
        "kind": kind,
        "path": str(current),
        "exists": current.exists() and current.is_dir(),
        "is_persisted": bool(sub.get("path")),
        "updated_at": sub.get("updated_at"),
        "source": sub.get("source"),
        "history": sub.get("history", []),
        "env_override": "ATTENDANCE_AUTO_UPLOAD_DIR" if kind == "attendance" else "DAILY_PACKS_AUTO_UPLOAD_DIR",
    }

@app.get("/api/auto-upload/config")
async def auto_upload_config_get(kind: str = "attendance"):
    if kind not in ("attendance", "daily_packs"):
        raise HTTPException(status_code=400, detail="kind must be 'attendance' or 'daily_packs'")
    return _auto_upload_kind_payload(kind)

@app.get("/api/auto-upload/config/all")
async def auto_upload_config_all():
    return {
        "attendance": _auto_upload_kind_payload("attendance"),
        "daily_packs": _auto_upload_kind_payload("daily_packs"),
    }

@app.post("/api/auto-upload/config")
async def auto_upload_config_set(payload: dict = Body(...)):
    raw = (payload.get("path") or "").strip()
    kind = (payload.get("kind") or "attendance").strip()
    if kind not in ("attendance", "daily_packs"):
        raise HTTPException(status_code=400, detail="kind must be 'attendance' or 'daily_packs'")
    if not raw:
        raise HTTPException(status_code=400, detail="path is required")
    p = _set_auto_upload_path(raw, source=payload.get("source", "ui"), kind=kind)
    return {
        "ok": True,
        "kind": kind,
        "path": str(p),
        "exists": p.exists() and p.is_dir(),
    }


# ============================================================================
# Structured API v1 — key-protected endpoints for app/web/CLI clients
# ============================================================================
@app.get("/api/v1/ping")
async def v1_ping(key_label: str = Depends(require_api_key)):
    return {"ok": True, "key": key_label, "server_time": datetime.now().isoformat(timespec="seconds")}


@app.get("/api/v1/status/precheck")
async def v1_status_precheck(
    type: str = Query(..., description="'attendance' or 'production'"),
    date: str = Query(..., description="ISO YYYY-MM-DD — the production/record date the file is FOR"),
    sha256: str = Query(..., min_length=64, max_length=64, description="lowercase hex SHA-256 of the file the desktop is about to send"),
    key_label: str = Depends(require_api_key),
):
    """Single-call decision endpoint for the desktop agent.

    Replaces the previous /api/v1/status/check + /api/v1/status/check-sha pair.
    The data table is the source of truth; uploaded_file_registry is treated
    as a derived index that gets reconciled on every call.

    Flow:
      1. Probe the right data table by `type`:
           attendance → attendance_records WHERE record_date = date
           production → daily_packs (then daily_pack_items) WHERE record_date = date
      2. If row count == 0:
           DELETE FROM uploaded_file_registry WHERE target_date=date OR sha256=sha
           → {"action": "upload", "reason": "no_data"}
      3. If rows exist, look up uploaded_file_registry by sha256:
           hash matches and target_date matches → {"action": "skip",            "reason": "same_file"}
           hash absent or different             → {"action": "confirm_replace", "reason": "different_file"}

    The desktop agent should:
      - "upload"          → POST the file to /api/v1/pdf/upload or /api/v1/xlsx/upload
      - "skip"            → do nothing
      - "confirm_replace" → ask the user; if confirmed, upload (server will overwrite)
      - "paused"          → admin paused ingestion; wait and retry later
    """
    if INGESTION_PAUSED:
        return {"action": "paused", "deactivated": True,
                "reason": "app deactivated by admin — agent should go idle",
                "type": type, "date": date, "sha256": sha256}
    if type not in ("attendance", "production"):
        raise HTTPException(status_code=400, detail="type must be 'attendance' or 'production'")
    if not re.fullmatch(r"[0-9a-f]{64}", sha256):
        raise HTTPException(status_code=400, detail="sha256 must be 64 lowercase hex chars")
    target_date = _parse_iso_date_or_400(date)
    iso = target_date.isoformat()

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            # 1. Data probe — table choice is driven by request type.
            if type == "attendance":
                cur.execute(
                    "SELECT COUNT(*), MAX(upload_date) FROM attendance_records WHERE record_date = %s",
                    (target_date,),
                )
                row = cur.fetchone() or (0, None)
                data_count = int(row[0] or 0)
                latest = row[1]
            else:
                cur.execute(
                    "SELECT number_of_packs, updated_at FROM daily_packs WHERE record_date = %s",
                    (target_date,),
                )
                pack_row = cur.fetchone()
                if pack_row:
                    data_count = 1
                    latest = pack_row[1]
                else:
                    cur.execute(
                        "SELECT COUNT(*), MAX(uploaded_at) FROM daily_pack_items WHERE production_date = %s",
                        (target_date,),
                    )
                    item_row = cur.fetchone() or (0, None)
                    data_count = int(item_row[0] or 0)
                    latest = item_row[1]

            # 2. No data → reconcile registry, instruct upload.
            if data_count == 0:
                cur.execute(
                    "DELETE FROM uploaded_file_registry WHERE target_date = %s OR sha256 = %s",
                    (target_date, sha256),
                )
                conn.commit()
                return {
                    "action": "upload",
                    "reason": "no_data",
                    "type": type,
                    "date": iso,
                    "sha256": sha256,
                }

            # 3. Data exists — registry decides skip vs confirm-replace.
            cur.execute(
                "SELECT target_date, status FROM uploaded_file_registry WHERE sha256 = %s",
                (sha256,),
            )
            reg_row = cur.fetchone()

    if reg_row and reg_row[0] == target_date and reg_row[1] in ("loaded", "received"):
        return {
            "action": "skip",
            "reason": "same_file",
            "type": type,
            "date": iso,
            "sha256": sha256,
            "uploaded_at": latest.isoformat() if latest else None,
        }

    return {
        "action": "confirm_replace",
        "reason": "different_file",
        "type": type,
        "date": iso,
        "sha256": sha256,
        "existing_records": data_count,
        "uploaded_at": latest.isoformat() if latest else None,
    }


PI_FALLBACK_UPLOAD_DIR = BASE_DIR / "auto_uploads"

def _resolve_upload_target_dir() -> Path:
    """Pick the directory to write incoming PDFs into.

    Defensive: if the configured path is unusable on the Pi (Windows-only
    path, or a path whose parent resolves to something relative like '.'),
    fall back to a known-good directory under BASE_DIR/auto_uploads.
    """
    cfg = AUTO_UPLOAD_DIR
    cfg_str = str(cfg)
    if _looks_like_windows_path(cfg_str):
        return PI_FALLBACK_UPLOAD_DIR
    if cfg.is_file() or cfg.suffix.lower() == ".pdf":
        cfg = cfg.parent
    if not cfg.is_absolute():
        return PI_FALLBACK_UPLOAD_DIR
    return cfg

# ---------------------------------------------------------------------------
# File-retention helpers (added 2026-05-01)
#
# Two rules the operator asked for:
#   1. ONE FILE PER DAY — when a new file is uploaded, any existing file
#      whose filename encodes the same production date (older copies, name
#      variations, etc.) is removed after the new upload succeeds.
#   2. 30-DAY RETENTION — files in the watched folders older than ~30 days
#      get cleaned up on demand, BUT only when the corresponding date's
#      data is confirmed saved in the database. Files for which the DB has
#      nothing get retained even past the threshold so the operator doesn't
#      lose unprocessed input.
# ---------------------------------------------------------------------------

def _extract_date_from_filename(name: str):
    """Pull YYYY-MM-DD (or YY.MM.DD) out of a filename. Handles full-width
    digits (２６.０４.２９) by NFKC-folding first, and accepts a few separator
    variants (`.`, `-`, `_`, `/`). Two-digit years are interpreted as 2000s."""
    if not name:
        return None
    s = unicodedata.normalize("NFKC", name)
    # Try 4-digit year first, then 2-digit
    candidates = (
        re.search(r"(\d{4})[\.\-_\s/]?(\d{1,2})[\.\-_\s/]?(\d{1,2})", s),
        re.search(r"(\d{2})[\.\-_\s/](\d{1,2})[\.\-_\s/](\d{1,2})", s),
    )
    for m in candidates:
        if not m:
            continue
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100:
                y += 2000
            return datetime(y, mo, d).date()
        except (ValueError, TypeError):
            continue
    return None


def _delete_same_date_older_files(folder: Path, keep_path: Path, target_date) -> list[str]:
    """Walk `folder`, delete every file (except `keep_path`) whose filename
    encodes `target_date`. Returns the list of removed filenames so the caller
    can include it in the upload response."""
    removed: list[str] = []
    if folder is None or not folder.exists() or target_date is None:
        return removed
    keep_resolved = keep_path.resolve()
    for f in folder.iterdir():
        try:
            if not f.is_file():
                continue
            if f.resolve() == keep_resolved:
                continue
            d = _extract_date_from_filename(f.name)
            if d == target_date:
                f.unlink()
                removed.append(f.name)
        except OSError:
            # don't crash the upload because of a stale handle / perm error
            continue
    return removed


def _db_has_data_for_date(d, kind: str) -> bool:
    """Cheap presence check for retention safety. Returns True if at least
    one row keyed on `d` exists in the relevant tables for `kind`."""
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            if kind == "attendance":
                cur.execute("SELECT 1 FROM attendance_records WHERE record_date = %s LIMIT 1", (d,))
            elif kind == "daily_packs":
                # Either daily_pack_items or daily_packs proves the day got saved.
                cur.execute(
                    "SELECT 1 FROM daily_pack_items WHERE production_date = %s "
                    "UNION ALL SELECT 1 FROM daily_packs WHERE record_date = %s LIMIT 1",
                    (d, d),
                )
            else:
                return False
            return cur.fetchone() is not None


def _scan_old_files(folder: Path, kind: str, days: int) -> list[dict]:
    """List files in `folder` along with extracted date + DB-data presence.
    `kind` is `attendance` or `daily_packs` and decides which tables to check."""
    out: list[dict] = []
    if folder is None or not folder.exists():
        return out
    cutoff = datetime.utcnow().date() - timedelta(days=days)
    for f in folder.iterdir():
        if not f.is_file():
            continue
        d = _extract_date_from_filename(f.name)
        is_old = (d is not None and d <= cutoff)
        # Files with no parseable date fall back to mtime
        if d is None:
            try:
                mtime = datetime.utcfromtimestamp(f.stat().st_mtime).date()
                is_old = mtime <= cutoff
                d = mtime
            except OSError:
                d = None
                is_old = False
        try:
            size = f.stat().st_size
        except OSError:
            size = 0
        in_db = _db_has_data_for_date(d, kind) if d is not None else False
        out.append({
            "filename": f.name,
            "extracted_date": d.isoformat() if d else None,
            "size": size,
            "is_older_than_threshold": is_old,
            "data_in_db": in_db,
            "safe_to_delete": is_old and in_db,
        })
    return sorted(out, key=lambda r: (r["extracted_date"] or "", r["filename"]))


# ---------------------------------------------------------------------------
# Uploaded-file registry — SHA-256 dedupe + lifecycle tracking. The desktop
# agent calls /api/v1/status/check (date) and /api/v1/status/check-sha (hash)
# before sending bytes; the upload endpoints below also dedupe server-side.
# ---------------------------------------------------------------------------
def _sha256_hex(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _registry_lookup_sha(sha256: str) -> dict | None:
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                "SELECT sha256, file_type, file_name, file_size, target_date, status, "
                "received_at, loaded_at, moved_path, record_count "
                "FROM uploaded_file_registry WHERE sha256 = %s",
                (sha256,),
            )
            return cur.fetchone()


def _registry_record_received(
    sha256: str,
    file_type: str,
    file_name: str,
    file_size: int,
    target_date,
    source_key: str | None,
) -> None:
    """Idempotent insert. Safe to call after _registry_lookup_sha returned None."""
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO uploaded_file_registry "
                "(sha256, file_type, file_name, file_size, target_date, source_key, status) "
                "VALUES (%s, %s, %s, %s, %s, %s, 'received') "
                "ON CONFLICT (sha256) DO NOTHING",
                (sha256, file_type, file_name, file_size, target_date, source_key),
            )
        conn.commit()


def _registry_mark_loaded(
    sha256: str,
    record_count: int | None,
    moved_path: str | None,
    target_date=None,
) -> None:
    """Called by the extract step once rows reach the DB. Flips status to
    'loaded', stamps loaded_at, records final path under done/."""
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE uploaded_file_registry SET "
                "status = 'loaded', loaded_at = NOW(), "
                "record_count = COALESCE(%s, record_count), "
                "moved_path = COALESCE(%s, moved_path), "
                "target_date = COALESCE(%s, target_date) "
                "WHERE sha256 = %s",
                (record_count, moved_path, target_date, sha256),
            )
        conn.commit()


def _parse_iso_date_or_400(s: str):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="date must be ISO YYYY-MM-DD")


def _resolve_packs_target_dir() -> Path:
    """Pick the directory to write incoming daily-packs files into. Mirrors
    `_resolve_upload_target_dir` but for the daily-packs watched folder."""
    cfg = DAILY_PACKS_AUTO_UPLOAD_DIR
    cfg_str = str(cfg)
    if _looks_like_windows_path(cfg_str):
        return BASE_DIR / "auto_uploads" / "daily_packs"
    if cfg.is_file():
        cfg = cfg.parent
    if not cfg.is_absolute():
        return BASE_DIR / "auto_uploads" / "daily_packs"
    return cfg


@app.post("/api/v1/xlsx/upload")
async def v1_xlsx_upload(file: UploadFile = File(...), key_label: str = Depends(require_api_key)):
    """Save an .xlsx into the daily-packs watched folder. Idempotent by filename.
    Companion to /api/v1/pdf/upload — used by the desktop client to push the
    daily フルキャスト Excel up so /api/daily-packs/auto-extract-excel can find it.

    One-file-per-day rule: after a successful save, any other file in the
    same folder whose filename encodes the same date is removed so only the
    latest copy lives in the watched folder.
    """
    if INGESTION_PAUSED:
        raise HTTPException(status_code=423, detail="Server ingestion is paused by admin.")
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="filename must end in .xlsx or .xlsm")
    target_dir = _resolve_packs_target_dir()
    target_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename).name
    dest = target_dir / safe_name
    content = await file.read()
    if not content[:4] == b"PK\x03\x04":
        raise HTTPException(status_code=400, detail="upload is not a valid .xlsx (zip magic missing)")
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="file too large (>50 MB)")
    sha = _sha256_hex(content)
    existing = _registry_lookup_sha(sha)
    if existing is not None:
        raise HTTPException(
            status_code=409,
            detail={
                "status": "duplicate",
                "reason": "sha256 already registered",
                "sha256": sha,
                "registry": {
                    "file_name": existing.get("file_name"),
                    "file_type": existing.get("file_type"),
                    "target_date": existing["target_date"].isoformat() if existing.get("target_date") else None,
                    "status": existing.get("status"),
                    "received_at": existing["received_at"].isoformat() if existing.get("received_at") else None,
                    "loaded_at": existing["loaded_at"].isoformat() if existing.get("loaded_at") else None,
                    "moved_path": existing.get("moved_path"),
                    "record_count": existing.get("record_count"),
                },
            },
        )
    dest.write_bytes(content)
    # One-file-per-day cleanup
    target_date = _extract_date_from_filename(safe_name)
    removed_same_date = _delete_same_date_older_files(target_dir, dest, target_date)
    _registry_record_received(sha, "production", safe_name, len(content), target_date, key_label)
    return {
        "ok": True,
        "filename": safe_name,
        "size": len(content),
        "sha256": sha,
        "stored_in": str(target_dir),
        "received_from_key": key_label,
        "extracted_date": target_date.isoformat() if target_date else None,
        "removed_same_date_files": removed_same_date,
    }


@app.post("/api/v1/pdf/upload")
async def v1_pdf_upload(file: UploadFile = File(...), key_label: str = Depends(require_api_key)):
    """Save a PDF into the auto-upload watched folder. Idempotent by filename.

    Falls back to BASE_DIR/auto_uploads if the configured path can't be used
    on the Pi (e.g. a Windows-only path like E:\\...).

    One-file-per-day rule: on success, other files in the same folder whose
    filename encodes the same date are removed (older copies / re-uploads
    with different naming) so the latest one is the only one parseable by
    the auto-extract endpoint.
    """
    if INGESTION_PAUSED:
        raise HTTPException(status_code=423, detail="Server ingestion is paused by admin.")
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="filename must end in .pdf")
    target_dir = _resolve_upload_target_dir()
    target_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename).name
    dest = target_dir / safe_name
    content = await file.read()
    validate_pdf_content(content)
    sha = _sha256_hex(content)
    existing = _registry_lookup_sha(sha)
    if existing is not None:
        raise HTTPException(
            status_code=409,
            detail={
                "status": "duplicate",
                "reason": "sha256 already registered",
                "sha256": sha,
                "registry": {
                    "file_name": existing.get("file_name"),
                    "file_type": existing.get("file_type"),
                    "target_date": existing["target_date"].isoformat() if existing.get("target_date") else None,
                    "status": existing.get("status"),
                    "received_at": existing["received_at"].isoformat() if existing.get("received_at") else None,
                    "loaded_at": existing["loaded_at"].isoformat() if existing.get("loaded_at") else None,
                    "moved_path": existing.get("moved_path"),
                    "record_count": existing.get("record_count"),
                },
            },
        )
    dest.write_bytes(content)
    target_date = _extract_date_from_filename(safe_name)
    removed_same_date = _delete_same_date_older_files(target_dir, dest, target_date)
    _registry_record_received(sha, "attendance", safe_name, len(content), target_date, key_label)
    return {
        "ok": True,
        "filename": safe_name,
        "size": len(content),
        "sha256": sha,
        "stored_in": str(target_dir),
        "received_from_key": key_label,
        "extracted_date": target_date.isoformat() if target_date else None,
        "removed_same_date_files": removed_same_date,
    }

@app.get("/api/v1/pdf/list")
async def v1_pdf_list(key_label: str = Depends(require_api_key)):
    if not AUTO_UPLOAD_DIR.exists():
        return {"path": str(AUTO_UPLOAD_DIR), "exists": False, "files": []}
    files = []
    for p in sorted(AUTO_UPLOAD_DIR.iterdir(), key=lambda x: x.name, reverse=True):
        if p.is_file() and p.suffix.lower() == ".pdf":
            st = p.stat()
            files.append({
                "filename": p.name,
                "size": st.st_size,
                "modified": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
            })
    return {"path": str(AUTO_UPLOAD_DIR), "exists": True, "count": len(files), "files": files}

@app.get("/api/v1/pdf/retrieve/{filename}")
async def v1_pdf_retrieve(filename: str, key_label: str = Depends(require_api_key)):
    safe = Path(filename).name
    target = AUTO_UPLOAD_DIR / safe
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail=f"{safe} not found in {AUTO_UPLOAD_DIR}")
    return FileResponse(target, media_type="application/pdf", filename=safe)


@app.get("/api/v1/xlsx/list")
async def v1_xlsx_list(key_label: str = Depends(require_api_key)):
    """List daily-packs Excel files already on the server. Each entry carries
    the parsed date from the filename so the desktop client can match a target
    report_date to a file before triggering /api/daily-packs/auto-extract-excel.
    """
    folder = DAILY_PACKS_AUTO_UPLOAD_DIR
    if not folder.exists():
        return {"path": str(folder), "exists": False, "files": []}
    files = []
    for p in sorted(folder.iterdir(), key=lambda x: x.name, reverse=True):
        if (
            p.is_file()
            and p.suffix.lower() in (".xlsx", ".xlsm")
            and not p.name.startswith("~$")
        ):
            st = p.stat()
            d = _extract_date_from_filename(p.name)
            files.append({
                "filename": p.name,
                "size": st.st_size,
                "modified": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
                "extracted_date": d.isoformat() if d else None,
            })
    return {"path": str(folder), "exists": True, "count": len(files), "files": files}

@app.post("/api/v1/pdf/auto-upload")
async def v1_pdf_auto_upload(
    save: bool = False,
    date: str | None = None,
    key_label: str = Depends(require_api_key),
):
    """Trigger preview/save of a PDF in the watched folder (key-protected).

    Without `date`: picks the latest PDF (legacy behavior — backward-compatible).
    With `date=YYYY-MM-DD`: picks the PDF whose filename encodes that date.
    Returns a structured `file_not_available` 404 when the requested date has
    no PDF in the watched folder.
    """
    if INGESTION_PAUSED:
        raise HTTPException(status_code=423, detail="Server ingestion is paused by admin.")
    return await attendance_auto_upload(save=save, date=date)


@app.get("/api/v1/done-files/warnings")
async def v1_done_warnings(
    limit: int = 200,
    kind: str | None = None,
    key_label: str = Depends(require_api_key),
):
    """Return recent move-to-done warnings (key-protected).

    Each warning is one JSON line written by `_log_done_warning` whenever a
    successful DB save was followed by a failed file move (filesystem error,
    cross-device rename failure, race with manual file deletion, etc.). The DB
    state is correct — only the file is still sitting in the watched folder
    instead of `done/`. The desktop client can poll this endpoint to surface
    the warning to the operator so they can retry the move (or just drag the
    file into Done themselves).

    Query params:
      - `limit` — max entries to return (newest last). Clamped to [1, 2000].
      - `kind`  — optional filter `attendance` | `daily_packs`.
    """
    n = max(1, min(int(limit or 200), 2000))
    if kind is not None and kind not in ("attendance", "daily_packs"):
        raise HTTPException(status_code=400, detail="kind must be 'attendance' or 'daily_packs'")
    if not _DONE_WARNINGS_LOG.exists():
        return {
            "path":     str(_DONE_WARNINGS_LOG),
            "exists":   False,
            "count":    0,
            "warnings": [],
        }
    entries: list[dict] = []
    try:
        with _DONE_WARNINGS_LOG.open("r", encoding="utf-8") as f:
            tail = f.readlines()[-n:]
        for line in tail:
            line = line.strip()
            if not line:
                continue
            try:
                e = json.loads(line)
            except Exception:
                continue
            if kind is not None and e.get("kind") != kind:
                continue
            entries.append(e)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"could not read warnings log: {exc}") from exc
    return {
        "path":     str(_DONE_WARNINGS_LOG),
        "exists":   True,
        "count":    len(entries),
        "warnings": entries,
    }


# ============================================================================
# /admin panel — server health, access log, login alerts, password-gated
# ============================================================================
def _require_admin_session(request: Request) -> None:
    token = request.cookies.get("admin_session")
    if not _admin_session_valid(token):
        raise HTTPException(status_code=401, detail="Admin login required.")


# ============================================================================
# GenbaLink user-facing auth endpoints (separate from /admin).
# Login form posts here; the page render of /login is served below.
# ============================================================================
def _gbl_require_user(request: Request) -> dict:
    user = _gbl_request_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Login required.")
    return user


def _gbl_require_admin(request: Request) -> dict:
    user = _gbl_require_user(request)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin only.")
    return user


@app.get("/login", response_class=HTMLResponse, include_in_schema=False)
async def gbl_login_page(request: Request):
    # Already logged in? Skip the form and bounce to next/.
    if _gbl_request_user(request) is not None:
        from fastapi.responses import RedirectResponse
        nxt = request.query_params.get("next") or "/dashboard"
        return RedirectResponse(url=nxt, status_code=302)
    return FileResponse(STATIC_DIR / "login.html")


@app.post("/api/auth/login")
async def gbl_api_login(request: Request, payload: dict = Body(...)):
    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    nxt = str(payload.get("next") or "/dashboard").strip() or "/dashboard"
    if not username or not password:
        raise HTTPException(status_code=400, detail="Username and password are required.")
    user = _gbl_find_user(username)
    if not user or not _gbl_verify_password(password, user.get("password_hash", "")):
        # Same vague message for both cases — don't leak which usernames exist.
        raise HTTPException(status_code=401, detail="Invalid username or password.")
    token = _gbl_session_create(user["username"], bool(user.get("is_admin")))
    # Only allow same-origin redirects to prevent open redirect.
    if not nxt.startswith("/"):
        nxt = "/dashboard"
    resp = JSONResponse({"ok": True, "redirect": nxt, "username": user["username"], "is_admin": bool(user.get("is_admin"))})
    resp.set_cookie(
        "gbl_session", token,
        httponly=True, samesite="lax", max_age=GBL_SESSION_TTL_SECONDS, path="/",
        # secure=True is set by the proxy when HTTPS is fronted by Nginx; FastAPI
        # has no easy way to know the fronting scheme reliably, so we leave it
        # off here and rely on Nginx adding `Secure` on the way out (or the
        # operator can flip this when HTTPS is live).
    )
    return resp


@app.post("/api/auth/logout")
async def gbl_api_logout(request: Request):
    _gbl_session_revoke(request.cookies.get("gbl_session"))
    resp = JSONResponse({"ok": True, "redirect": "/login"})
    resp.delete_cookie("gbl_session", path="/")
    return resp


@app.get("/logout", include_in_schema=False)
async def gbl_logout_page(request: Request):
    _gbl_session_revoke(request.cookies.get("gbl_session"))
    from fastapi.responses import RedirectResponse
    resp = RedirectResponse(url="/login", status_code=302)
    resp.delete_cookie("gbl_session", path="/")
    return resp


# ---- Browser auto-login (desktop agent → /pdf/* without a login form) -------
# Allowed redirect targets for /auto-login. Same-origin relative paths only;
# extend this tuple as more browser-facing pages need agent deep-links.
AUTO_LOGIN_REDIRECT_PREFIXES = ("/pdf/", "/dashboard")


@app.post("/api/v1/auth/login-token")
async def v1_auth_login_token(request: Request, key_label: str = Depends(require_api_key)):
    """Mint a 60s single-use token the agent can put in a browser URL.

    The agent authenticates here with its X-API-Key HEADER (never a URL), then
    opens /auto-login?token=...&redirect=... in the user's default browser.
    Keeps the long-lived API key out of browser history / Referer / logs.
    """
    token = _login_token_create(key_label)
    return {"token": token, "expires_in": _LOGIN_TOKEN_TTL_SECONDS,
            "redirect_to": "/auto-login"}


@app.get("/auto-login", include_in_schema=False)
async def auto_login(request: Request,
                     token: str = Query(...),
                     redirect: str = Query(...)):
    """Exchange a one-time login token for a gbl_session cookie + 302 redirect.

    Public path (no session required) — the agent's browser hits this first.
    Open-redirect hardened: same-origin relative paths from an allowlist only.
    """
    from fastapi.responses import RedirectResponse
    started = time.time()
    ip = (request.headers.get("cf-connecting-ip")
          or request.headers.get("x-real-ip")
          or (request.client.host if request.client else "?"))
    agent_version = request.headers.get("x-agent-version")

    def _fail(status: int, detail: str):
        _record_agent("/auto-login", "GET", status, ip, None,
                      f"redirect={redirect}", int((time.time() - started) * 1000),
                      agent_version)
        raise HTTPException(status_code=status, detail=detail)

    # 1. Open-redirect guard — same-origin relative paths only.
    if not redirect.startswith("/") or "://" in redirect or redirect.startswith("//"):
        _fail(400, "redirect must be a relative path")
    if redirect.startswith("/auto-login"):
        _fail(400, "redirect may not loop back to /auto-login")
    if not any(redirect.startswith(p) for p in AUTO_LOGIN_REDIRECT_PREFIXES):
        _fail(400, "redirect not in allowlist")

    # 2. Validate + consume the one-time token (this is the auth check).
    key_label = _login_token_consume(token)
    if key_label is None:
        _fail(401, "Missing, expired, or already-used token")

    # 3. Mint a normal gbl_session under a service identity (the key label).
    #    The agent key identifies the machine, not a human — sessions carry the
    #    label so the audit trail shows which agent key logged the browser in.
    session_token = _gbl_session_create(username=key_label, is_admin=False)

    # 4. Redirect + set the same cookie shape /api/auth/login issues. No
    #    secure= here — Nginx adds `Secure` on the HTTPS hop (see /api/auth/login).
    resp = RedirectResponse(url=redirect, status_code=302)
    resp.set_cookie("gbl_session", session_token, httponly=True,
                    samesite="lax", max_age=GBL_SESSION_TTL_SECONDS, path="/")
    _record_agent("/auto-login", "GET", 302, ip, key_label,
                  f"redirect={redirect}", int((time.time() - started) * 1000),
                  agent_version)
    return resp


@app.get("/api/auth/whoami")
async def gbl_api_whoami(request: Request):
    user = _gbl_request_user(request)
    if not user:
        return {"authenticated": False}
    return {
        "authenticated": True,
        "username": user["username"],
        "is_admin": bool(user.get("is_admin")),
        "expires_at": user["expires_at"],
    }


# ----- User management (admin only) -----
@app.get("/api/auth/users")
async def gbl_api_users_list(request: Request):
    _gbl_require_admin(request)
    return {"users": [
        {
            "username": u["username"],
            "is_admin": bool(u.get("is_admin")),
            "created_at": u.get("created_at"),
        } for u in GBL_USERS
    ]}


@app.post("/api/auth/users")
async def gbl_api_users_create(request: Request, payload: dict = Body(...)):
    _gbl_require_admin(request)
    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    is_admin = bool(payload.get("is_admin"))
    if not username or not password:
        raise HTTPException(status_code=400, detail="Username and password required.")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters.")
    if not re.match(r"^[A-Za-z0-9_.-]{2,32}$", username):
        raise HTTPException(status_code=400, detail="Username must be 2–32 chars: letters, digits, _ . -")
    if _gbl_find_user(username):
        raise HTTPException(status_code=400, detail=f"User '{username}' already exists.")
    GBL_USERS.append({
        "username": username,
        "password_hash": _gbl_hash_password(password),
        "is_admin": is_admin,
        "created_at": datetime.now().isoformat(),
    })
    _gbl_save_users(GBL_USERS)
    return {"ok": True, "username": username}


@app.delete("/api/auth/users/{username}")
async def gbl_api_users_delete(request: Request, username: str):
    actor = _gbl_require_admin(request)
    target = _gbl_find_user(username)
    if not target:
        raise HTTPException(status_code=404, detail="User not found.")
    if target["username"].lower() == actor["username"].lower():
        raise HTTPException(status_code=400, detail="Cannot delete your own account.")
    # Prevent removing the last admin.
    remaining_admins = [u for u in GBL_USERS if u.get("is_admin") and u["username"].lower() != target["username"].lower()]
    if target.get("is_admin") and not remaining_admins:
        raise HTTPException(status_code=400, detail="Cannot remove the last admin.")
    GBL_USERS[:] = [u for u in GBL_USERS if u["username"].lower() != target["username"].lower()]
    _gbl_save_users(GBL_USERS)
    return {"ok": True}


@app.post("/api/auth/change-password")
async def gbl_api_change_password(request: Request, payload: dict = Body(...)):
    user = _gbl_require_user(request)
    current = str(payload.get("current_password") or "")
    new_pw = str(payload.get("new_password") or "")
    if len(new_pw) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters.")
    record = _gbl_find_user(user["username"])
    if not record or not _gbl_verify_password(current, record.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Current password is incorrect.")
    record["password_hash"] = _gbl_hash_password(new_pw)
    _gbl_save_users(GBL_USERS)
    return {"ok": True}

async def _admin_root_impl(request: Request):
    token = request.cookies.get("admin_session")
    # no-store so the admin console is never served stale after an update
    # (avoids "I don't see the new tab" until a hard refresh).
    headers = {"Cache-Control": "no-store, must-revalidate"}
    if not _admin_session_valid(token):
        return FileResponse(STATIC_DIR / "admin_login.html", headers=headers)
    return FileResponse(STATIC_DIR / "admin.html", headers=headers)

@app.get("/admin", response_class=HTMLResponse, include_in_schema=False)
async def admin_root_no_slash(request: Request):
    return await _admin_root_impl(request)

@app.get("/admin/", response_class=HTMLResponse)
async def admin_root(request: Request):
    return await _admin_root_impl(request)

@app.post("/admin/login")
async def admin_login(password: str = Form(...)):
    expected = ADMIN_CONFIG.get("password", "")
    if not expected or not hmac.compare_digest(expected, password):
        return JSONResponse({"ok": False, "error": "Wrong password."}, status_code=401)
    token = ADMIN_CONFIG.get("session_token") or secrets.token_urlsafe(24)
    ADMIN_CONFIG["session_token"] = token
    _save_json(ADMIN_CONFIG_PATH, ADMIN_CONFIG)
    resp = JSONResponse({"ok": True, "redirect": "/admin"})
    resp.set_cookie("admin_session", token, httponly=True, samesite="lax", max_age=60 * 60 * 12, path="/admin")
    return resp

@app.post("/admin/logout")
async def admin_logout():
    ADMIN_CONFIG["session_token"] = secrets.token_urlsafe(24)
    _save_json(ADMIN_CONFIG_PATH, ADMIN_CONFIG)
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("admin_session", path="/admin")
    return resp

@app.get("/admin/api/status")
async def admin_status(request: Request):
    _require_admin_session(request)
    load = _sample_system_load()
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                cur.fetchone()
        db_ok = True
        db_err = None
    except Exception as exc:
        db_ok = False
        db_err = str(exc)
    auto_cfg = _load_json(AUTO_UPLOAD_CONFIG_PATH, {})
    return {
        "server_time": datetime.now().isoformat(timespec="seconds"),
        "system": load,
        "db": {"ok": db_ok, "error": db_err, "host": DATABASE_HOST, "name": DATABASE_NAME},
        "auto_upload": {
            "path": str(AUTO_UPLOAD_DIR),
            "exists": AUTO_UPLOAD_DIR.exists() and AUTO_UPLOAD_DIR.is_dir(),
            "updated_at": auto_cfg.get("updated_at") if isinstance(auto_cfg, dict) else None,
        },
        "api_keys_loaded": [k for k in API_KEYS.keys() if not k.startswith("_")],
        "access_log_path": str(ACCESS_LOG_PATH),
        "known_clients": len(KNOWN_CLIENTS),
        "alert_count": len(NEW_LOGIN_ALERTS),
    }

@app.get("/admin/api/access-log")
async def admin_access_log(request: Request, limit: int = 100):
    _require_admin_session(request)
    limit = max(1, min(limit, _RECENT_MAX))
    with _access_lock:
        items = list(_recent_access[-limit:])
    return {"count": len(items), "items": list(reversed(items))}

@app.get("/admin/api/debug-log")
async def admin_debug_log_state(request: Request, limit: int = 50):
    """Verbose API debug log: current toggle + last N entries from api_debug.jsonl."""
    _require_admin_session(request)
    limit = max(1, min(int(limit or 50), 500))
    file_size = 0
    entries: list[dict] = []
    if API_DEBUG_LOG_PATH.exists():
        try:
            file_size = API_DEBUG_LOG_PATH.stat().st_size
            with API_DEBUG_LOG_PATH.open("r", encoding="utf-8") as fh:
                tail = fh.readlines()[-limit:]
            for raw in tail:
                raw = raw.strip()
                if not raw:
                    continue
                try:
                    entries.append(json.loads(raw))
                except Exception:
                    continue
        except Exception as exc:
            return {
                "enabled": DEBUG_API_LOG_ENABLED,
                "path": str(API_DEBUG_LOG_PATH),
                "error": str(exc),
                "items": [],
            }
    return {
        "enabled": DEBUG_API_LOG_ENABLED,
        "path": str(API_DEBUG_LOG_PATH),
        "file_size_bytes": file_size,
        "count": len(entries),
        "items": list(reversed(entries)),
    }

@app.post("/admin/api/debug-log")
async def admin_debug_log_toggle(request: Request, payload: dict = Body(...)):
    """Flip the verbose-debug toggle live (no restart). Persists in admin_config.json."""
    _require_admin_session(request)
    enabled = bool(payload.get("enabled"))
    global DEBUG_API_LOG_ENABLED
    DEBUG_API_LOG_ENABLED = enabled
    ADMIN_CONFIG["debug_api_log"] = enabled
    try:
        _save_json(ADMIN_CONFIG_PATH, ADMIN_CONFIG)
    except Exception as exc:
        print(f"[API_DEBUG_LOG] persist failed: {exc}")
    print(f"[API_DEBUG_LOG] toggled enabled={enabled} via /admin/api/debug-log")
    return {"enabled": DEBUG_API_LOG_ENABLED}


# ---- AutoDoc tab — doctor.py run history -----------------------------------
# Reads logs/internal_health_logs.log (written every 5h by attendance-doctor.timer)
# and returns the last N runs parsed into structured sections so the admin UI
# can show "what doctor did" and highlight problems it auto-solved.

_DOCTOR_LOG_PATH = BASE_DIR / "logs" / "internal_health_logs.log"
_DOCTOR_CLONE_STATE = BASE_DIR / "logs" / "last_db_clone.json"

def _classify_doctor_line(line: str) -> str:
    """Bucket a doctor log line by severity."""
    upper = line.upper()
    if any(t in upper for t in ("ERROR", "FAILED", "FATAL", "INVALID", "MISSING", "WRONG_TYPE")):
        return "err"
    if "WARNING" in upper:
        return "warn"
    return "ok"

def _parse_doctor_runs(text: str, limit: int) -> list[dict]:
    """Split the log into runs and structure each into sections. Newest first."""
    import re as _re
    # Each run starts with `=== doctor run TIMESTAMP ===` and ends with `=== end TIMESTAMP ===`.
    chunks = _re.split(r"^=== doctor run (\S+) ===\s*$", text, flags=_re.MULTILINE)
    # split() yields [prefix, ts1, body1, ts2, body2, ...]
    runs: list[dict] = []
    for i in range(1, len(chunks) - 1, 2):
        started = chunks[i]
        body = chunks[i + 1]
        # body ends with `=== end TIMESTAMP ===`
        end_match = _re.search(r"^=== end (\S+) ===\s*$", body, flags=_re.MULTILINE)
        ended = end_match.group(1) if end_match else None
        if end_match:
            body = body[: end_match.start()]
        # Split body into sections by `[section_name]` lines
        sections: list[dict] = []
        current: dict | None = None
        for raw in body.splitlines():
            if not raw.strip():
                continue
            if raw.startswith("[") and raw.endswith("]"):
                if current is not None:
                    sections.append(current)
                current = {"name": raw[1:-1], "lines": [], "status": "ok"}
                continue
            if current is None:
                # lines before the first [section] header (e.g. FATAL traceback)
                current = {"name": "preamble", "lines": [], "status": "ok"}
            line = raw.strip()
            current["lines"].append(line)
            sev = _classify_doctor_line(line)
            if sev == "err" or (sev == "warn" and current["status"] != "err"):
                current["status"] = sev
        if current is not None:
            sections.append(current)
        # Per-run rollup
        sev_rank = {"ok": 0, "warn": 1, "err": 2}
        overall = max((s["status"] for s in sections), key=lambda s: sev_rank.get(s, 0), default="ok")
        # Duration in seconds (best-effort; both are ISO)
        duration = None
        try:
            if ended:
                duration = int((datetime.fromisoformat(ended) - datetime.fromisoformat(started)).total_seconds())
        except Exception:
            pass
        runs.append({
            "started_at": started,
            "ended_at": ended,
            "duration_seconds": duration,
            "overall_status": overall,
            "sections": sections,
        })
    runs.reverse()  # newest first
    return runs[:limit]

@app.get("/admin/api/doctor-runs")
async def admin_doctor_runs(request: Request, limit: int = 20):
    """Last N doctor.py runs from internal_health_logs.log, parsed into
    sections + severity for the AutoDoc admin tab. Also returns the most
    recent DB-clone state so the UI can show 'last clone' at a glance."""
    _require_admin_session(request)
    limit = max(1, min(int(limit or 20), 100))
    text = ""
    size = 0
    if _DOCTOR_LOG_PATH.exists():
        try:
            size = _DOCTOR_LOG_PATH.stat().st_size
            # Only read the tail — 256KB easily covers ~80 runs.
            with _DOCTOR_LOG_PATH.open("rb") as fh:
                fh.seek(max(0, size - 256 * 1024))
                text = fh.read().decode("utf-8", "replace")
        except Exception as exc:
            return {"error": str(exc), "log_path": str(_DOCTOR_LOG_PATH), "runs": [], "count": 0}
    runs = _parse_doctor_runs(text, limit)
    last_clone = None
    if _DOCTOR_CLONE_STATE.exists():
        try:
            last_clone = json.loads(_DOCTOR_CLONE_STATE.read_text(encoding="utf-8"))
        except Exception:
            last_clone = None
    # Pull the systemd timer's next-run hint (best-effort).
    next_run = None
    try:
        import subprocess as _sp
        out = _sp.run(
            ["systemctl", "list-timers", "attendance-doctor.timer", "--no-pager"],
            capture_output=True, text=True, timeout=5,
        )
        for ln in out.stdout.splitlines():
            if "attendance-doctor.timer" in ln:
                # Format: NEXT LEFT LAST PASSED UNIT ACTIVATES
                parts = ln.split()
                if len(parts) >= 4:
                    next_run = " ".join(parts[:4])
                break
    except Exception:
        pass
    return {
        "log_path": str(_DOCTOR_LOG_PATH),
        "file_size_bytes": size,
        "count": len(runs),
        "next_run": next_run,
        "last_clone": last_clone,
        "runs": runs,
    }


def _detect_lan_ip() -> str:
    """Best-effort primary LAN IPv4 (no traffic sent)."""
    try:
        import socket
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
        finally:
            s.close()
        if ip and not ip.startswith("127."):
            return ip
    except Exception:
        pass
    return "192.168.0.2"  # last-known fallback

@app.get("/admin/api/agent-status")
async def admin_agent_status(request: Request):
    """Live Desktop Agent status (admin only)."""
    _require_admin_session(request)
    return _agent_status_summary()

@app.get("/admin/api/agent-log")
async def admin_agent_log(request: Request, limit: int = 100):
    """Tail of the Desktop Agent request log + live status (admin only)."""
    _require_admin_session(request)
    limit = max(1, min(int(limit or 100), 1000))
    file_size = 0
    entries: list[dict] = []
    if AGENT_LOG_PATH.exists():
        try:
            file_size = AGENT_LOG_PATH.stat().st_size
            with AGENT_LOG_PATH.open("r", encoding="utf-8") as fh:
                tail = fh.readlines()[-limit:]
            for raw in tail:
                raw = raw.strip()
                if not raw:
                    continue
                try:
                    entries.append(json.loads(raw))
                except Exception:
                    continue
        except Exception as exc:
            return {"path": str(AGENT_LOG_PATH), "error": str(exc),
                    "status": _agent_status_summary(), "items": []}
    return {
        "path": str(AGENT_LOG_PATH),
        "file_size_bytes": file_size,
        "count": len(entries),
        "status": _agent_status_summary(),
        "items": list(reversed(entries)),
    }

@app.post("/admin/api/ingestion-toggle")
async def admin_ingestion_toggle(request: Request, payload: dict = Body(...)):
    """Pause/resume Desktop Agent ingestion (admin). Persists in admin_config.json.
    Paused → precheck returns {"action":"paused"} and uploads return HTTP 423."""
    _require_admin_session(request)
    paused = bool(payload.get("paused"))
    global INGESTION_PAUSED
    INGESTION_PAUSED = paused
    ADMIN_CONFIG["ingestion_paused"] = paused
    try:
        _save_json(ADMIN_CONFIG_PATH, ADMIN_CONFIG)
    except Exception as exc:
        print(f"[AGENT] persist failed: {exc}")
    print(f"[AGENT] ingestion_paused={paused} via /admin/api/ingestion-toggle")
    return {"paused": INGESTION_PAUSED}

# ---- Desktop-app update channel -------------------------------------------
def _app_update_public() -> dict:
    m = dict(APP_UPDATE_META)
    has = bool(m.get("filename"))
    return {
        "available": has,
        "version": m.get("version"),
        "filename": m.get("filename"),
        "sha256": m.get("sha256"),
        "size": m.get("size"),
        "notes": m.get("notes"),
        "mandatory": bool(m.get("mandatory", False)),
        "updated_at": m.get("updated_at"),
        "download_url": "/api/v1/app-update/download" if has else None,
    }

@app.get("/api/v1/app-update")
async def v1_app_update(key_label: str = Depends(require_api_key)):
    """Desktop agent polls this to learn if a newer app build is published.
    Compare `version` to the running version; if newer, GET download_url,
    verify sha256, apply, restart, then report the new X-Agent-Version."""
    return _app_update_public()

@app.get("/api/v1/app-update/download")
async def v1_app_update_download(key_label: str = Depends(require_api_key)):
    """Stream the published desktop-app update file."""
    fn = APP_UPDATE_META.get("filename")
    if not fn:
        raise HTTPException(status_code=404, detail="No app update published.")
    fp = APP_UPDATE_DIR / fn
    if not fp.exists():
        raise HTTPException(status_code=404, detail="Update file missing on server.")
    return FileResponse(fp, filename=fn, media_type="application/octet-stream")

@app.get("/admin/api/app-update")
async def admin_app_update_get(request: Request):
    """Current published update metadata (admin)."""
    _require_admin_session(request)
    return _app_update_public()

@app.post("/admin/api/app-update")
async def admin_app_update_set(
    request: Request,
    file: UploadFile | None = File(default=None),
    version: str = Form(default=""),
    notes: str = Form(default=""),
    mandatory: str = Form(default=""),
    clear: str = Form(default=""),
    local_filename: str = Form(default=""),
):
    """Publish (or clear) the desktop-app update file (admin, multipart).

    Three modes:
      * clear=1            → unpublish
      * local_filename=X   → FAST: register a file already in app_updates/
                             (scp it onto the Pi over LAN — no HTTP upload)
      * file=<multipart>   → upload (streamed to disk in chunks, hashed live)
    """
    _require_admin_session(request)
    global APP_UPDATE_META
    if clear.strip().lower() in ("1", "true", "yes", "on"):
        old = APP_UPDATE_META.get("filename")
        if old:
            try: (APP_UPDATE_DIR / old).unlink(missing_ok=True)
            except Exception as exc: print(f"[APP_UPDATE] unlink failed: {exc}")
        APP_UPDATE_META = {}
        _save_json(APP_UPDATE_CONFIG_PATH, APP_UPDATE_META)
        return {"ok": True, "cleared": True, "update": _app_update_public()}
    CAP = 300 * 1024 * 1024

    def _finalize(name: str, sha: str, size: int, registered: bool):
        global APP_UPDATE_META
        prev = APP_UPDATE_META.get("filename")
        if prev and prev != name:
            try: (APP_UPDATE_DIR / prev).unlink(missing_ok=True)
            except Exception: pass
        APP_UPDATE_META = {
            "filename": name,
            "version": version.strip() or datetime.now().strftime("%Y.%m.%d-%H%M"),
            "notes": notes.strip()[:1000],
            "mandatory": mandatory.strip().lower() in ("1", "true", "yes", "on"),
            "sha256": sha, "size": size,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        _save_json(APP_UPDATE_CONFIG_PATH, APP_UPDATE_META)
        print(f"[APP_UPDATE] {'registered' if registered else 'published'} "
              f"{name} v{APP_UPDATE_META['version']} ({size} bytes)")
        return {"ok": True, "registered": registered, "update": _app_update_public()}

    # ---- FAST PATH: register a file already on the server ------------------
    # Admin scp's the (big) installer into app_updates/ over LAN, then just
    # registers it here — zero HTTP upload, no Cloudflare/nginx body limit.
    if local_filename.strip():
        name = Path(local_filename.strip()).name
        fp = APP_UPDATE_DIR / name
        if not fp.is_file():
            raise HTTPException(status_code=404,
                detail=f"'{name}' not found in app_updates/ — scp it onto the Pi first.")
        h = hashlib.sha256(); size = 0
        with fp.open("rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                h.update(chunk); size += len(chunk)
        return _finalize(name, h.hexdigest(), size, registered=True)

    # ---- HTTP upload: stream to disk in chunks, hash live (no full RAM) ----
    if file is None or not file.filename:
        raise HTTPException(status_code=400, detail="No file or local_filename provided.")
    safe_name = Path(file.filename).name
    tmp = APP_UPDATE_DIR / (safe_name + ".part")
    h = hashlib.sha256(); size = 0
    try:
        with tmp.open("wb") as out:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > CAP:
                    raise HTTPException(status_code=413, detail="file too large (>300 MB)")
                h.update(chunk); out.write(chunk)
    except HTTPException:
        try: tmp.unlink(missing_ok=True)
        except Exception: pass
        raise
    except Exception as exc:
        try: tmp.unlink(missing_ok=True)
        except Exception: pass
        raise HTTPException(status_code=500, detail=f"upload write failed: {exc}")
    if size == 0:
        try: tmp.unlink(missing_ok=True)
        except Exception: pass
        raise HTTPException(status_code=400, detail="Empty file.")
    tmp.replace(APP_UPDATE_DIR / safe_name)
    return _finalize(safe_name, h.hexdigest(), size, registered=False)

@app.get("/admin/api/app-update/local-files")
async def admin_app_update_local_files(request: Request):
    """List files currently sitting in app_updates/ (for the fast register path)."""
    _require_admin_session(request)
    out = []
    try:
        for p in sorted(APP_UPDATE_DIR.iterdir()):
            if p.is_file() and not p.name.endswith(".part"):
                st = p.stat()
                out.append({"name": p.name, "size": st.st_size,
                            "mtime": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds")})
    except Exception as exc:
        return {"dir": str(APP_UPDATE_DIR), "error": str(exc), "files": []}
    return {"dir": str(APP_UPDATE_DIR), "files": out}

def _ip_is_lan(ip: str) -> bool:
    """True only for RFC1918 / loopback — the LAN uploader must not be usable
    through the public Cloudflare tunnel."""
    if not ip:
        return False
    ip = ip.strip()
    if ip in ("127.0.0.1", "::1", "localhost"):
        return True
    if ip.startswith(("192.168.", "10.")):
        return True
    if ip.startswith("172."):
        try:
            return 16 <= int(ip.split(".")[1]) <= 31
        except Exception:
            return False
    return False

_LAN_UPLOAD_HTML = """<!doctype html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>LAN App Update Upload — GENBA FMS</title>
<style>
:root{--bg:#0f1720;--card:#16212e;--line:#27384a;--ink:#e7eef6;--mut:#8aa0b4;--acc:#0d9488}
*{box-sizing:border-box} body{margin:0;background:var(--bg);color:var(--ink);
font:15px/1.5 "Noto Sans JP",system-ui,Segoe UI,sans-serif;display:grid;place-items:center;min-height:100vh;padding:1rem}
.card{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:1.4rem 1.5rem;width:min(560px,100%)}
h1{font-size:1.15rem;margin:.1rem 0 .2rem} .sub{color:var(--mut);font-size:.83rem;margin-bottom:1.1rem}
label{display:block;font-size:.78rem;color:var(--mut);margin:.7rem 0 .25rem}
input[type=text],input[type=password]{width:100%;padding:.55rem .65rem;background:#0f1925;border:1px solid var(--line);
color:inherit;border-radius:8px}
#drop{margin-top:.3rem;border:2px dashed var(--line);border-radius:10px;padding:1.1rem;text-align:center;color:var(--mut);cursor:pointer}
#drop.hot{border-color:var(--acc);color:var(--ink)}
.row{display:flex;gap:.6rem;align-items:center;flex-wrap:wrap;margin-top:.8rem}
button{font:inherit;font-weight:700;padding:.6rem 1.1rem;border:0;border-radius:9px;cursor:pointer;
background:linear-gradient(135deg,#0d9488,#0b6f58);color:#fff}
button.lang{padding:.3rem .65rem;font-size:.72rem;background:transparent;border:1px solid var(--line);color:var(--mut)}
button.lang:hover{color:var(--ink);border-color:var(--acc)}
button:disabled{opacity:.6;cursor:wait}
.head{display:flex;justify-content:space-between;align-items:flex-start;gap:.5rem}
.bar{height:10px;background:#0f1925;border:1px solid var(--line);border-radius:6px;overflow:hidden;margin-top:.9rem;display:none}
.bar>i{display:block;height:100%;width:0;background:var(--acc);transition:width .15s}
.msg{margin-top:.8rem;font-size:.85rem;white-space:pre-wrap;word-break:break-word}
.ok{color:#34d399}.err{color:#f87171} code{background:#0f1925;padding:.05rem .35rem;border-radius:4px}
</style>
<link rel=preconnect href="https://fonts.googleapis.com">
<link rel=preconnect href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@400;700&display=swap" rel=stylesheet>
</head><body><form class=card id=f onsubmit="return false">
<div class=head>
  <div>
    <h1 data-i18n=title>⚡ LAN App-Update Upload</h1>
    <div class=sub data-i18n=sub>Direct to the Pi over the local network — bypasses the Cloudflare 100&nbsp;MB limit. Big installers OK (≤300&nbsp;MB).</div>
  </div>
  <button type=button class=lang id=lang>日本語</button>
</div>
<label data-i18n=l_pw>Admin password</label><input type=password id=pw autocomplete=current-password>
<label data-i18n=l_file>Update file</label>
<div id=drop data-i18n=drop>Drop file here, or click to choose<input type=file id=file hidden></div>
<div id=fname class=sub style="margin:.4rem 0 0"></div>
<div class=row>
<div style="flex:1;min-width:160px"><label data-i18n=l_ver>Version</label><input type=text id=ver placeholder="e.g. 2.4.14" data-i18n-ph=ph_ver></div>
<div style="flex:2;min-width:180px"><label data-i18n=l_notes>Notes (optional)</label><input type=text id=notes></div>
<label style="display:flex;align-items:center;gap:.35rem;margin-top:1.2rem"><input type=checkbox id=mand> <span data-i18n=l_mand>mandatory</span></label>
</div>
<div class=row><button id=go data-i18n=btn_go>Upload &amp; publish</button>
<span class=sub id=hint data-i18n=hint>file is auto-published when upload completes</span></div>
<div class=bar id=bar><i id=bari></i></div>
<div class=msg id=msg></div>
</form><script>
var T={
 en:{title:'⚡ LAN App-Update Upload',
     sub:'Direct to the Pi over the local network — bypasses the Cloudflare 100\\u00a0MB limit. Big installers OK (≤300\\u00a0MB).',
     l_pw:'Admin password', l_file:'Update file',
     drop:'Drop file here, or click to choose',
     l_ver:'Version', ph_ver:'e.g. 2.4.14',
     l_notes:'Notes (optional)', l_mand:'mandatory',
     btn_go:'Upload & publish',
     hint:'file is auto-published when upload completes',
     need_pw:'Enter the admin password.',
     need_file:'Choose a file.',
     auth:'Authenticating…',
     login_fail:'Login failed (wrong password?).',
     login_err:'Login error: ',
     uploading:'Uploading…',
     processing:' — processing on server…',
     published:'✅ Published: ',
     failed:'❌ Failed: HTTP ',
     net_err:'❌ Network error during upload.',
     lang_btn:'日本語'},
 ja:{title:'⚡ LAN アプリ更新アップロード',
     sub:'ローカルネットワーク経由でPiへ直接アップロード — Cloudflareの100\\u00a0MB制限を回避します。大きなインストーラー（最大300\\u00a0MB）にも対応。',
     l_pw:'管理者パスワード', l_file:'更新ファイル',
     drop:'ここにファイルをドロップ、またはクリックして選択',
     l_ver:'バージョン', ph_ver:'例: 2.4.14',
     l_notes:'メモ（任意）', l_mand:'強制更新',
     btn_go:'アップロードして公開',
     hint:'アップロード完了後、自動で公開されます',
     need_pw:'管理者パスワードを入力してください。',
     need_file:'ファイルを選択してください。',
     auth:'認証中…',
     login_fail:'ログインに失敗しました（パスワードが違いますか？）',
     login_err:'ログインエラー: ',
     uploading:'アップロード中…',
     processing:' — サーバーで処理中…',
     published:'✅ 公開しました: ',
     failed:'❌ 失敗しました: HTTP ',
     net_err:'❌ アップロード中にネットワークエラーが発生しました。',
     lang_btn:'English'}
};
var LANG=(localStorage.getItem('lan_lang')||(navigator.language&&navigator.language.indexOf('ja')===0?'ja':'en'));
function t(k){return (T[LANG]&&T[LANG][k])||T.en[k]||k}
function applyLang(){
 document.documentElement.lang=LANG;
 document.querySelectorAll('[data-i18n]').forEach(function(el){
   var k=el.getAttribute('data-i18n'); var v=t(k);
   if(el.id==='drop'){ el.firstChild?(el.firstChild.nodeValue=v):(el.textContent=v); }
   else{ el.textContent=v; }
 });
 document.querySelectorAll('[data-i18n-ph]').forEach(function(el){
   el.setAttribute('placeholder', t(el.getAttribute('data-i18n-ph')));
 });
 document.getElementById('lang').textContent=t('lang_btn');
}
document.getElementById('lang').onclick=function(){LANG=(LANG==='en'?'ja':'en');localStorage.setItem('lan_lang',LANG);applyLang();};
applyLang();

var F=document.getElementById('file'),D=document.getElementById('drop'),M=document.getElementById('msg'),
B=document.getElementById('bar'),BI=document.getElementById('bari'),G=document.getElementById('go'),
FN=document.getElementById('fname');
D.onclick=function(){F.click()};
F.onchange=function(){FN.textContent=F.files[0]?(F.files[0].name+'  ('+(F.files[0].size/1048576).toFixed(1)+' MB)'):''};
['dragover','dragenter'].forEach(function(e){D.addEventListener(e,function(ev){ev.preventDefault();D.classList.add('hot')})});
['dragleave','drop'].forEach(function(e){D.addEventListener(e,function(ev){ev.preventDefault();D.classList.remove('hot')})});
D.addEventListener('drop',function(ev){if(ev.dataTransfer.files[0]){F.files=ev.dataTransfer.files;F.onchange()}});
function say(s,c){M.textContent=s;M.className='msg '+(c||'')}
G.onclick=async function(){
 var pw=document.getElementById('pw').value, f=F.files[0];
 if(!pw){say(t('need_pw'),'err');return}
 if(!f){say(t('need_file'),'err');return}
 G.disabled=true; say(t('auth'));
 try{
  var lr=await fetch('/admin/login',{method:'POST',credentials:'same-origin',
    headers:{'Content-Type':'application/x-www-form-urlencoded'},body:'password='+encodeURIComponent(pw)});
  if(!lr.ok){say(t('login_fail'),'err');G.disabled=false;return}
 }catch(e){say(t('login_err')+e,'err');G.disabled=false;return}
 var fd=new FormData();fd.append('file',f);fd.append('version',document.getElementById('ver').value.trim());
 fd.append('notes',document.getElementById('notes').value.trim());
 fd.append('mandatory',document.getElementById('mand').checked?'1':'0');
 B.style.display='block';BI.style.width='0';say(t('uploading'));
 var x=new XMLHttpRequest();x.open('POST','/admin/api/app-update',true);x.withCredentials=true;
 x.upload.onprogress=function(e){if(e.lengthComputable){var p=Math.round(e.loaded/e.total*100);
   BI.style.width=p+'%';say(t('uploading')+' '+p+'%'+(p>=100?t('processing'):''))}};
 x.onload=function(){G.disabled=false;
   if(x.status>=200&&x.status<300){var j={};try{j=JSON.parse(x.responseText)}catch(_){}
     var u=(j&&j.update)||{};BI.style.width='100%';
     say(t('published')+(u.filename||'')+'  v'+(u.version||'')+'\\nsha256 '+(u.sha256||'').slice(0,16)+'…  '+
       (((u.size||0)/1048576).toFixed(1))+' MB','ok')}
   else{say(t('failed')+x.status+'\\n'+x.responseText.slice(0,300),'err')}};
 x.onerror=function(){G.disabled=false;say(t('net_err'),'err')};
 x.send(fd);
};
</script></body></html>"""

@app.get("/lan-upload", response_class=HTMLResponse, include_in_schema=False)
async def lan_upload_page(request: Request):
    """LAN-only GUI to push a big app-update file straight to the Pi
    (bypasses the Cloudflare body-size cap). RFC1918/loopback clients only."""
    # Any Cloudflare signature ⇒ it arrived via the public tunnel ⇒ deny
    # (cloudflared reaches nginx from 127.0.0.1, so an IP check alone can't
    # distinguish tunnel traffic — these headers can).
    if request.headers.get("cf-connecting-ip") or request.headers.get("cf-ray") \
       or request.headers.get("cf-ipcountry"):
        return HTMLResponse(
            "<h2>LAN only</h2><p>This uploader is not available over the public "
            "internet. Open it on the Pi's LAN address, e.g. "
            "<code>http://192.168.0.2/lan-upload</code>.</p>",
            status_code=403,
        )
    ip = (request.headers.get("x-real-ip")
          or request.headers.get("x-forwarded-for", "").split(",")[0].strip()
          or (request.client.host if request.client else ""))
    if not _ip_is_lan(ip):
        return HTMLResponse(
            "<h2>LAN only</h2><p>This uploader is reachable only from the local "
            "network (you came from <code>%s</code>). Use it on the Pi's LAN "
            "address, e.g. <code>http://192.168.0.2/lan-upload</code>.</p>" % (ip or "?"),
            status_code=403,
        )
    return HTMLResponse(_LAN_UPLOAD_HTML)

@app.get("/admin/agent-spec")
async def admin_agent_spec(request: Request):
    """Serve the desktop-agent integration spec (Markdown) for admins."""
    _require_admin_session(request)
    if not DESKTOP_SPEC_PATH.exists():
        raise HTTPException(status_code=404, detail="Spec file not found.")
    return PlainTextResponse(
        DESKTOP_SPEC_PATH.read_text(encoding="utf-8"),
        media_type="text/markdown; charset=utf-8",
    )

@app.get("/admin/partner-spec")
async def admin_partner_spec(request: Request):
    """Serve the partner-facing API doc (Markdown) for admins."""
    _require_admin_session(request)
    if not PARTNER_SPEC_PATH.exists():
        raise HTTPException(status_code=404, detail="Spec file not found.")
    return PlainTextResponse(
        PARTNER_SPEC_PATH.read_text(encoding="utf-8"),
        media_type="text/markdown; charset=utf-8",
    )

@app.get("/admin/agentapi-spec")
async def admin_agentapi_spec(request: Request):
    """Serve the unified AI-Agent API doc (Markdown) for admins."""
    _require_admin_session(request)
    spec = BASE_DIR / "AGENT_API.md"
    if not spec.exists():
        raise HTTPException(status_code=404, detail="Spec file not found.")
    return PlainTextResponse(spec.read_text(encoding="utf-8"),
                             media_type="text/markdown; charset=utf-8")

@app.get("/admin/api/partner-status")
async def admin_partner_status(request: Request):
    _require_admin_session(request)
    return _partner_status()

@app.get("/admin/api/partner-log")
async def admin_partner_log(request: Request, limit: int = 100):
    _require_admin_session(request)
    return _partner_log_tail(limit)

@app.post("/admin/api/partner-enable")
async def admin_partner_enable(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    return _partner_set_enabled(bool(payload.get("enabled")))

@app.post("/admin/api/partner-add")
async def admin_partner_add(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    ips = payload.get("ip_allowlist") or []
    if isinstance(ips, str):
        ips = [s for s in ips.replace(",", " ").split() if s]
    return _partner_add(name, int(payload.get("rate_per_min") or 60), ips,
                        int(payload.get("monthly_quota") or 0))

@app.post("/admin/api/partner-update")
async def admin_partner_update(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    pid = (payload.get("id") or "").strip()
    if not pid:
        raise HTTPException(status_code=400, detail="id required")
    ips = payload.get("ip_allowlist")
    if isinstance(ips, str):
        ips = [s for s in ips.replace(",", " ").split() if s]
    return _partner_update(
        pid,
        disabled=payload.get("disabled"),
        rate_per_min=payload.get("rate_per_min"),
        ip_allowlist=ips,
        monthly_quota=payload.get("monthly_quota"),
    )

@app.post("/admin/api/partner-revoke")
async def admin_partner_revoke(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    pid = (payload.get("id") or "").strip()
    if not pid:
        raise HTTPException(status_code=400, detail="id required")
    return _partner_revoke(pid)


# ---- Agent API admin (token CRUD) — `agentapi-*` to avoid colliding with the
#      pre-existing desktop-agent `/admin/api/agent-status|agent-log` endpoints.
@app.get("/admin/api/agentapi-status")
async def admin_agentapi_status(request: Request):
    _require_admin_session(request)
    return _agent_status()

@app.get("/admin/api/agentapi-log")
async def admin_agentapi_log(request: Request, limit: int = 100):
    _require_admin_session(request)
    return _agent_log_tail(limit)

@app.post("/admin/api/agentapi-enable")
async def admin_agentapi_enable(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    return _agent_set_enabled(bool(payload.get("enabled")))

@app.post("/admin/api/agentapi-add")
async def admin_agentapi_add(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name required")
    ips = payload.get("ip_allowlist") or []
    if isinstance(ips, str):
        ips = [s for s in ips.replace(",", " ").split() if s]
    we = payload.get("write_enabled")
    return _agent_add(name, int(payload.get("rate_per_min") or 60), ips,
                      True if we is None else bool(we))

@app.post("/admin/api/agentapi-update")
async def admin_agentapi_update(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    aid = (payload.get("id") or "").strip()
    if not aid:
        raise HTTPException(status_code=400, detail="id required")
    ips = payload.get("ip_allowlist")
    if isinstance(ips, str):
        ips = [s for s in ips.replace(",", " ").split() if s]
    return _agent_update(aid, disabled=payload.get("disabled"),
                         rate_per_min=payload.get("rate_per_min"),
                         ip_allowlist=ips, write_enabled=payload.get("write_enabled"))

@app.post("/admin/api/agentapi-revoke")
async def admin_agentapi_revoke(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    aid = (payload.get("id") or "").strip()
    if not aid:
        raise HTTPException(status_code=400, detail="id required")
    return _agent_revoke(aid)

@app.get("/admin/api/agentapi-connect-pw")
async def admin_agentapi_connect_pw_get(request: Request):
    """The dedicated MCP-connector approval password (separate from the admin
    console password) + the connector URL to hand to ChatGPT/Claude."""
    _require_admin_session(request)
    return {"password": ADMIN_CONFIG.get("mcp_connect_password", ""),
            "mcp_url": "https://link.genbafms.com/mcp"}

@app.post("/admin/api/agentapi-connect-pw")
async def admin_agentapi_connect_pw_set(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    pw = str(payload.get("password") or "").strip()
    ADMIN_CONFIG["mcp_connect_password"] = pw
    _save_json(ADMIN_CONFIG_PATH, ADMIN_CONFIG)
    return {"ok": True, "set": bool(pw)}


# ---- Change-request management (Part B) -----------------------------------
@app.get("/admin/api/change-requests")
async def admin_change_requests(request: Request, status: str | None = None,
                                agent: str | None = None, limit: int = 100):
    _require_admin_session(request)
    items = _change_requests.list_requests(agent_id=agent, status=status, limit=limit)
    return {"count": len(items), "counts": _change_requests.status_counts(), "items": items}

@app.get("/admin/api/change-requests/{req_id}")
async def admin_change_request_detail(request: Request, req_id: int):
    _require_admin_session(request)
    row = _change_requests.get_request(req_id)
    if not row:
        raise HTTPException(status_code=404, detail="request not found")
    return row

def _cr_admin(fn, *args, **kwargs):
    try:
        return fn(*args, **kwargs)
    except _change_requests.ChangeRequestError as exc:
        raise HTTPException(status_code=exc.status, detail=exc.message)

@app.post("/admin/api/change-requests/{req_id}/evaluate")
async def admin_change_request_evaluate(request: Request, req_id: int, payload: dict = Body(...)):
    _require_admin_session(request)
    return _cr_admin(_change_requests.evaluate, req_id, payload.get("feasible"),
                     payload.get("risk_level"), payload.get("timeline_estimate", ""),
                     payload.get("notes", ""), "admin")

@app.post("/admin/api/change-requests/{req_id}/decision")
async def admin_change_request_decision(request: Request, req_id: int, payload: dict = Body(...)):
    _require_admin_session(request)
    return _cr_admin(_change_requests.decide, req_id, payload.get("decision", ""),
                     payload.get("notes", ""), "admin")

@app.post("/admin/api/change-requests/{req_id}/status")
async def admin_change_request_status(request: Request, req_id: int, payload: dict = Body(...)):
    _require_admin_session(request)
    return _cr_admin(_change_requests.set_status, req_id, payload.get("status", ""), "admin")

@app.post("/admin/api/change-requests/{req_id}/messages")
async def admin_change_request_message(request: Request, req_id: int, payload: dict = Body(...)):
    _require_admin_session(request)
    _cr_admin(_change_requests.add_message, req_id, "admin", payload.get("message", ""), "message", None, None)
    return _change_requests.get_request(req_id)

@app.get("/admin/api/api-status")
async def admin_api_status(request: Request, limit: int = 200):
    """API request/response status — every /api/* hit captured by the access
    middleware, plus rolled-up endpoint stats. Drives the Admin → API Status tab."""
    _require_admin_session(request)
    limit = max(1, min(limit, _RECENT_MAX))
    with _access_lock:
        snapshot = list(_recent_access)
    # Filter to API endpoints only.
    api_items = [it for it in snapshot if (it.get("path") or "").startswith("/api/")
                 or (it.get("path") or "").startswith("/admin/api/")]
    # Per-endpoint roll-up.
    by_path: dict[str, dict] = {}
    for it in api_items:
        path = it.get("path") or "?"
        st   = int(it.get("status") or 0)
        dur  = int(it.get("duration_ms") or 0)
        row  = by_path.setdefault(path, {
            "path": path, "count": 0, "errors_4xx": 0, "errors_5xx": 0,
            "total_ms": 0, "max_ms": 0, "last_status": None, "last_ts": None,
        })
        row["count"]     += 1
        row["total_ms"]  += dur
        row["max_ms"]     = max(row["max_ms"], dur)
        row["last_status"] = st
        row["last_ts"]    = it.get("ts")
        if 400 <= st < 500: row["errors_4xx"] += 1
        if st >= 500:       row["errors_5xx"] += 1
    endpoints = []
    for row in by_path.values():
        row["avg_ms"] = round(row["total_ms"] / row["count"]) if row["count"] else 0
        endpoints.append(row)
    endpoints.sort(key=lambda r: r["count"], reverse=True)
    # Summary
    total = len(api_items)
    err4  = sum(1 for it in api_items if 400 <= int(it.get("status") or 0) < 500)
    err5  = sum(1 for it in api_items if int(it.get("status") or 0) >= 500)
    durs  = [int(it.get("duration_ms") or 0) for it in api_items]
    durs_sorted = sorted(durs)
    p50 = durs_sorted[len(durs_sorted)//2] if durs_sorted else 0
    p95 = durs_sorted[int(len(durs_sorted)*0.95)] if durs_sorted else 0
    avg = round(sum(durs)/len(durs)) if durs else 0
    return {
        "summary": {
            "total": total, "errors_4xx": err4, "errors_5xx": err5,
            "avg_ms": avg, "p50_ms": p50, "p95_ms": p95,
            "endpoint_count": len(endpoints),
            "buffer_size": _RECENT_MAX,
            "buffer_used": len(snapshot),
        },
        "endpoints": endpoints,
        "recent": list(reversed(api_items[-limit:])),
    }

@app.get("/admin/api/alerts")
async def admin_alerts(request: Request):
    _require_admin_session(request)
    return {"count": len(NEW_LOGIN_ALERTS), "items": list(reversed(NEW_LOGIN_ALERTS))}

@app.post("/admin/api/alerts/clear")
async def admin_alerts_clear(request: Request):
    _require_admin_session(request)
    NEW_LOGIN_ALERTS.clear()
    return {"ok": True}

@app.get("/admin/api/api-keys")
async def admin_api_keys_get(request: Request):
    _require_admin_session(request)
    # Return masked values — full values only via the on-disk file (root-readable).
    masked = {}
    for k, v in API_KEYS.items():
        if k.startswith("_"):
            continue
        if isinstance(v, str) and len(v) >= 8:
            masked[k] = v[:4] + "…" + v[-4:]
        else:
            masked[k] = "(unset)"
    return {"keys": masked, "file": str(API_KEYS_PATH)}


# ============================================================================
# Announcement banner — public GET, admin POST. Drives the BETA-style banner
# on /attendance/console (and any other page that subscribes).
# ============================================================================
_ANNOUNCEMENT_ALLOWED_COLORS = {"amber", "red", "blue", "green", "gray", "purple"}

@app.get("/api/announcement")
async def announcement_get():
    cfg = _load_json(ANNOUNCEMENT_PATH, {}) or {}
    return cfg

@app.post("/admin/api/announcement")
async def announcement_set(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    current = _load_json(ANNOUNCEMENT_PATH, {}) or {}
    color = (payload.get("color") or current.get("color") or "amber").lower()
    if color not in _ANNOUNCEMENT_ALLOWED_COLORS:
        raise HTTPException(status_code=400, detail=f"color must be one of {sorted(_ANNOUNCEMENT_ALLOWED_COLORS)}")
    new_cfg = {
        "enabled":     bool(payload.get("enabled", current.get("enabled", True))),
        "type":        (payload.get("type") or current.get("type") or "info").lower(),
        "color":       color,
        "badge":       (payload.get("badge") or current.get("badge") or "INFO").upper()[:24],
        "text_en":     (payload.get("text_en") or "")[:600],
        "text_ja":     (payload.get("text_ja") or "")[:600],
        "ends_at":     (payload.get("ends_at") or current.get("ends_at") or "") or None,
        "dismissible": bool(payload.get("dismissible", current.get("dismissible", True))),
        "updated_at":  datetime.now().isoformat(timespec="seconds"),
        "presets":     current.get("presets", {}),
    }
    _save_json(ANNOUNCEMENT_PATH, new_cfg)
    return {"ok": True, "announcement": new_cfg}


# ============================================================================
# Process list — top tasks by CPU and MEM (for admin Overview)
# ============================================================================
@app.get("/admin/api/processes")
async def admin_processes(request: Request, top: int = 10):
    _require_admin_session(request)
    if psutil is None:
        return {"available": False, "by_cpu": [], "by_mem": [], "self": None}
    top = max(1, min(top, 50))
    procs = []
    # Prime CPU% so the second read returns a real value.
    for p in psutil.process_iter(attrs=["pid", "name", "username"]):
        try:
            p.cpu_percent(interval=None)
        except Exception:
            pass
    time.sleep(0.25)
    for p in psutil.process_iter(attrs=["pid", "name", "username", "create_time"]):
        try:
            cpu = p.cpu_percent(interval=None)
            mem = p.memory_info().rss
            procs.append({
                "pid":     p.info["pid"],
                "name":    (p.info["name"] or "")[:40],
                "user":    (p.info["username"] or "")[:20],
                "cpu":     round(cpu, 1),
                "mem_mb":  round(mem / (1024 * 1024), 1),
                "started": datetime.fromtimestamp(p.info["create_time"]).isoformat(timespec="seconds"),
            })
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    by_cpu = sorted(procs, key=lambda x: x["cpu"], reverse=True)[:top]
    by_mem = sorted(procs, key=lambda x: x["mem_mb"], reverse=True)[:top]
    self_proc = None
    try:
        sp = psutil.Process(os.getpid())
        self_proc = {
            "pid":    sp.pid,
            "name":   sp.name(),
            "cpu":    round(sp.cpu_percent(interval=None), 1),
            "mem_mb": round(sp.memory_info().rss / (1024 * 1024), 1),
            "threads": sp.num_threads(),
            "started": datetime.fromtimestamp(sp.create_time()).isoformat(timespec="seconds"),
        }
    except Exception:
        pass
    return {
        "available": True,
        "total_processes": len(procs),
        "by_cpu": by_cpu,
        "by_mem": by_mem,
        "self":   self_proc,
    }


# ============================================================================
# Visitors / IP labels — group access log by IP, with editable labels
# stored in known_ips.json so admin can mark "Office", "Home", etc.
# ============================================================================
def _load_ip_labels() -> dict:
    raw = _load_json(KNOWN_IPS_PATH, {}) or {}
    return raw if isinstance(raw, dict) else {}

def _save_ip_labels(d: dict) -> None:
    _save_json(KNOWN_IPS_PATH, d)

@app.get("/admin/api/ip-labels")
async def admin_ip_labels(request: Request):
    _require_admin_session(request)
    return {"items": _load_ip_labels()}

@app.post("/admin/api/ip-labels")
async def admin_ip_labels_set(request: Request, payload: dict = Body(...)):
    _require_admin_session(request)
    ip = (payload.get("ip") or "").strip()
    if not ip:
        raise HTTPException(status_code=400, detail="ip is required")
    labels = _load_ip_labels()
    if payload.get("delete"):
        labels.pop(ip, None)
    else:
        labels[ip] = {
            "label": (payload.get("label") or "")[:60] or "Unnamed",
            "owner": (payload.get("owner") or "")[:60],
            "color": (payload.get("color") or "blue")[:20],
        }
    _save_ip_labels(labels)
    return {"ok": True, "items": labels}

@app.get("/admin/api/visitors")
async def admin_visitors(request: Request):
    _require_admin_session(request)
    labels = _load_ip_labels()
    with _access_lock:
        items = list(_recent_access)
    by_ip: dict[str, dict] = {}
    for it in items:
        ip = it.get("ip") or "?"
        rec = by_ip.setdefault(ip, {
            "ip": ip,
            "label": labels.get(ip, {}).get("label"),
            "owner": labels.get(ip, {}).get("owner"),
            "color": labels.get(ip, {}).get("color"),
            "first_seen": it["ts"],
            "last_seen":  it["ts"],
            "hits": 0,
            "errors_4xx": 0,
            "errors_5xx": 0,
            "devices": set(),
            "paths":   set(),
        })
        rec["last_seen"] = it["ts"]
        if it["ts"] < rec["first_seen"]:
            rec["first_seen"] = it["ts"]
        rec["hits"] += 1
        st = it.get("status") or 0
        if 400 <= st < 500: rec["errors_4xx"] += 1
        elif st >= 500:     rec["errors_5xx"] += 1
        rec["devices"].add(it.get("device") or "")
        rec["paths"].add(it.get("path") or "")
    out = []
    for rec in by_ip.values():
        rec["devices"] = sorted(d for d in rec["devices"] if d)[:5]
        rec["paths_unique"] = len(rec["paths"])
        rec.pop("paths", None)
        out.append(rec)
    out.sort(key=lambda r: r["last_seen"], reverse=True)
    return {"count": len(out), "items": out}


# ============================================================================
# Security signals — bot/abuse hints from the in-memory access ring.
# Cheap heuristics: hits/min, 4xx-heavy IPs, suspicious paths.
# ============================================================================
_SUSPICIOUS_PATH_HINTS = ("/.env", "/wp-", "/.git", "/phpmyadmin", "/xmlrpc", "/etc/passwd", "/.aws", "/admin.php")

@app.get("/admin/api/security")
async def admin_security(request: Request):
    _require_admin_session(request)
    with _access_lock:
        items = list(_recent_access)
    if not items:
        return {"window_size": 0, "items": []}
    now = datetime.now()
    def _age(it) -> float:
        try:
            return (now - datetime.fromisoformat(it["ts"])).total_seconds()
        except Exception:
            return 1e9
    last_5m  = [it for it in items if _age(it) <= 300]
    last_15m = [it for it in items if _age(it) <= 900]
    last_60m = [it for it in items if _age(it) <= 3600]
    by_ip_5m: dict[str, int] = {}
    err_by_ip: dict[str, int] = {}
    for it in last_5m:
        by_ip_5m[it["ip"]] = by_ip_5m.get(it["ip"], 0) + 1
    for it in last_60m:
        if (it.get("status") or 0) >= 400:
            err_by_ip[it["ip"]] = err_by_ip.get(it["ip"], 0) + 1
    suspicious = []
    for it in items:
        p = (it.get("path") or "").lower()
        if any(h in p for h in _SUSPICIOUS_PATH_HINTS):
            suspicious.append({"ts": it["ts"], "ip": it["ip"], "path": it["path"], "status": it["status"]})
    return {
        "window_size": len(items),
        "hits_5m":  len(last_5m),
        "hits_15m": len(last_15m),
        "hits_60m": len(last_60m),
        "top_ips_5m":  sorted(([{"ip": k, "hits": v} for k, v in by_ip_5m.items()]), key=lambda x: x["hits"], reverse=True)[:10],
        "top_4xx_60m": sorted(([{"ip": k, "errors": v} for k, v in err_by_ip.items()]), key=lambda x: x["errors"], reverse=True)[:10],
        "suspicious_paths": suspicious[-30:],
    }


@app.post("/api/convert")
async def convert_pdf(file: UploadFile = File(...)):
    """
    Convert PDF to Excel
    Returns Excel file for download
    """
    try:
        original_name, _file_size, records, metadata = await parse_uploaded_pdf(file)
        record_date = metadata.get("record_date")
        pack_count = metadata.get("number_of_packs")

        # Check mismatches against any rows already saved for this date BEFORE we overwrite them.
        mismatches = find_attendance_mismatches(record_date, records)

        excel_filename, _excel_path, records_processed, batch_id = export_records_to_excel(
            original_name,
            records,
            record_date=record_date,
        )

        if record_date is not None and pack_count is not None:
            upsert_pack_count_for_date(record_date, pack_count, source=original_name)

        return {
            "status": "success",
            "filename": excel_filename,
            "records_processed": records_processed,
            "roster_records": len(records),
            "batch_id": batch_id,
            "message": "PDFが正常に変換され、Excelファイルが生成されました",
            "download_url": build_download_url(excel_filename),
            "record_date": record_date.isoformat() if record_date else None,
            "number_of_packs": pack_count,
            "mismatches": mismatches,
            "mismatch_count": len(mismatches),
            "timestamp": datetime.now().isoformat()
        }
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error while converting PDF: {exc}") from exc

@app.post("/api/convert-multiple")
async def convert_multiple_pdfs(files: list[UploadFile] = File(...)):
    """Convert multiple PDFs and return one zip bundle of Excel files."""
    if not files:
        raise HTTPException(status_code=400, detail="Please upload at least one PDF file.")

    try:
        exports: list[tuple[str, Path, int]] = []
        per_file_results: list[dict[str, object]] = []
        total_records_processed = 0
        total_mismatch_count = 0

        for file in files:
            original_name, _file_size, records, metadata = await parse_uploaded_pdf(file)
            record_date = metadata.get("record_date")
            pack_count = metadata.get("number_of_packs")

            mismatches = find_attendance_mismatches(record_date, records)
            total_mismatch_count += len(mismatches)

            excel_filename, excel_path, records_processed, _batch_id = export_records_to_excel(
                original_name,
                records,
                record_date=record_date,
            )

            if record_date is not None and pack_count is not None:
                upsert_pack_count_for_date(record_date, pack_count, source=original_name)

            exports.append((excel_filename, excel_path, len(records)))
            total_records_processed += records_processed
            per_file_results.append({
                "source_filename": original_name,
                "export_filename": excel_filename,
                "record_date": record_date.isoformat() if record_date else None,
                "number_of_packs": pack_count,
                "mismatches": mismatches,
                "mismatch_count": len(mismatches),
            })

        bundle_filename = create_bundle_filename()
        bundle_path = EXPORT_DIR / bundle_filename
        with zipfile.ZipFile(bundle_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for excel_filename, excel_path, _row_count in exports:
                archive.write(excel_path, arcname=excel_filename)

        return {
            "status": "success",
            "filename": bundle_filename,
            "file_count": len(exports),
            "records_processed": total_records_processed,
            "message": f"{len(exports)} PDF files were converted and bundled into one ZIP file.",
            "download_url": build_download_url(bundle_filename),
            "files": per_file_results,
            "mismatch_count": total_mismatch_count,
            "timestamp": datetime.now().isoformat(),
        }
    except HTTPException:
        raise
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error while converting PDFs: {exc}") from exc

def _process_batch_job(job_id: str) -> None:
    """Worker executed in a BackgroundTasks thread — converts each staged file one at a time."""
    job_dir = BATCH_UPLOAD_DIR / job_id
    try:
        with BATCH_JOBS_LOCK:
            job = BATCH_JOBS.get(job_id)
            if job is None:
                return
            files = list(job.get("files") or [])
            job["status"] = "running"
            job["started_at"] = datetime.now().isoformat()
            job["started_at_ts"] = time.time()

        exports: list[tuple[str, Path]] = []
        total_records = 0
        total_mismatch_count = 0

        for index, entry in enumerate(files):
            _wait_for_resources(job_id)

            _batch_update_file(job_id, index, status="processing")
            _batch_set(job_id, current_index=index, current_file=entry["source_filename"])

            stored_path = Path(entry["stored_path"])
            try:
                records = apply_employee_roster(parse_pdf_data(stored_path))
                metadata = extract_pdf_metadata(stored_path)
                record_date = metadata.get("record_date")
                pack_count = metadata.get("number_of_packs")

                mismatches = find_attendance_mismatches(record_date, records)
                total_mismatch_count += len(mismatches)

                excel_filename, excel_path, records_processed, _batch_id = export_records_to_excel(
                    entry["source_filename"],
                    records,
                    record_date=record_date,
                )

                if record_date is not None and pack_count is not None:
                    upsert_pack_count_for_date(record_date, pack_count, source=entry["source_filename"])

                exports.append((excel_filename, excel_path))
                total_records += records_processed

                _batch_update_file(
                    job_id,
                    index,
                    status="done",
                    export_filename=excel_filename,
                    records_processed=records_processed,
                    record_date=record_date.isoformat() if record_date else None,
                    number_of_packs=pack_count,
                    mismatch_count=len(mismatches),
                    mismatches=mismatches,
                )
            except ValueError as exc:
                _batch_update_file(job_id, index, status="error", error=str(exc))
            except Exception as exc:  # noqa: BLE001
                _batch_update_file(job_id, index, status="error", error=f"unexpected error: {exc}")

            with BATCH_JOBS_LOCK:
                job = BATCH_JOBS.get(job_id)
                if job is not None:
                    job["processed"] = index + 1
                    job["total_mismatch_count"] = total_mismatch_count
                    job["total_records"] = total_records

        # Bundle successful exports into a single zip.
        successful_exports = [pair for pair in exports if pair[1].exists()]
        bundle_filename = None
        if successful_exports:
            bundle_filename = create_bundle_filename()
            bundle_path = EXPORT_DIR / bundle_filename
            with zipfile.ZipFile(bundle_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for excel_filename, excel_path in successful_exports:
                    archive.write(excel_path, arcname=excel_filename)

        finished_at = datetime.now().isoformat()
        with BATCH_JOBS_LOCK:
            job = BATCH_JOBS.get(job_id)
            if job is not None:
                error_count = sum(1 for f in job.get("files") or [] if f.get("status") == "error")
                job["status"] = "done" if error_count == 0 else "done_with_errors"
                job["bundle_filename"] = bundle_filename
                job["download_url"] = build_download_url(bundle_filename) if bundle_filename else None
                job["finished_at"] = finished_at
                job["finished_at_ts"] = time.time()
                job["current_file"] = None
                job["total_mismatch_count"] = total_mismatch_count
                job["total_records"] = total_records
                job["error_count"] = error_count
    except Exception as exc:  # noqa: BLE001
        with BATCH_JOBS_LOCK:
            job = BATCH_JOBS.get(job_id)
            if job is not None:
                job["status"] = "error"
                job["error"] = f"worker crashed: {exc}"
                job["finished_at"] = datetime.now().isoformat()
                job["finished_at_ts"] = time.time()
    finally:
        # Remove staged PDF inputs — the exports live under EXPORT_DIR and are served separately.
        if job_dir.exists():
            shutil.rmtree(job_dir, ignore_errors=True)

@app.post("/api/convert-batch")
async def convert_batch(background_tasks: BackgroundTasks, files: list[UploadFile] = File(...)):
    """Accept a batch of PDFs, stage them to disk, and kick off a background conversion job."""
    _batch_cleanup_expired()

    if not files:
        raise HTTPException(status_code=400, detail="Please upload at least one PDF file.")

    job_id = uuid4().hex
    job_dir = BATCH_UPLOAD_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    staged: list[dict] = []
    total_bytes = 0
    try:
        for file in files:
            original_name = Path(file.filename or "").name
            if not original_name:
                raise HTTPException(status_code=400, detail="One of the uploaded files is missing a filename.")
            if Path(original_name).suffix.lower() != ".pdf":
                raise HTTPException(status_code=400, detail=f"Only PDF files are supported (got '{original_name}').")

            content = await file.read()
            if not content:
                raise HTTPException(status_code=400, detail=f"'{original_name}' is empty.")
            if b"%PDF" not in content[:1024]:
                raise HTTPException(status_code=400, detail=f"'{original_name}' is not a valid PDF.")

            total_bytes += len(content)
            if total_bytes > BATCH_MAX_TOTAL_BYTES:
                raise HTTPException(
                    status_code=413,
                    detail=(
                        f"Batch exceeds the {BATCH_MAX_TOTAL_BYTES // (1024 * 1024)} MB limit "
                        f"(reached {total_bytes // (1024 * 1024)} MB while adding '{original_name}')."
                    ),
                )

            stored_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid4().hex}.pdf"
            stored_path = job_dir / stored_name
            stored_path.write_bytes(content)

            staged.append({
                "index": len(staged),
                "source_filename": original_name,
                "stored_path": str(stored_path),
                "size_bytes": len(content),
                "status": "queued",
                "export_filename": None,
                "records_processed": 0,
                "record_date": None,
                "number_of_packs": None,
                "mismatch_count": 0,
                "mismatches": [],
                "error": None,
            })
    except HTTPException:
        shutil.rmtree(job_dir, ignore_errors=True)
        raise
    except Exception as exc:  # noqa: BLE001
        shutil.rmtree(job_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"Unexpected error while staging uploads: {exc}") from exc

    created_at = datetime.now().isoformat()
    with BATCH_JOBS_LOCK:
        BATCH_JOBS[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "files": staged,
            "total_files": len(staged),
            "processed": 0,
            "total_bytes": total_bytes,
            "total_records": 0,
            "total_mismatch_count": 0,
            "error_count": 0,
            "current_index": None,
            "current_file": None,
            "bundle_filename": None,
            "download_url": None,
            "created_at": created_at,
            "created_at_ts": time.time(),
            "started_at": None,
            "started_at_ts": None,
            "finished_at": None,
            "finished_at_ts": None,
            "throttled": False,
            "cpu_percent": None,
            "ram_percent": None,
            "ram_available_mb": None,
            "error": None,
        }

    background_tasks.add_task(_process_batch_job, job_id)

    return {
        "job_id": job_id,
        "status": "queued",
        "total_files": len(staged),
        "total_bytes": total_bytes,
        "max_total_bytes": BATCH_MAX_TOTAL_BYTES,
        "created_at": created_at,
    }

@app.get("/api/convert-batch/{job_id}/status")
async def convert_batch_status(job_id: str):
    """Return a live snapshot of a batch job for the frontend progress panel."""
    snapshot = _batch_snapshot(job_id)
    if snapshot is None:
        raise HTTPException(status_code=404, detail="Batch job not found (it may have expired).")

    # Refresh live CPU/RAM on read so the UI polling gets current values.
    load = _sample_system_load()
    snapshot["cpu_percent"] = load["cpu_percent"]
    snapshot["ram_percent"] = load["ram_percent"]
    snapshot["ram_available_mb"] = load["ram_available_mb"]

    started_ts = snapshot.get("started_at_ts")
    finished_ts = snapshot.get("finished_at_ts")
    if started_ts is not None:
        end_ts = finished_ts if finished_ts is not None else time.time()
        snapshot["elapsed_seconds"] = round(end_ts - started_ts, 2)
    else:
        snapshot["elapsed_seconds"] = 0.0

    return snapshot

@app.get("/api/system/load")
async def system_load():
    """Lightweight system load snapshot used by the dashboard/status pills."""
    return _sample_system_load()

@app.get("/api/download/{filename}")
async def download_export(filename: str):
    """
    Download a generated Excel file or ZIP bundle.
    """
    file_path = get_export_file_path(filename)
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    media_type = (
        "application/zip"
        if file_path.suffix.lower() == ".zip"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return FileResponse(
        file_path,
        media_type=media_type,
        filename=file_path.name
    )

@app.get("/api/dashboard/summary")
async def dashboard_summary(month: str | None = None):
    """
    Get dashboard summary statistics
    """
    resolved_month = resolve_month(month)
    rows = list_attendance_rows(resolved_month)
    active_rows = [row for row in rows if row["has_data"]]
    working_minutes = [
        minutes
        for minutes in (parse_working_hours_minutes(row["working_hours"]) for row in active_rows)
        if minutes is not None
    ]

    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT COUNT(*) AS batch_count, MAX(upload_date) AS latest_upload
                FROM attendance_records
                WHERE month_year = %s
                """,
                (resolved_month,),
            )
            batch_row = cursor.fetchone()

    year, month_value = resolved_month.split("-")
    return {
        "month": resolved_month,
        "total_employees": len({row["employee_code"] for row in active_rows}),
        "total_records": len(active_rows),
        "average_working_hours": format_average_hours(working_minutes),
        "days_in_month": monthrange(int(year), int(month_value))[1],
        "upload_batches": batch_row["batch_count"] if batch_row else 0,
        "latest_upload": batch_row["latest_upload"] if batch_row else None,
    }

@app.get("/api/dashboard/employees")
async def dashboard_employees(month: str | None = None):
    """
    Get all employees with summary
    """
    resolved_month = resolve_month(month)
    rows = list_attendance_rows(resolved_month)
    by_employee: dict[str, dict[str, object]] = {}

    for row in rows:
        if not row["has_data"]:
            continue

        code = row["employee_code"]
        summary = by_employee.setdefault(code, {
            "code": code,
            "name": row["full_name"],
            "days_worked": 0,
            "working_minutes": [],
        })
        summary["days_worked"] = int(summary["days_worked"]) + 1
        minutes = parse_working_hours_minutes(row["working_hours"])
        if minutes is not None:
            summary["working_minutes"].append(minutes)

    employees = []
    for code in [employee["employee_code"] for employee in EMPLOYEE_ROSTER]:
        summary = by_employee.get(code)
        if not summary:
            continue
        employees.append({
            "id": code,
            "code": code,
            "name": summary["name"],
            "days_worked": summary["days_worked"],
            "avg_working_hours": format_average_hours(summary["working_minutes"]),
        })

    return {
        "month": resolved_month,
        "employees": employees,
        "total_count": len(employees),
    }

@app.get("/api/dashboard/employee/{employee_code}")
async def dashboard_employee_detail(employee_code: str, month: str | None = None):
    """
    Get individual employee details
    """
    resolved_month = resolve_month(month)
    rows = [
        row for row in list_attendance_rows(resolved_month)
        if row["employee_code"] == employee_code
    ]
    working_minutes = [
        minutes
        for minutes in (parse_working_hours_minutes(row["working_hours"]) for row in rows)
        if minutes is not None
    ]
    late_count = sum(
        1
        for row in rows
        if row["has_data"] and (time_to_minutes(row["commute_time"]) or 0) > 9 * 60
    )
    early_count = sum(
        1
        for row in rows
        if row["has_data"] and row["leave_time"] and (time_to_minutes(row["leave_time"]) or 0) < 17 * 60
    )

    return {
        "employee_id": employee_code,
        "name": EMPLOYEE_ROSTER_BY_ID.get(employee_code, ""),
        "month": resolved_month,
        "records": [
            {
                "source_filename": row["source_filename"],
                "converted_at": row["converted_at"],
                "commute_time": row["commute_time"],
                "leave_time": row["leave_time"],
                "working_hours": row["working_hours"],
                "has_data": bool(row["has_data"]),
            }
            for row in rows
        ],
        "monthly_data": {
            "total_hours": round(sum(working_minutes) / 60, 2) if working_minutes else 0,
            "average_daily": format_average_hours(working_minutes),
            "late_count": late_count,
            "early_count": early_count,
        }
    }

@app.get("/api/daily-packs/{record_date}")
async def get_daily_packs(record_date: str):
    """Return the pack count saved for a single date, or 0 if none recorded yet."""
    try:
        parsed = datetime.strptime(record_date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="record_date must be YYYY-MM-DD") from exc

    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT number_of_packs, note, updated_at FROM daily_packs WHERE record_date = %s",
                (parsed,),
            )
            row = cursor.fetchone()

    if row is None:
        return {"record_date": record_date, "number_of_packs": 0, "note": None, "exists": False}
    return {
        "record_date": record_date,
        "number_of_packs": row[0],
        "note": row[1],
        "updated_at": row[2].isoformat() if row[2] else None,
        "exists": True,
    }


def fetch_existing_daily_pack(record_date) -> dict | None:
    """Return the current daily_packs row for a date, or None if absent."""
    if record_date is None:
        return None
    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                "SELECT number_of_packs, note, updated_at FROM daily_packs WHERE record_date = %s",
                (record_date,),
            )
            row = cursor.fetchone()
    if not row:
        return None
    return {
        "number_of_packs": row["number_of_packs"],
        "note": row["note"],
        "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None,
    }


def fetch_existing_temp_staff(record_date) -> list[dict]:
    """Return saved temp_staff rows for a date (used to detect overwrite cases)."""
    if record_date is None:
        return []
    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT id, company, headcount, start_time, leave_time,
                       leave_next_day, hours_per_person, total_hours
                FROM temp_staff
                WHERE record_date = %s
                ORDER BY id
                """,
                (record_date,),
            )
            rows = cursor.fetchall()
    out: list[dict] = []
    for row in rows:
        out.append({
            "id": row["id"],
            "company": row["company"],
            "headcount": row["headcount"],
            "start_time": str(row["start_time"])[:5] if row["start_time"] else None,
            "leave_time": str(row["leave_time"])[:5] if row["leave_time"] else None,
            "leave_next_day": bool(row["leave_next_day"]),
            "hours_per_person": float(row["hours_per_person"]) if row["hours_per_person"] is not None else 0.0,
            "total_hours": float(row["total_hours"]) if row["total_hours"] is not None else 0.0,
        })
    return out


def shift_to_prod_date(shift_date):
    """Daily Packs convention: the PDF prints the shift date (製造分 day),
    but the pack count + フルキャスト hours belong to prod_date = shift + 1.

    ### CONFIG: change the +1 offset here if the business rule ever changes.
    # e.g. to save under the SAME date as the PDF, replace the return line with:
    #     return shift_date
    # e.g. to save two days later, use timedelta(days=2), etc.
    """
    if shift_date is None:
        return None
    return shift_date + timedelta(days=1)


@app.post("/api/daily-packs/auto-extract")
async def auto_extract_daily_packs():
    """Pick the latest PDF in the daily-packs watched folder and run the same
    extraction that /api/daily-packs/extract-pdf-multi runs on uploaded files.
    Returns one `results[]` entry so the frontend can reuse the same render path.
    """
    folder, single_file = _resolve_watched_target(DAILY_PACKS_AUTO_UPLOAD_DIR)
    if folder is None:
        path_str = str(DAILY_PACKS_AUTO_UPLOAD_DIR)
        if _looks_like_windows_path(path_str):
            raise HTTPException(status_code=404, detail=f"Configured path is a Windows-only path the Pi cannot read: {path_str}")
        raise HTTPException(status_code=404, detail=f"Watched path not reachable: {path_str}")
    picked = single_file if single_file is not None else _pick_latest_pdf_in(folder)
    if picked is None:
        raise HTTPException(status_code=404, detail=f"No PDF files found in {folder}.")

    entry: dict = {
        "source_filename": picked.name,
        "record_date": None,
        "shift_date": None,
        "number_of_packs": None,
        "found_date": False,
        "found_packs": False,
        "compared_count": 0,
        "mismatch_count": 0,
        "mismatches": [],
        "parse_error": None,
        "error": None,
        "fullcast_rows": [],
        "existing_pack": None,
        "existing_fullcast": [],
    }
    try:
        metadata = extract_pdf_metadata(picked)
        shift_date = metadata.get("record_date")
        record_date = shift_to_prod_date(shift_date)
        pack_count = metadata.get("number_of_packs")
        fullcast_rows = metadata.get("fullcast_rows") or []
        entry["shift_date"] = shift_date.isoformat() if shift_date else None
        entry["record_date"] = record_date.isoformat() if record_date else None
        entry["number_of_packs"] = pack_count
        entry["found_date"] = record_date is not None
        entry["found_packs"] = pack_count is not None
        entry["fullcast_rows"] = fullcast_rows
        entry["existing_pack"] = fetch_existing_daily_pack(record_date) if record_date else None
        entry["existing_fullcast"] = fetch_existing_temp_staff(record_date) if record_date else []
    except Exception as exc:
        entry["error"] = str(exc)

    return {
        "picked_filename": picked.name,
        "picked_size": picked.stat().st_size,
        "results": [entry],
        "found_count": 1 if entry["found_date"] else 0,
        "total_mismatch_count": 0,
    }


@app.post("/api/daily-packs/extract-pdf")
async def extract_daily_packs_from_pdf(file: UploadFile = File(...)):
    """Extract date, pack count, AND time mismatches from an uploaded PDF without saving.

    The frontend uses this to pre-fill the Daily packs form so the user can review
    the values and decide whether to save them. It also surfaces any time mismatches
    between the PDF rows and the existing database entries for the same date.
    """
    _original_name, file_path, _content = await save_uploaded_pdf(file)
    metadata = extract_pdf_metadata(file_path)
    shift_date = metadata.get("record_date")
    record_date = shift_to_prod_date(shift_date)
    pack_count = metadata.get("number_of_packs")
    fullcast_rows = metadata.get("fullcast_rows") or []

    # Tab 3 only needs date + pack count + フルキャスト rows + existing-data check.
    # Skip the expensive attendance table parse and mismatch scan to keep the
    # request well under Cloudflare's 100s gateway timeout on bulk uploads.
    existing_pack = fetch_existing_daily_pack(record_date)
    existing_fullcast = fetch_existing_temp_staff(record_date)

    return {
        "record_date": record_date.isoformat() if record_date else None,
        "shift_date": shift_date.isoformat() if shift_date else None,
        "number_of_packs": pack_count,
        "found_date": record_date is not None,
        "found_packs": pack_count is not None,
        "compared_count": 0,
        "mismatch_count": 0,
        "mismatches": [],
        "parse_error": None,
        "fullcast_rows": fullcast_rows,
        "existing_pack": existing_pack,
        "existing_fullcast": existing_fullcast,
    }


@app.post("/api/daily-packs/extract-pdf-multi")
async def extract_daily_packs_from_pdfs(files: list[UploadFile] = File(...)):
    """Extract date + pack count + attendance mismatches from several PDFs in one call.

    Mirrors /api/daily-packs/extract-pdf but runs over a batch. Each entry in the
    returned list is self-describing so the frontend can render per-file state and
    let the user bulk-save all successful extractions at once.
    """
    if not files:
        raise HTTPException(status_code=400, detail="Please upload at least one PDF file.")

    results: list[dict] = []
    found_count = 0
    total_mismatch_count = 0

    for file in files:
        original_name = Path(file.filename or "").name or "unnamed.pdf"
        entry: dict = {
            "source_filename": original_name,
            "record_date": None,
            "shift_date": None,
            "number_of_packs": None,
            "found_date": False,
            "found_packs": False,
            "compared_count": 0,
            "mismatch_count": 0,
            "mismatches": [],
            "parse_error": None,
            "error": None,
            "fullcast_rows": [],
            "existing_pack": None,
            "existing_fullcast": [],
        }
        try:
            _original, file_path, _content = await save_uploaded_pdf(file)
            metadata = extract_pdf_metadata(file_path)
            shift_date = metadata.get("record_date")
            record_date = shift_to_prod_date(shift_date)
            pack_count = metadata.get("number_of_packs")
            fullcast_rows = metadata.get("fullcast_rows") or []

            # Skip heavy attendance parse (not shown in Tab 3) to stay under the
            # Cloudflare gateway timeout on bulk uploads.
            entry.update({
                "record_date": record_date.isoformat() if record_date else None,
                "shift_date": shift_date.isoformat() if shift_date else None,
                "number_of_packs": pack_count,
                "found_date": record_date is not None,
                "found_packs": pack_count is not None,
                "compared_count": 0,
                "mismatch_count": 0,
                "mismatches": [],
                "fullcast_rows": fullcast_rows,
                "existing_pack": fetch_existing_daily_pack(record_date),
                "existing_fullcast": fetch_existing_temp_staff(record_date),
            })
            if entry["found_date"] and entry["found_packs"]:
                found_count += 1
        except HTTPException as exc:
            entry["error"] = exc.detail if isinstance(exc.detail, str) else "validation failed"
        except Exception as exc:  # noqa: BLE001
            entry["error"] = f"unexpected error: {exc}"

        results.append(entry)

    return {
        "total_files": len(files),
        "extractable_count": found_count,
        "total_mismatch_count": total_mismatch_count,
        "results": results,
    }


@app.post("/api/daily-packs/bulk")
async def upsert_daily_packs_bulk(payload: dict = Body(...)):
    """Upsert many (date, pack count) rows at once — used by the daily packs batch UI.

    Request: {"entries": [{"record_date": "YYYY-MM-DD", "number_of_packs": 123, "note": "..."}, ...]}
    Each entry is validated and persisted independently; the response reports per-entry
    success so the frontend can flag the ones that failed without losing the rest.
    """
    entries = payload.get("entries")
    if not isinstance(entries, list) or not entries:
        raise HTTPException(status_code=400, detail="entries must be a non-empty list")

    results: list[dict] = []
    saved_count = 0
    failed_count = 0

    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            for raw in entries:
                record_date = raw.get("record_date") if isinstance(raw, dict) else None
                number_of_packs = raw.get("number_of_packs") if isinstance(raw, dict) else None
                note = raw.get("note") if isinstance(raw, dict) else None

                entry_result: dict = {
                    "record_date": record_date,
                    "saved": False,
                    "number_of_packs": None,
                    "error": None,
                }
                try:
                    if not record_date:
                        raise ValueError("record_date is required")
                    parsed_date = datetime.strptime(record_date, "%Y-%m-%d").date()
                    packs_value = int(number_of_packs)
                    if packs_value < 0:
                        raise ValueError("number_of_packs cannot be negative")

                    cursor.execute(
                        """
                        INSERT INTO daily_packs (record_date, number_of_packs, note)
                        VALUES (%s, %s, %s)
                        ON CONFLICT (record_date) DO UPDATE
                            SET number_of_packs = EXCLUDED.number_of_packs,
                                note = EXCLUDED.note,
                                updated_at = NOW()
                        RETURNING number_of_packs
                        """,
                        (parsed_date, packs_value, note),
                    )
                    row = cursor.fetchone()
                    entry_result["saved"] = True
                    entry_result["number_of_packs"] = row[0]
                    saved_count += 1
                except (TypeError, ValueError) as exc:
                    entry_result["error"] = str(exc)
                    failed_count += 1
                except Exception as exc:  # noqa: BLE001
                    entry_result["error"] = f"database error: {exc}"
                    failed_count += 1

                results.append(entry_result)
        connection.commit()

    return {
        "total": len(entries),
        "saved": saved_count,
        "failed": failed_count,
        "results": results,
    }


@app.post("/api/daily-packs")
async def upsert_daily_packs(payload: dict = Body(...)):
    """Insert or update the pack count for a single date (one row per date)."""
    record_date = payload.get("record_date")
    number_of_packs = payload.get("number_of_packs")
    note = payload.get("note")

    if not record_date:
        raise HTTPException(status_code=400, detail="record_date is required")
    try:
        parsed_date = datetime.strptime(record_date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="record_date must be YYYY-MM-DD") from exc
    try:
        packs_value = int(number_of_packs)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="number_of_packs must be an integer") from exc
    if packs_value < 0:
        raise HTTPException(status_code=400, detail="number_of_packs cannot be negative")

    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO daily_packs (record_date, number_of_packs, note)
                VALUES (%s, %s, %s)
                ON CONFLICT (record_date) DO UPDATE
                    SET number_of_packs = EXCLUDED.number_of_packs,
                        note = EXCLUDED.note,
                        updated_at = NOW()
                RETURNING number_of_packs, updated_at
                """,
                (parsed_date, packs_value, note),
            )
            saved = cursor.fetchone()
        connection.commit()

    return {
        "record_date": record_date,
        "number_of_packs": saved[0],
        "updated_at": saved[1].isoformat() if saved[1] else None,
        "saved": True,
    }


@app.get("/api/daily-packs/{record_date}/history")
async def daily_packs_history(record_date: str, limit: int = 50):
    """Change history for one date's pack count. Powered by the
    daily_packs_audit_trigger postgres trigger; one row per change.
    Newest first.

    Session-gated on GenbaLink hosts (path does not start with any
    GBL_PUBLIC_PREFIXES entry — falls through to the cookie check)."""
    try:
        parsed_date = datetime.strptime(record_date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="record_date must be YYYY-MM-DD") from exc
    limit = max(1, min(int(limit or 50), 500))
    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT id, action, old_packs, new_packs, old_note, new_note,
                       changed_at, changed_by, source_hint
                FROM daily_packs_history
                WHERE record_date = %s
                ORDER BY changed_at DESC, id DESC
                LIMIT %s
                """,
                (parsed_date, limit),
            )
            rows = cursor.fetchall()
    # Normalise timestamps for JSON
    for r in rows:
        if r.get("changed_at") and hasattr(r["changed_at"], "isoformat"):
            r["changed_at"] = r["changed_at"].isoformat(timespec="seconds")
    return {
        "record_date": record_date,
        "count": len(rows),
        "history": rows,
    }


# ============================================================================
# Daily Packs Excel — per-product breakdown + end-time prediction
# ============================================================================
PRODUCTION_RATES_PATH = BASE_DIR / "production_rates.json"
DEFAULT_RATE_PER_HOUR = 2000

def _normalize_product_key(name: str) -> str:
    """Lower + NFKC normalize + strip whitespace, for fuzzy product-name lookup."""
    if not name:
        return ""
    s = unicodedata.normalize("NFKC", str(name))
    return "".join(s.split()).lower()

def _load_rates() -> dict:
    raw = _load_json(PRODUCTION_RATES_PATH, {}) or {}
    if not isinstance(raw, dict):
        raw = {}
    raw.setdefault("_default_rate_per_hour", DEFAULT_RATE_PER_HOUR)
    raw.setdefault("products", {})
    return raw

def _save_rates(d: dict) -> None:
    d["updated_at"] = datetime.now().isoformat(timespec="seconds")
    _save_json(PRODUCTION_RATES_PATH, d)

def _rate_for_product(rates: dict, name: str) -> tuple[int, bool]:
    """Return (rate_per_hour, used_default). Try exact, then NFKC-normalized match."""
    products = rates.get("products", {}) or {}
    if name in products and products[name]:
        return int(products[name]), False
    key = _normalize_product_key(name)
    norm_map = {_normalize_product_key(k): v for k, v in products.items() if v}
    if key in norm_map:
        return int(norm_map[key]), False
    return int(rates.get("_default_rate_per_hour") or DEFAULT_RATE_PER_HOUR), True


@app.get("/api/production-rates")
async def production_rates_get():
    return _load_rates()

@app.post("/api/production-rates")
async def production_rates_set(payload: dict = Body(...)):
    cur = _load_rates()
    if "_default_rate_per_hour" in payload:
        try:
            cur["_default_rate_per_hour"] = max(1, int(payload["_default_rate_per_hour"]))
        except Exception as exc:
            raise HTTPException(status_code=400, detail="_default_rate_per_hour must be a positive integer") from exc
    if "products" in payload and isinstance(payload["products"], dict):
        out: dict[str, int] = {}
        for k, v in payload["products"].items():
            try:
                rv = int(v)
            except Exception:
                continue
            if rv > 0:
                out[str(k)] = rv
        cur["products"] = out
    _save_rates(cur)
    return {"ok": True, "rates": cur}


def _excel_cell_int(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v != v:
            return None
        return int(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None

def _excel_cell_str(v) -> str:
    return "" if v is None else str(v).strip()

# Excel epoch is 1900 with the historical leap-year bug → 1899-12-30 base.
EXCEL_EPOCH = datetime(1899, 12, 30)
def _xlsx_serial_to_date(n):
    try:
        return (EXCEL_EPOCH + timedelta(days=int(n))).date()
    except Exception:
        return None

def _nfkc(s) -> str:
    """NFKC normalize + strip. Folds fullwidth Ｎ→N, Ｙ→Y so header detection
    works on either form. Returns empty string for None."""
    if s is None:
        return ""
    return unicodedata.normalize("NFKC", str(s)).strip()

# Cells whose text matches one of these are layout labels, not products —
# skip even if they pass the "looks like Japanese text" check.
_PACK_NOISE_NAMES = {
    "店舗数", "合計", "合 計", "山梨", "長野", "松本", "石倉",
    "天気", "温度", "気温", "晴れ", "曇り", "雨", "雪", "晴れ曇り雨雪",
    "仮確定", "入力者", "入り数", "全合計",
    "N便", "Y便", "N便計", "Y便計", "N合計", "Y合計", "Ｎ便", "Ｙ便",
    "Ｎ便計", "Ｙ便計", "Ｎ合計", "Ｙ合計",
    "商品名", "商 品 名", "２便製造分", "３便製造分", "2便製造分", "3便製造分",
    "受注集計表", "製造日",
}

def _is_product_name(s) -> bool:
    """A product-name cell is a non-empty string of ≥3 chars that contains at
    least one non-ASCII (Japanese) character. Rejects single-letter IDs (A,
    B, …) and numeric IDs from column 0 of the 入力画面 layout."""
    s = _nfkc(s)
    if not s or len(s) < 3:
        return False
    if s.replace(",", "").replace(".", "").replace("-", "").isdigit():
        return False
    if s in _PACK_NOISE_NAMES:
        return False
    return any(ord(ch) > 127 for ch in s)

def _find_pack_header_window(rows, max_scan=40):
    """Anchor on the row containing N便計, then extend forward up to 2 rows
    so the batch-label row (Ｎ便/Ｙ便) and the region row (山梨/長野/松本)
    are included. Returns (top_row, bottom_row_exclusive)."""
    anchor = None
    for i in range(min(max_scan, len(rows))):
        for c in (rows[i] or ()):
            if _nfkc(c) == "N便計":
                anchor = i
                break
        if anchor is not None:
            break
    if anchor is None:
        return None, None
    return anchor, min(anchor + 3, len(rows))

def _resolve_pack_columns(rows, top, bottom):
    """Within the header window, locate column indices for header tokens."""
    out = {
        "n_total_col": None, "y_total_col": None, "grand_col": None, "ppc_col": None,
        "n_label_col": None, "y_label_col": None,
        "yamanashi": [], "nagano": [], "matsumoto": [],
    }
    for r in range(top, bottom):
        row = rows[r] or ()
        for j, c in enumerate(row):
            t = _nfkc(c)
            if not t:
                continue
            if   t == "N便計": out["n_total_col"] = j
            elif t == "Y便計": out["y_total_col"] = j
            elif t == "全合計": out["grand_col"]  = j
            elif t == "入り数": out["ppc_col"]    = j
            elif t == "N便":   out["n_label_col"] = j
            elif t == "Y便":   out["y_label_col"] = j
            elif t == "山梨":  out["yamanashi"].append(j)
            elif t == "長野":  out["nagano"].append(j)
            elif t == "松本":  out["matsumoto"].append(j)
    return out

def _split_region_cols(cols):
    """山梨/長野/松本 each appear twice (under N便 and Y便). Smaller column
    index = N便, larger = Y便. If only one is present, fall back to comparing
    against the N便 / Y便 label columns."""
    region = {"n_yamanashi": None, "n_nagano": None, "n_matsumoto": None,
              "y_yamanashi": None, "y_nagano": None, "y_matsumoto": None}
    for src, key_n, key_y in (("yamanashi", "n_yamanashi", "y_yamanashi"),
                              ("nagano",    "n_nagano",    "y_nagano"),
                              ("matsumoto", "n_matsumoto", "y_matsumoto")):
        sorted_cols = sorted(set(cols[src]))
        if len(sorted_cols) >= 2:
            region[key_n] = sorted_cols[0]
            region[key_y] = sorted_cols[1]
        elif len(sorted_cols) == 1:
            c = sorted_cols[0]
            ylabel = cols.get("y_label_col")
            if ylabel is not None and c >= ylabel:
                region[key_y] = c
            else:
                region[key_n] = c
    return region

def _value_near_label(rows, i, j, kind):
    """Return the value cell for a label at (i,j). Excel layouts put the value
    either to the right (same row) OR directly below (same column), so we
    check both. `kind` selects the parser: 'date', 'text', 'int', 'temp'."""
    def _cells_to_right(r, max_d=8):
        rrow = rows[r] or ()
        for k in range(j + 1, min(j + 1 + max_d, len(rrow))):
            yield rrow[k]
    def _cells_below(c, max_d=3):
        for r in range(i + 1, min(i + 1 + max_d, len(rows))):
            rrow = rows[r] or ()
            if c < len(rrow):
                yield rrow[c]
    candidates = list(_cells_to_right(i)) + list(_cells_below(j))
    for v in candidates:
        if v is None: continue
        if isinstance(v, bool): continue
        if kind == "date":
            if isinstance(v, datetime):
                return v.date().isoformat()
            if isinstance(v, (int, float)) and v > 30000:
                d = _xlsx_serial_to_date(v)
                if d: return d.isoformat()
            s = _nfkc(v)
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
                try: return datetime.strptime(s, fmt).date().isoformat()
                except Exception: continue
        elif kind == "text":
            s = _nfkc(v)
            # Only reject layout labels that would loop back to themselves; keep
            # weather words like 曇り, 晴れ, 雨, 雪 — those *are* valid values.
            if s and s != "入力者" and s != "天気" and s != "温度" and s != "気温" and s != "晴れ曇り雨雪":
                return s
        elif kind in ("int", "temp"):
            n = _excel_cell_int(v)
            if n is not None:
                return n
    return None

def _scan_pack_meta(rows, header_top):
    """Scan the whole sheet for 製造日 / 入力者 / 天気 / 温度. Values may sit
    to the right of the label or directly below it depending on the sheet."""
    meta = {"production_date": None, "input_by": None, "weather": None, "temperature": None}
    for i, row in enumerate(rows):
        for j, c in enumerate(row or ()):
            t = _nfkc(c)
            if not t:
                continue
            if t == "製造日" and meta["production_date"] is None:
                v = _value_near_label(rows, i, j, "date")
                if v: meta["production_date"] = v
            elif t == "入力者" and not meta["input_by"]:
                v = _value_near_label(rows, i, j, "text")
                if v: meta["input_by"] = v
            elif t == "天気" and not meta["weather"]:
                v = _value_near_label(rows, i, j, "text")
                if v and v != "晴れ曇り雨雪": meta["weather"] = v
            elif t in ("温度", "気温") and meta["temperature"] is None:
                v = _value_near_label(rows, i, j, "temp")
                if v is not None: meta["temperature"] = v
    return meta

def _parse_pack_sheet(ws, sheet_name):
    rows = list(ws.iter_rows(values_only=True))
    top, bottom = _find_pack_header_window(rows)
    if top is None:
        raise ValueError(f"Could not find header row (N便計 / Y便計) on sheet '{sheet_name}'.")
    cols = _resolve_pack_columns(rows, top, bottom)
    region_cols = _split_region_cols(cols)
    meta = _scan_pack_meta(rows, top)

    products: list[dict] = []
    blank_streak = 0
    for i in range(bottom, len(rows)):
        row = rows[i] or ()
        name = ""
        for c in row:
            t = _nfkc(c)
            if _is_product_name(t):
                name = t
                break

        def _at(col):
            return row[col] if (col is not None and col < len(row)) else None

        n_total = _excel_cell_int(_at(cols["n_total_col"]))
        y_total = _excel_cell_int(_at(cols["y_total_col"]))
        grand   = _excel_cell_int(_at(cols["grand_col"]))
        ppc     = _excel_cell_int(_at(cols["ppc_col"]))
        no_qty = (grand is None and n_total is None and y_total is None)

        if not name and no_qty:
            blank_streak += 1
            if blank_streak >= 4:
                break
            continue
        blank_streak = 0
        # Skip rows with no product name OR no quantities at all
        if not name or no_qty:
            continue

        item = {
            "product_name": name,
            "product_key":  _normalize_product_key(name),
            "n_total": n_total, "y_total": y_total, "grand_total": grand, "packs_per_case": ppc,
        }
        for key, col in region_cols.items():
            item[key] = _excel_cell_int(_at(col))
        products.append(item)

    if not products:
        raise ValueError(f"Header found on '{sheet_name}' but no product rows extracted.")

    return {
        "meta": meta,
        "products": products,
        "header_row": top,
        "region_cols": region_cols,
        "header_cols": {
            "N便計": cols["n_total_col"], "Y便計": cols["y_total_col"],
            "全合計": cols["grand_col"],   "入り数": cols["ppc_col"],
        },
        "sheet_name": sheet_name,
    }

# ---------------------------------------------------------------------------
# Sub-parsers added 2026-05-01: read three more pieces from the workbook so
# the daily-packs flow can persist:
#   1) Section-2 right-side totals (Ｎ合計 / Ｙ合計 / sum) on 入力画面
#   2) フルキャスト manual rows on 人時生産性 (R78/R79 area for Section 2,
#      plus the matching block for Section 1 lower in the same sheet)
#   3) Section-2 production plan (Aライン / Bライン / Cライン) on 製造予定表 ＮＹ
#
# All three search by Japanese LABEL, never by hardcoded coordinates — the
# operator's workbook gets minor cell shifts month over month.
# ---------------------------------------------------------------------------

def _norm_jp(s) -> str:
    """NFKC + strip whitespace including ideographic space — cheaper variant
    of _nfkc for label matching where we also want to fold full/half spaces."""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s))
    return t.replace("　", "").replace(" ", "").strip()


def _parse_section_totals_from_input_sheet(ws) -> dict | None:
    """Find Ｎ合計 / Ｙ合計 by label in the top header rows of 入力画面 and
    return their numeric values + a combined sum. Tolerates the screenshot's
    layout where the two cells sit far to the right of the per-product grid
    and may be one or more rows above the value row.

    Returns {n_total, y_total, combined, n_label_at:[r,c], y_label_at:[r,c]}
    or None if labels not found.
    """
    n_at = y_at = None
    # Scan first ~10 rows × all columns for the labels
    max_r = min(ws.max_row, 12)
    max_c = min(ws.max_column, 80)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            t = _norm_jp(v)
            if not t:
                continue
            if n_at is None and t == "N合計":
                n_at = (r, c)
            elif y_at is None and t == "Y合計":
                y_at = (r, c)
            if n_at and y_at:
                break
        if n_at and y_at:
            break
    if n_at is None or y_at is None:
        return None
    # Walk down the column under each label until we find the first numeric.
    def first_num_below(r0, c0):
        for r in range(r0 + 1, min(ws.max_row + 1, r0 + 8)):
            v = ws.cell(r, c0).value
            iv = _excel_cell_int(v)
            if iv is not None and iv >= 0:
                return iv, r
        return None, None
    n_total, n_row = first_num_below(*n_at)
    y_total, y_row = first_num_below(*y_at)
    return {
        "n_total": n_total or 0,
        "y_total": y_total or 0,
        "combined": (n_total or 0) + (y_total or 0),
        "n_label_at": list(n_at),
        "y_label_at": list(y_at),
        "n_value_at": [n_row, n_at[1]] if n_row else None,
        "y_value_at": [y_row, y_at[1]] if y_row else None,
    }


def _parse_fullcast_rows_from_jisei_sheet(ws) -> list[dict]:
    """Read フルキャスト rows from the 人時生産性 sheet.

    Layout (verified on the real 夜勤用日報 workbook):
      • 人時生産性計算　製造２課 starts at R1; section ends with two rows
        whose col A = headcount (e.g. 8 then 1) and col C = total worked time
        (timedelta), then a 人時生産性 calc row in col A.
      • 人時生産性計算　製造１課 begins later in the same sheet with the same
        shape; its フルキャスト rows are the two rows above its 人時生産性 row.

    We locate each section by scanning for the label '人時生産性\\n…' in col A,
    walk back up to two rows that look like fullcast (small int in col A,
    timedelta in col C) and emit them.
    """
    out: list[dict] = []
    # Find rows whose col A starts with '人時生産性' (the calc label rows)
    section_anchors = []
    for r in range(1, min(ws.max_row + 1, 400)):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "人時生産性" in v and "計算" not in v:
            section_anchors.append(r)
    # Also look for the section title rows (人時生産性計算　製造X課) to label
    section_titles: list[tuple[int, str]] = []  # (row, "section_label")
    for r in range(1, min(ws.max_row + 1, 400)):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "人時生産性計算" in v:
            section_titles.append((r, v))

    def label_for_anchor(anchor_row: int) -> str:
        # The closest section title above this anchor
        prev = [t for t in section_titles if t[0] < anchor_row]
        if not prev:
            return ""
        title = prev[-1][1]
        if "２課" in title or "2課" in title:
            return "製造2課"
        if "１課" in title or "1課" in title:
            return "製造1課"
        return _norm_jp(title)

    def _hhmm_from_cell(v):
        """Read a 人時生産性 cell that may be timedelta, datetime.time, or
        datetime.datetime (Excel's 1900-base) and return 'HH:MM' (or None)."""
        if v is None:
            return None
        if isinstance(v, timedelta):
            secs = int(v.total_seconds()) % (24 * 3600)
            return f"{secs // 3600:02d}:{(secs % 3600) // 60:02d}"
        if hasattr(v, "hour") and hasattr(v, "minute"):
            return f"{v.hour:02d}:{v.minute:02d}"
        return None

    for anchor in section_anchors:
        # Walk up from anchor-1 collecting candidate fullcast rows until
        # we hit an employee row (col B has a name string) or the section title.
        # Column layout (verified on 夜勤用日報 workbook, rows 84-85):
        #   A = headcount         (e.g. 5)
        #   B = blank
        #   C = start_time        (timedelta — Excel renders it as "19:00")
        #   D = leave_time        (datetime — Excel renders the time part, e.g. "04:00")
        #   E = total worked time (timedelta — the real labor: headcount × duration)
        candidates = []
        r = anchor - 1
        while r > 0 and len(candidates) < 4:
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            c = ws.cell(r, 3).value
            d = ws.cell(r, 4).value
            e = ws.cell(r, 5).value
            if isinstance(a, (int, float)) and 0 < a <= 50 and (b is None or _norm_jp(b) == ""):
                start_hhmm = _hhmm_from_cell(c)
                leave_hhmm = _hhmm_from_cell(d)
                total_secs = 0
                leave_next_day = False
                if isinstance(e, timedelta):
                    total_secs = int(e.total_seconds())
                    leave_next_day = e.days >= 1
                # Heuristic for leave-next-day when col E is missing: leave time
                # numerically less than start time means it crosses midnight.
                if start_hhmm and leave_hhmm and not leave_next_day:
                    sh, sm = (int(x) for x in start_hhmm.split(":"))
                    lh, lm = (int(x) for x in leave_hhmm.split(":"))
                    if (lh * 60 + lm) <= (sh * 60 + sm):
                        leave_next_day = True
                candidates.append({
                    "row": r, "headcount": int(a),
                    "start_time": start_hhmm, "leave_time": leave_hhmm,
                    "leave_next_day": leave_next_day,
                    "total_seconds": total_secs,
                })
                r -= 1
                continue
            if isinstance(b, str) and b.strip():
                break
            if isinstance(a, str) and ("人時生産性計算" in a or "個人コード" in a):
                break
            r -= 1
        # Reverse so output is in workbook order (top-down)
        for cand in reversed(candidates):
            head = cand["headcount"]
            secs = cand["total_seconds"]
            per_person_h = round((secs / head) / 3600.0, 2) if (head and secs) else None
            out.append({
                "section_label": label_for_anchor(anchor),
                "headcount": head,
                "start_time": cand["start_time"],
                "leave_time": cand["leave_time"],
                "leave_next_day": cand["leave_next_day"],
                "total_seconds": secs,
                "total_hours": round(secs / 3600.0, 2),
                "hours_per_person": per_person_h,
                "source_row": cand["row"],
            })
    return out


_PLAN_LINE_HEADERS = {
    "Aライン": "A", "Aライン計": "A",
    "Bライン": "B", "Bライン計": "B",
    "Cライン": "C", "Cライン計": "C",
}


def _parse_production_plan_sheet(ws) -> dict:
    """Parse 製造予定表 ＮＹ (Section-2 plan).

    Header columns (verified on the real workbook): item-name in col A; we
    locate '確定' / 'Ｙ便' / 'Ｎ+Ｙ' / '製造\\n時間' / 'タイムテーブル' / 'p/h'
    / '必要\\n人員' by label in the row directly under each Aライン/Bライン/Cライン
    section header. Then walk rows until we hit a 'X line 計' row (the line total).
    """
    rows = ws.max_row
    cols_max = ws.max_column

    # Find each line section's header rows (label row + the row right under it
    # which holds the column titles for that line)
    line_blocks: list[dict] = []
    for r in range(1, rows + 1):
        a = _norm_jp(ws.cell(r, 1).value)
        if a in ("Aライン", "Bライン", "Cライン"):
            line_blocks.append({"line": _PLAN_LINE_HEADERS[a], "header_row": r, "title_row": r + 1})

    plan: dict = {"section_start_time": None, "lines": {"A": [], "B": [], "C": []}, "totals": {"A": 0, "B": 0, "C": 0}}

    last_col_map: dict = {}  # carry across line blocks — B/C don't always repeat every column header
    for blk in line_blocks:
        # Read the column-title row to find which column is which
        col_map = dict(last_col_map)  # inherit from previous line; per-line title overrides
        for c in range(1, cols_max + 1):
            t = _norm_jp(ws.cell(blk["title_row"], c).value)
            if not t:
                continue
            if t == "確定":             col_map["n_qty"] = c
            elif t == "Y便":            col_map["y_qty"] = c
            elif t in ("N+Y", "Ｎ+Ｙ"):  col_map["combined"] = c
            elif t == "製造時間":        col_map["takt"] = c
            elif t == "タイムテーブル":   col_map["start"] = c
            elif t == "切替時間":        col_map["changeover"] = c
            elif t == "p/h" or t.lower() == "p/h":
                                       col_map["pph_target"] = c
            elif t == "必要人員":        col_map["required_staff"] = c
            elif t in ("製造予定数", "製造\n予定数"):
                                       col_map["planned_qty"] = c
            elif t == "アイテム":        col_map["item_name"] = c
        last_col_map = dict(col_map)
        # Item-name column defaults to col 1 if not flagged explicitly
        col_item = col_map.get("item_name", 1)
        # Walk rows from title_row+1 down until we hit 'X計' or a blank row
        items = []
        line_total = 0
        r = blk["title_row"] + 1
        while r <= rows:
            name_v = ws.cell(r, col_item).value
            name = _norm_jp(name_v)
            if not name:
                r += 1
                if r - blk["title_row"] > 30:
                    break
                continue
            if name in ("Aライン計", "Bライン計", "Cライン計", "Ａライン計", "Ｂライン計", "Ｃライン計"):
                # Read total qty from the planned_qty column (or combined) on this row
                tot_col = col_map.get("planned_qty") or col_map.get("combined")
                if tot_col:
                    line_total = _excel_cell_int(ws.cell(r, tot_col).value) or 0
                break
            # Also stop if we hit the next section header or "２便計" / "３便計" / blank far down
            if name in ("Aライン", "Bライン", "Cライン", "２便計", "３便計", "二便計", "三便計"):
                break
            # Build an item record
            def _at(col_key):
                col = col_map.get(col_key)
                if col is None:
                    return None
                return ws.cell(r, col).value
            def _int(col_key):
                return _excel_cell_int(_at(col_key))
            def _time_str(col_key):
                v = _at(col_key)
                if v is None:
                    return None
                if hasattr(v, "hour") and hasattr(v, "minute"):
                    return f"{v.hour:02d}:{v.minute:02d}"
                return _norm_jp(v) or None
            items.append({
                "item_name": name,
                "planned_qty": _int("planned_qty") or _int("combined") or 0,
                "n_qty":       _int("n_qty") or 0,
                "y_qty":       _int("y_qty") or 0,
                "combined":    _int("combined") or 0,
                "start_time":  _time_str("start"),
                "takt_time":   _time_str("takt"),
                "pph_target":  _int("pph_target"),
                "required_staff": _int("required_staff"),
            })
            r += 1
        plan["lines"][blk["line"]] = items
        plan["totals"][blk["line"]] = line_total
    # Section start time: pick the LATEST first-item start across all
    # non-empty lines. The workbook sometimes flips which line carries the
    # "main" (night) production vs the early prep — but the actual canonical
    # night-shift start is always the latest of the line-firsts (typically
    # 19:20). Falls back to None if no line has items with a start time.
    candidates = []
    for ln in ("A", "B", "C"):
        items = plan["lines"].get(ln) or []
        if items:
            t = items[0].get("start_time")
            if t and re.match(r"^\d{1,2}:\d{2}$", t):
                candidates.append((t, ln))
    if candidates:
        # Sort by HH*60+MM ascending, take the LATEST
        def to_min(s): h, m = s.split(":"); return int(h) * 60 + int(m)
        candidates.sort(key=lambda c: to_min(c[0]))
        plan["section_start_time"] = candidates[-1][0]
        plan["section_start_source_line"] = candidates[-1][1]
    return plan


# Sheets to try in order. 入力画面 is the canonical input view in the real
# 日報 workbook (matches what the PDF prints). 受注集計表 is a secondary
# summary view used in some older files.
_PACK_SHEET_PREFERENCE = ("入力画面", "受注集計表")

def parse_daily_pack_excel(file_path: Path) -> dict:
    """Parse the daily-packs Excel by trying preferred sheets first, then any
    other sheet that happens to contain the N便計/Y便計 header pair.
    Layout-tolerant: NFKC-folds fullwidth Ｎ/Ｙ, allows the header to span up
    to 3 rows, and skips single-letter ID columns when picking product names.

    Also pulls (when present):
      • section_totals  — Ｎ合計 / Ｙ合計 / sum from 入力画面
      • fullcast        — フルキャスト rows from 人時生産性 (Section 1 + 2)
      • production_plan — A/B/C lines from 製造予定表 ＮＹ (Section 2)
    """
    # We need read_only=False so we can call ws.cell() directly with random
    # access for the auxiliary parsers (they label-search the whole sheet).
    wb = load_workbook(file_path, data_only=True, read_only=False, keep_vba=False)
    sheets_in_order = (
        [s for s in _PACK_SHEET_PREFERENCE if s in wb.sheetnames]
        + [s for s in wb.sheetnames if s not in _PACK_SHEET_PREFERENCE]
    )
    base: dict | None = None
    tried = []
    for sheet_name in sheets_in_order:
        try:
            base = _parse_pack_sheet(wb[sheet_name], sheet_name)
            base["_input_sheet_name"] = sheet_name
            break
        except Exception as exc:
            tried.append(f"{sheet_name}: {exc}")
            continue
    if base is None:
        raise ValueError(
            "No sheet matched the daily-packs layout. Tried: "
            + " | ".join(tried[:5])
        )

    # Section totals (right-side Ｎ合計 / Ｙ合計) — from the same sheet we just parsed.
    try:
        if base.get("_input_sheet_name") in wb.sheetnames:
            base["section_totals"] = _parse_section_totals_from_input_sheet(wb[base["_input_sheet_name"]])
    except Exception:
        base["section_totals"] = None

    # Fullcast rows on 人時生産性 (best-effort — not all workbooks have this sheet)
    try:
        if "人時生産性" in wb.sheetnames:
            base["fullcast"] = _parse_fullcast_rows_from_jisei_sheet(wb["人時生産性"])
        else:
            base["fullcast"] = []
    except Exception:
        base["fullcast"] = []

    # Production plan (A/B/C lines) — sheet name has an ideographic space.
    plan_sheet = next((s for s in wb.sheetnames if "製造予定表" in s and ("ＮＹ" in s or "NY" in s)), None)
    if plan_sheet:
        try:
            base["production_plan"] = _parse_production_plan_sheet(wb[plan_sheet])
        except Exception:
            base["production_plan"] = None
    else:
        base["production_plan"] = None

    return base


def _suggest_start_time(production_date) -> dict:
    """Choose 17:00 or 19:00 based on the median IN time on `production_date`.

    Heuristic: most workers arrive ~30 min before start. Snap median - 30 min
    to whichever known start (17:00 / 19:00) is closer.
    """
    target_date = production_date
    if isinstance(target_date, str):
        target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
    rows: list[str] = []
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT commute_time FROM attendance_records
                    WHERE record_date = %s AND commute_time IS NOT NULL AND commute_time <> ''
                    """,
                    (target_date,),
                )
                rows = [r[0] for r in cur.fetchall()]
    except Exception:
        rows = []
    minutes: list[int] = []
    for s in rows:
        m = re.match(r"^\s*(\d{1,2}):(\d{2})", str(s))
        if not m:
            continue
        mins = int(m.group(1)) * 60 + int(m.group(2))
        if mins < 5 * 60:        # treat 0–4 as overnight
            mins += 24 * 60
        minutes.append(mins)
    if not minutes:
        return {"start_time": "17:00", "source": "default", "median_in": None, "n": 0,
                "candidates": ["17:00", "19:00"], "note": "no attendance data — default 17:00"}
    minutes.sort()
    median = minutes[len(minutes) // 2]
    expected = median - 30
    s17 = abs(expected - 17 * 60)
    s19 = abs(expected - 19 * 60)
    chosen = "17:00" if s17 <= s19 else "19:00"
    mh, mm = divmod(median % (24 * 60), 60)
    return {"start_time": chosen, "source": "auto-detected",
            "median_in": f"{mh:02d}:{mm:02d}", "n": len(minutes),
            "candidates": ["17:00", "19:00"]}


@app.get("/api/daily-packs/start-time/suggest")
async def daily_packs_start_time(date: str):
    try:
        d = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc
    return _suggest_start_time(d)


def _build_prediction(products: list[dict], start_time: str, rates: dict) -> dict:
    total_qty = 0
    total_seconds = 0
    breakdown = []
    for p in products:
        qty = int(p.get("grand_total") or 0)
        rate, used_default = _rate_for_product(rates, p.get("product_name", "") or "")
        secs = int(round((qty / max(rate, 1)) * 3600)) if qty else 0
        total_qty += qty
        total_seconds += secs
        breakdown.append({
            "product_name":  p.get("product_name"),
            "qty":           qty,
            "rate_per_hour": rate,
            "rate_default":  used_default,
            "minutes":       round(secs / 60, 1),
        })
    try:
        sh, sm = [int(x) for x in (start_time or "17:00").split(":")]
    except Exception:
        sh, sm = 17, 0
    end_minutes = sh * 60 + sm + total_seconds // 60
    eh = (end_minutes // 60) % 24
    em = end_minutes % 60
    overnight = (end_minutes // 60) >= 24
    avg_rate = round(total_qty / (total_seconds / 3600), 1) if total_seconds else 0
    return {
        "start_time":     f"{sh:02d}:{sm:02d}",
        "end_time":       f"{eh:02d}:{em:02d}",
        "end_overnight":  overnight,
        "total_qty":      total_qty,
        "total_minutes":  round(total_seconds / 60, 1),
        "avg_rate":       avg_rate,
        "products_count": len(products),
        "breakdown":      breakdown,
    }


def _pick_latest_xlsx_in(folder: Path):
    if folder.exists() and folder.is_file() and folder.suffix.lower() in (".xlsx", ".xlsm"):
        return folder
    if not folder.exists() or not folder.is_dir():
        return None
    files = [p for p in folder.iterdir()
             if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm")
             and not p.name.startswith("~$")]
    if not files:
        return None
    files.sort(key=lambda p: (p.name, p.stat().st_mtime), reverse=True)
    return files[0]


def _pick_xlsx_for_date(folder: Path, target_date) -> Path | None:
    """Find the .xlsx/.xlsm in `folder` whose filename encodes `target_date`.
    Returns None if no match (caller turns this into a 'file_not_available' 404)."""
    if folder is None or not folder.exists() or not folder.is_dir() or target_date is None:
        return None
    matches = []
    for p in folder.iterdir():
        if not (p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm")):
            continue
        if p.name.startswith("~$"):
            continue
        d = _extract_date_from_filename(p.name)
        if d == target_date:
            matches.append(p)
    if not matches:
        return None
    matches.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return matches[0]


@app.post("/api/daily-packs/auto-extract-excel")
async def auto_extract_daily_pack_excel(date: str | None = None):
    """Pick an .xlsx in the daily-packs watched folder, parse it, and return
    preview + start-time suggestion + prediction. Mirrors the PDF
    `auto-extract` so the Excel segment can reuse the Auto-update pattern.

    Without `date`: picks the latest .xlsx (legacy behavior).
    With `date=YYYY-MM-DD`: picks the .xlsx whose filename encodes that date.
    Returns a structured `file_not_available` 404 if no file matches.
    """
    folder, single_file = _resolve_watched_target(DAILY_PACKS_AUTO_UPLOAD_DIR)
    if folder is None:
        path_str = str(DAILY_PACKS_AUTO_UPLOAD_DIR)
        if _looks_like_windows_path(path_str):
            raise HTTPException(status_code=404, detail=f"Configured path is a Windows-only path the Pi cannot read: {path_str}")
        raise HTTPException(status_code=404, detail=f"Watched path not reachable: {path_str}")
    target_date = None
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc
    if target_date is not None:
        if single_file is not None and single_file.suffix.lower() in (".xlsx", ".xlsm"):
            picked = single_file if _extract_date_from_filename(single_file.name) == target_date else None
        else:
            picked = _pick_xlsx_for_date(folder, target_date)
        if picked is None:
            raise HTTPException(
                status_code=404,
                detail={
                    "error": "file_not_available",
                    "kind": "daily_packs_xlsx",
                    "requested_date": target_date.isoformat(),
                    "watched_folder": str(folder),
                    "message": f"No daily-packs Excel for {target_date.isoformat()} in {folder}. "
                               f"Upload the .xlsx first via POST /api/v1/xlsx/upload.",
                },
            )
    else:
        picked = single_file if (single_file is not None and single_file.suffix.lower() in (".xlsx", ".xlsm")) else _pick_latest_xlsx_in(folder)
        if picked is None:
            raise HTTPException(status_code=404, detail=f"No .xlsx files found in {folder}")
    try:
        parsed = parse_daily_pack_excel(picked)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel parse failed: {exc}") from exc

    rates = _load_rates()
    pdate = parsed["meta"].get("production_date")
    start = _suggest_start_time(pdate) if pdate else {"start_time": "17:00", "source": "default", "n": 0, "candidates": ["17:00", "19:00"]}
    # If the workbook has a planned start time on 製造予定表, prefer that over the
    # heuristic — it's a confirmed value, not an inference.
    plan = parsed.get("production_plan") or {}
    if plan.get("section_start_time"):
        start = {"start_time": plan["section_start_time"], "source": "production_plan", "n": None,
                 "candidates": [start.get("start_time"), plan["section_start_time"]]}
    prediction = _build_prediction(parsed["products"], start["start_time"], rates)
    return {
        "source_filename": picked.name,
        "picked_size":     picked.stat().st_size,
        "meta":            parsed["meta"],
        "products":        parsed["products"],
        "start":           start,
        "prediction":      prediction,
        "section_totals":  parsed.get("section_totals"),
        "fullcast":        parsed.get("fullcast") or [],
        "production_plan": parsed.get("production_plan"),
    }


@app.post("/api/daily-packs/extract-excel")
async def extract_daily_pack_excel(file: UploadFile = File(...)):
    """Parse a .xlsx and return preview + start-time suggestion + prediction.
    Does NOT save — frontend confirms first."""
    name = (file.filename or "").lower()
    if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="filename must end in .xlsx")
    tmp_dir = BASE_DIR / "logs"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp = tmp_dir / f"_pack_xlsx_{int(time.time() * 1000)}.xlsx"
    content = await file.read()
    tmp.write_bytes(content)
    try:
        parsed = parse_daily_pack_excel(tmp)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Excel parse failed: {exc}") from exc
    finally:
        try:
            tmp.unlink()
        except Exception:
            pass

    rates = _load_rates()
    pdate = parsed["meta"].get("production_date")
    start = _suggest_start_time(pdate) if pdate else {"start_time": "17:00", "source": "default", "n": 0, "candidates": ["17:00", "19:00"]}
    plan = parsed.get("production_plan") or {}
    if plan.get("section_start_time"):
        start = {"start_time": plan["section_start_time"], "source": "production_plan", "n": None,
                 "candidates": [start.get("start_time"), plan["section_start_time"]]}
    prediction = _build_prediction(parsed["products"], start["start_time"], rates)
    return {
        "source_filename": file.filename,
        "meta":            parsed["meta"],
        "products":        parsed["products"],
        "start":           start,
        "prediction":      prediction,
        # New blocks (parity with /auto-extract-excel) so the drag-drop UI
        # populates the section-totals strip + フルキャスト card + A/B/C
        # plan preview AND the subsequent save carries them through.
        "section_totals":  parsed.get("section_totals"),
        "fullcast":        parsed.get("fullcast") or [],
        "production_plan": parsed.get("production_plan"),
    }


@app.post("/api/daily-packs/save-excel-batch")
async def save_daily_pack_excel(request: Request, payload: dict = Body(...)):
    """Save a parsed batch into daily_pack_items. Overwrites same-date batch."""
    # On GenbaLink hosts, require either a logged-in session (browser console)
    # or a valid X-API-Key (desktop agent, doctor loopback). The host-auth
    # middleware now lets this path through via GBL_PUBLIC_PREFIXES, so the
    # handler is the only auth gate on link.genbafms.com. Other hosts
    # (rnd.asiakawaii.com legacy vhost, 127.0.0.1 loopback) keep their
    # existing wide-open behavior — same model as auto-extract-excel.
    if _gbl_host_requires_auth(request):
        if _gbl_request_user(request) is None and _resolve_key_label(request.headers.get("x-api-key")) is None:
            raise HTTPException(status_code=401, detail="authentication required")
    pdate_str = (payload.get("production_date") or "").strip()
    products  = payload.get("products") or []
    start     = (payload.get("start_time") or "17:00").strip()
    end       = (payload.get("end_time") or "").strip() or None
    source    = (payload.get("source_filename") or "").strip() or None
    weather   = (payload.get("weather") or "").strip() or None
    temperature = payload.get("temperature")
    input_by  = (payload.get("input_by") or "").strip() or None

    if not pdate_str:
        raise HTTPException(status_code=400, detail="production_date is required")
    try:
        pdate = datetime.strptime(pdate_str, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="production_date must be YYYY-MM-DD") from exc
    if not products:
        raise HTTPException(status_code=400, detail="products[] is empty")

    rates = _load_rates()
    batch_id = str(uuid.uuid4())
    inserted = 0
    total_packs = 0
    # Compute per-product effective qty: prefer grand_total, fall back to
    # n_total + y_total (some workbooks leave 全合計 blank). The day-level
    # total is later overridden by section_totals.combined when present —
    # that's the authoritative right-side Ｎ合計+Ｙ合計 the operator confirms.
    def _effective_qty(p: dict) -> int:
        gt = p.get("grand_total")
        if isinstance(gt, (int, float)) and gt > 0:
            return int(gt)
        return int(p.get("n_total") or 0) + int(p.get("y_total") or 0)
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM daily_pack_items WHERE production_date = %s", (pdate,))
            for p in products:
                rate, _ = _rate_for_product(rates, p.get("product_name") or "")
                qty = _effective_qty(p)
                total_packs += qty
                est_secs = int(round((qty / max(rate, 1)) * 3600)) if qty else 0
                cur.execute(
                    """
                    INSERT INTO daily_pack_items (
                        batch_id, production_date, product_name, product_key,
                        n_yamanashi, n_nagano, n_matsumoto,
                        y_yamanashi, y_nagano, y_matsumoto,
                        n_total, y_total, grand_total, packs_per_case,
                        rate_per_hour, est_seconds, source_filename,
                        weather, temperature, input_by, start_time, end_time
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                    """,
                    (
                        batch_id, pdate,
                        (p.get("product_name") or "")[:200],
                        _normalize_product_key(p.get("product_name") or "")[:200],
                        p.get("n_yamanashi"), p.get("n_nagano"), p.get("n_matsumoto"),
                        p.get("y_yamanashi"), p.get("y_nagano"), p.get("y_matsumoto"),
                        p.get("n_total"), p.get("y_total"), p.get("grand_total"), p.get("packs_per_case"),
                        rate, est_secs, source,
                        weather, temperature, input_by, start, end,
                    ),
                )
                inserted += 1
            # Also upsert the legacy daily_packs summary so the rest of the app
            # (productivity, gantt, summary report) keeps working off a single
            # number_of_packs per date — same row the PDF "Confirm & Save" writes.
            # Note format: "[source] filename · N products · batch <id8>" so the
            # Daily Packs UI can show how a row was entered (manual / pdf /
            # pdf-auto / excel / excel-auto).
            note_source = (payload.get("source_method") or "excel").strip() or "excel"
            note_parts = [f"[{note_source}]"]
            if source: note_parts.append(source)
            note_parts.append(f"{inserted} products")
            note_parts.append(f"batch {batch_id[:8]}")
            # Pull the parsed extras the UI passed through (or compute fresh
            # totals from products[] as a fallback for legacy callers).
            section_totals = payload.get("section_totals") or {}
            n_total_val = section_totals.get("n_total") if section_totals else None
            y_total_val = section_totals.get("y_total") if section_totals else None
            sec_start_t = (payload.get("section_start_time") or
                           (payload.get("production_plan") or {}).get("section_start_time"))
            # Authoritative day-level pack count: when the right-side Ｎ合計+Ｙ合計
            # was confirmed in the workbook, use it instead of the per-product sum.
            if section_totals and isinstance(section_totals.get("combined"), int) and section_totals["combined"] > 0:
                total_packs = int(section_totals["combined"])
            cur.execute(
                """
                INSERT INTO daily_packs (record_date, number_of_packs, note,
                                         n_total, y_total, section_start_time)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (record_date) DO UPDATE
                    SET number_of_packs    = EXCLUDED.number_of_packs,
                        note               = EXCLUDED.note,
                        n_total            = COALESCE(EXCLUDED.n_total, daily_packs.n_total),
                        y_total            = COALESCE(EXCLUDED.y_total, daily_packs.y_total),
                        section_start_time = COALESCE(EXCLUDED.section_start_time, daily_packs.section_start_time),
                        updated_at = NOW()
                """,
                (pdate, total_packs, " ".join(note_parts), n_total_val, y_total_val, sec_start_t),
            )

            # Persist フルキャスト rows into temp_staff. CRITICAL date rule
            # (memory: project_date_rules):
            #   record_date for temp_staff = SHIFT date = production_date − 1
            # The Excel's production_date is the day the packs are booked to;
            # the workers actually clocked in the night before. Every other
            # /api/temp-staff/{date} consumer (gantt, summary, manual fullcast
            # tab) keys on the shift date.
            shift_date = pdate - timedelta(days=1)
            if payload.get("skip_fullcast"):
                pass
            else:
                fullcast = payload.get("fullcast") or []
                if fullcast:
                    cur.execute("DELETE FROM temp_staff WHERE record_date = %s", (shift_date,))
                    for fc in fullcast:
                        head = int(fc.get("headcount") or 0)
                        secs = int(fc.get("total_seconds") or 0)
                        if head <= 0:
                            continue
                        # Each fullcast row carries its own start/leave time read
                        # straight from 人時生産性 — col C and col D of the row.
                        # Fall back to section start / 10:00 only when missing.
                        start_t = (fc.get("start_time") or sec_start_t or "19:00")[:5]
                        leave_t = (fc.get("leave_time") or "10:00")[:5]
                        leave_nxt = bool(fc.get("leave_next_day", True))
                        per_person_h = fc.get("hours_per_person")
                        if per_person_h is None and secs and head:
                            per_person_h = round((secs / head) / 3600.0, 2)
                        # If total_seconds is missing on the parsed row (older
                        # files), fall back to start/leave HH:MM math.
                        if (not secs) and start_t and leave_t and head:
                            sh, sm = (int(x) for x in start_t.split(":"))
                            lh, lm = (int(x) for x in leave_t.split(":"))
                            duration_min = (lh * 60 + lm + (24 * 60 if leave_nxt else 0)) - (sh * 60 + sm)
                            if duration_min > 0:
                                secs = duration_min * 60 * head
                                per_person_h = round(duration_min / 60.0, 2)
                        if not secs or not per_person_h:
                            continue
                        total_h = round(secs / 3600.0, 2)
                        cur.execute(
                            """
                            INSERT INTO temp_staff (record_date, company, headcount,
                                start_time, leave_time, leave_next_day,
                                hours_per_person, total_hours, note)
                            VALUES (%s, 'フルキャスト', %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (shift_date, head, start_t, leave_t, leave_nxt, per_person_h, total_h,
                             f"[excel-auto] {source or ''}".strip()),
                        )

            # Persist the production plan (A/B/C lines) — replace this date's rows.
            plan = payload.get("production_plan") or {}
            plan_lines = (plan.get("lines") or {}) if isinstance(plan, dict) else {}
            if plan_lines:
                cur.execute("DELETE FROM production_plan WHERE record_date = %s", (pdate,))
                for line_code, items in plan_lines.items():
                    for item in (items or []):
                        cur.execute(
                            """
                            INSERT INTO production_plan (
                                record_date, line_code, item_name,
                                planned_qty, n_qty, y_qty, combined_qty,
                                start_time, takt_time, pph_target, required_staff,
                                source_filename
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON CONFLICT (record_date, line_code, item_name) DO UPDATE SET
                                planned_qty    = EXCLUDED.planned_qty,
                                n_qty          = EXCLUDED.n_qty,
                                y_qty          = EXCLUDED.y_qty,
                                combined_qty   = EXCLUDED.combined_qty,
                                start_time     = EXCLUDED.start_time,
                                takt_time      = EXCLUDED.takt_time,
                                pph_target     = EXCLUDED.pph_target,
                                required_staff = EXCLUDED.required_staff,
                                source_filename= EXCLUDED.source_filename,
                                saved_at = NOW()
                            """,
                            (pdate, line_code[:1], (item.get("item_name") or "")[:200],
                             int(item.get("planned_qty") or 0),
                             int(item.get("n_qty") or 0),
                             int(item.get("y_qty") or 0),
                             int(item.get("combined") or 0),
                             item.get("start_time"),
                             item.get("takt_time"),
                             item.get("pph_target"),
                             item.get("required_staff"),
                             source),
                        )
        conn.commit()
    # File now lives in DB → move source Excel to <packs-watched>/done/ so the
    # next Auto-update doesn't re-pick it. Best-effort: any error is logged but
    # never rolls back the just-committed batch. We resolve the source by name
    # in the watched folder (the payload only carries the basename).
    moved_to_done: str | None = None
    if source:
        candidate = DAILY_PACKS_AUTO_UPLOAD_DIR / Path(source).name
        if candidate.exists() and candidate.is_file():
            moved = _move_to_done(candidate, "daily_packs")
            if moved is not None:
                moved_to_done = str(moved)
    return {
        "ok": True,
        "batch_id": batch_id,
        "production_date": pdate.isoformat(),
        # フルキャスト attendance + manual-entry tab key on the shift date
        # (= production_date − 1). Surfacing it here so the frontend can
        # navigate to the matching tab without re-deriving it.
        "shift_date": shift_date.isoformat(),
        "rows_saved": inserted,
        "number_of_packs": total_packs,
        "n_total": section_totals.get("n_total") if section_totals else None,
        "y_total": section_totals.get("y_total") if section_totals else None,
        "section_start_time": sec_start_t,
        "fullcast_saved": (0 if payload.get("skip_fullcast") else len(payload.get("fullcast") or [])),
        "plan_lines_saved": {k: len(v or []) for k, v in plan_lines.items()},
        "moved_to_done": bool(moved_to_done),
        "done_path": moved_to_done,
    }


# ---------------------------------------------------------------------------
# Done-folder management — list / restore / delete files that have already
# been ingested into the DB. Three endpoints, all browser-facing (no API key)
# to match the rest of the management surface. The Done subfolders are
# auto-created the first time _move_to_done() runs.
# ---------------------------------------------------------------------------
def _validate_done_kind(kind: str) -> str:
    if kind not in ("attendance", "daily_packs"):
        raise HTTPException(status_code=400, detail="kind must be 'attendance' or 'daily_packs'")
    return kind


def _safe_done_path(kind: str, filename: str) -> Path:
    """Resolve Done folder + a basename-only filename. Rejects path-traversal
    attempts (anything containing a separator or resolving outside Done)."""
    base = (filename or "").strip()
    if not base or base != Path(base).name:
        raise HTTPException(status_code=400, detail="filename must be a plain basename")
    done = _done_dir_for(kind)
    target = (done / base).resolve()
    try:
        target.relative_to(done.resolve())
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="filename escapes Done folder") from exc
    return target


@app.get("/api/done-files/list")
async def done_files_list():
    """Return both Done folders' contents in one shot for the management UI."""
    return {
        "attendance": {
            "path":  str(_done_dir_for("attendance")),
            "files": _list_done_files("attendance"),
        },
        "daily_packs": {
            "path":  str(_done_dir_for("daily_packs")),
            "files": _list_done_files("daily_packs"),
        },
    }


@app.post("/api/done-files/restore")
async def done_files_restore(payload: dict = Body(...)):
    """Move a Done file back to its watched (pending) folder so the next
    Auto-update can re-process it. Overwrites a same-named file in the parent."""
    kind = _validate_done_kind((payload.get("kind") or "").strip())
    src = _safe_done_path(kind, payload.get("filename") or "")
    if not src.exists() or not src.is_file():
        raise HTTPException(status_code=404, detail=f"{src.name} not found in Done")
    parent = _parent_dir_for(kind)
    parent.mkdir(parents=True, exist_ok=True)
    dest = parent / src.name
    if dest.exists():
        try:
            dest.unlink()
        except Exception:
            pass
    try:
        src.rename(dest)
    except OSError:
        shutil.copy2(src, dest)
        src.unlink()
    return {
        "ok": True,
        "kind": kind,
        "filename": src.name,
        "restored_to": str(dest),
    }


@app.post("/api/done-files/delete")
async def done_files_delete(payload: dict = Body(...)):
    """Permanently delete a file from the Done folder. The UI gates this with
    a two-step confirm; no soft-delete on the server side."""
    kind = _validate_done_kind((payload.get("kind") or "").strip())
    target = _safe_done_path(kind, payload.get("filename") or "")
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail=f"{target.name} not found in Done")
    target.unlink()
    return {"ok": True, "kind": kind, "filename": target.name, "deleted": True}


# ---------------------------------------------------------------------------
# Data Cleanup — operator-controlled deletion of a small set of dates from
# every date-keyed table. Hard-capped at _CLEANUP_MAX_DATES per request to make
# sure this is never accidentally turned into a "wipe everything" tool. There is
# no "all" mode — the operator must pick the specific dates.
# ---------------------------------------------------------------------------
_CLEANUP_TABLES: list[tuple[str, str]] = [
    ("attendance_records", "record_date"),
    ("daily_packs",        "record_date"),
    ("daily_pack_items",   "production_date"),
    ("temp_staff",         "record_date"),
    ("production_plan",    "record_date"),
]
_CLEANUP_MAX_DATES = 31  # Demo/testing window — covers a full month per request.


def _parse_cleanup_dates(raw) -> list:
    if isinstance(raw, str):
        items = [s.strip() for s in raw.split(",") if s.strip()]
    elif isinstance(raw, list):
        items = [str(s).strip() for s in raw if str(s).strip()]
    else:
        raise HTTPException(status_code=400, detail="dates must be a list or comma-separated string")
    out: list = []
    seen: set[str] = set()
    for s in items:
        if not _DAYOFF_DATE_RE.match(s):
            raise HTTPException(status_code=400, detail=f"bad date '{s}' (expected YYYY-MM-DD)")
        if s in seen:
            continue
        seen.add(s)
        try:
            out.append(datetime.strptime(s, "%Y-%m-%d").date())
        except ValueError:
            raise HTTPException(status_code=400, detail=f"invalid date '{s}'")
    if not out:
        raise HTTPException(status_code=400, detail="dates is empty")
    if len(out) > _CLEANUP_MAX_DATES:
        raise HTTPException(status_code=400,
            detail=f"too many dates: maximum {_CLEANUP_MAX_DATES} per request (got {len(out)})")
    return sorted(out)


# ---------------------------------------------------------------------------
# Feedback log — append-only text file driven by the demo banner's
# "Send feedback" button. No DB, no auth — just a plain newline-delimited
# file the operator can read directly. Hard-capped at 4 KB per message.
# ---------------------------------------------------------------------------
FEEDBACK_LOG_PATH = BASE_DIR / "logs" / "feedback.txt"
_FEEDBACK_LOCK = threading.Lock()


@app.post("/api/feedback")
async def feedback_submit(payload: dict = Body(...), request: Request = None):
    msg  = (payload.get("message") or "").strip()
    name = (payload.get("name") or "").strip()[:60]
    page = (payload.get("page") or "").strip()[:200]
    if not msg:
        raise HTTPException(status_code=400, detail="message required")
    if len(msg) > 4000:
        raise HTTPException(status_code=400, detail="message too long (max 4000 chars)")
    ip = request.client.host if request and request.client else "?"
    ua = (request.headers.get("user-agent") if request else "") or ""
    ts = datetime.utcnow().isoformat() + "Z"
    # NDJSON-style line so future tooling can parse each entry without
    # ambiguity (multi-line messages are kept inside the JSON string).
    entry = json.dumps({
        "ts": ts, "ip": ip, "name": name or None, "page": page or None,
        "ua": ua[:200], "message": msg,
    }, ensure_ascii=False)
    with _FEEDBACK_LOCK:
        FEEDBACK_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with FEEDBACK_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(entry + "\n")
    return {"ok": True, "stored_at": ts, "log_path": str(FEEDBACK_LOG_PATH)}


@app.get("/api/feedback/recent")
async def feedback_recent(limit: int = 50):
    """Return the last N feedback entries (newest first). Used by an admin
    UI later if desired; for now an operator can also just `tail -f` the
    feedback file directly."""
    if limit < 1 or limit > 500:
        raise HTTPException(status_code=400, detail="limit must be 1..500")
    if not FEEDBACK_LOG_PATH.exists():
        return {"entries": []}
    try:
        lines = FEEDBACK_LOG_PATH.read_text(encoding="utf-8").splitlines()
    except OSError:
        return {"entries": []}
    out = []
    for line in lines[-limit:]:
        try:
            out.append(json.loads(line))
        except Exception:
            out.append({"raw": line})
    out.reverse()
    return {"entries": out, "total": len(lines)}


@app.get("/api/data-status/{date}")
async def data_status(date: str):
    """Single-date health check across every data source the operator
    cares about. Returns row counts + a simple `present` boolean per kind
    so the UI can show "✓ have / ✗ missing" badges before triggering a
    report-generation flow.

    Date interpretation follows project_date_rules.md (memory):
      • attendance        keyed on attendance PDF date  = `date − 1`
      • daily_packs / pack_items / production_plan  keyed on `date`
      • temp_staff (フルキャスト) keyed on shift_date   = `date − 1`

    The response includes both `report_date` (= input) and `shift_date`
    (= input − 1) so the caller can see which day each table was queried
    against. Missing-data warnings name the specific kind so the operator
    knows what to fix before retrying.
    """
    try:
        d = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc
    shift_d = d - timedelta(days=1)

    def _count(query: str, *params) -> int:
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute(query, params)
                    row = cur.fetchone()
                    return int(row[0]) if row else 0
        except Exception:
            return 0

    n_attendance = _count("SELECT COUNT(*) FROM attendance_records WHERE record_date = %s", shift_d)
    n_daily_packs_summary = _count("SELECT COUNT(*) FROM daily_packs WHERE record_date = %s", d)
    n_pack_items   = _count("SELECT COUNT(*) FROM daily_pack_items WHERE production_date = %s", d)
    n_fullcast     = _count("SELECT COUNT(*) FROM temp_staff WHERE record_date = %s", shift_d)
    n_plan         = _count("SELECT COUNT(*) FROM production_plan WHERE record_date = %s", d)

    blocks = {
        "attendance": {
            "keyed_on": "attendance PDF date (= report_date − 1)",
            "lookup_date": shift_d.isoformat(),
            "rows": n_attendance,
            "present": n_attendance > 0,
        },
        "daily_packs": {
            "keyed_on": "production date (= report_date)",
            "lookup_date": d.isoformat(),
            "rows_summary": n_daily_packs_summary,
            "rows_items": n_pack_items,
            "present": (n_daily_packs_summary > 0) or (n_pack_items > 0),
        },
        "fullcast": {
            "keyed_on": "shift_date (= report_date − 1)",
            "lookup_date": shift_d.isoformat(),
            "rows": n_fullcast,
            "present": n_fullcast > 0,
        },
        "production_plan": {
            "keyed_on": "production date (= report_date)",
            "lookup_date": d.isoformat(),
            "rows": n_plan,
            "present": n_plan > 0,
        },
    }
    missing = [k for k, v in blocks.items() if not v["present"]]
    return {
        "report_date": d.isoformat(),
        "shift_date":  shift_d.isoformat(),
        "all_present": len(missing) == 0,
        "missing":     missing,
        "blocks":      blocks,
    }


def _apply_range_filter(files: list[dict], range_from: str | None, range_to: str | None) -> list[dict]:
    """When a date range is set, filter files to those whose extracted_date
    is in [range_from..range_to] inclusive. `safe_to_delete` is recomputed:
    in-range + DB has data → safe. The threshold (days) becomes irrelevant
    because the operator is being explicit about which window to clean.
    Files outside the range are dropped from the list."""
    if not (range_from or range_to):
        return files
    out = []
    for f in files:
        d = f.get("extracted_date")
        if not d:
            continue
        if range_from and d < range_from:
            continue
        if range_to and d > range_to:
            continue
        f = dict(f)
        f["range_filtered"] = True
        f["safe_to_delete"] = bool(f.get("data_in_db"))
        out.append(f)
    return out


@app.get("/api/cleanup/old-files")
async def cleanup_old_files(days: int = 30,
                            range_from: str | None = None,
                            range_to:   str | None = None):
    """List files in BOTH watched folders, with each file's extracted date,
    DB-data presence, and whether it's safe to delete.

    Two modes:
      • days-only (default): safe_to_delete = (older than `days`) AND DB has data
      • range-set: safe_to_delete = (extracted_date in [range_from..range_to]) AND DB has data
        Files outside the range are dropped. `days` is ignored when range is set.

    Files for which the DB has nothing are NEVER safe_to_delete — operators
    don't lose unprocessed input."""
    if days < 1 or days > 365:
        raise HTTPException(status_code=400, detail="days must be 1..365")
    for v in (range_from, range_to):
        if v and not _DAYOFF_DATE_RE.match(v):
            raise HTTPException(status_code=400, detail=f"bad range date '{v}' (expected YYYY-MM-DD)")
    if range_from and range_to and range_from > range_to:
        raise HTTPException(status_code=400, detail="range_from must be ≤ range_to")
    a_files = _scan_old_files(AUTO_UPLOAD_DIR,            "attendance",  days)
    d_files = _scan_old_files(DAILY_PACKS_AUTO_UPLOAD_DIR, "daily_packs", days)
    a_files = _apply_range_filter(a_files, range_from, range_to)
    d_files = _apply_range_filter(d_files, range_from, range_to)
    return {
        "days_threshold": days,
        "range_from": range_from,
        "range_to":   range_to,
        "today": datetime.utcnow().date().isoformat(),
        "attendance":  {"folder": str(AUTO_UPLOAD_DIR),            "files": a_files},
        "daily_packs": {"folder": str(DAILY_PACKS_AUTO_UPLOAD_DIR), "files": d_files},
    }


@app.post("/api/cleanup/old-files/delete")
async def cleanup_old_files_delete(payload: dict = Body(...)):
    """Delete files flagged safe_to_delete=true from the watched folders.
    Requires `confirm: true`. Re-runs the safety check server-side so a stale
    UI list can't trick the server into deleting a file whose DB data has
    since disappeared.

    Body: { confirm: true, days?: 30, range_from?: YYYY-MM-DD, range_to?: YYYY-MM-DD }
    Range-set semantics match the GET endpoint: when both are present, only
    files whose extracted_date is in [range_from..range_to] are eligible,
    and `days` is ignored.
    """
    if not payload.get("confirm"):
        raise HTTPException(status_code=400, detail="confirm: true is required")
    days = int(payload.get("days") or 30)
    if days < 1 or days > 365:
        raise HTTPException(status_code=400, detail="days must be 1..365")
    range_from = payload.get("range_from") or None
    range_to   = payload.get("range_to") or None
    for v in (range_from, range_to):
        if v and not _DAYOFF_DATE_RE.match(v):
            raise HTTPException(status_code=400, detail=f"bad range date '{v}'")
    if range_from and range_to and range_from > range_to:
        raise HTTPException(status_code=400, detail="range_from must be ≤ range_to")

    deleted = {"attendance": [], "daily_packs": []}
    skipped = {"attendance": [], "daily_packs": []}
    for kind, folder in (("attendance", AUTO_UPLOAD_DIR), ("daily_packs", DAILY_PACKS_AUTO_UPLOAD_DIR)):
        files = _scan_old_files(folder, kind, days)
        files = _apply_range_filter(files, range_from, range_to)
        for entry in files:
            f = folder / entry["filename"]
            if not entry["safe_to_delete"]:
                skipped[kind].append({"filename": entry["filename"],
                                      "reason": "no DB data" if not entry["data_in_db"] else "not old enough"})
                continue
            try:
                if f.is_file():
                    f.unlink()
                    deleted[kind].append(entry["filename"])
            except OSError as e:
                skipped[kind].append({"filename": entry["filename"], "reason": str(e)})
    return {
        "ok": True,
        "days_threshold": days,
        "range_from": range_from,
        "range_to":   range_to,
        "deleted": deleted,
        "skipped": skipped,
        "deleted_at": datetime.utcnow().isoformat() + "Z",
    }


@app.get("/api/cleanup/preview")
async def cleanup_preview(dates: str):
    """Count how many rows in each date-keyed table would be deleted for the
    given dates. Use this to drive a confirmation UI before calling delete."""
    target_dates = _parse_cleanup_dates(dates)
    out: dict = {"dates": [d.isoformat() for d in target_dates], "max_dates": _CLEANUP_MAX_DATES, "rows": {}}
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            for tbl, col in _CLEANUP_TABLES:
                per_date: dict[str, int] = {}
                total = 0
                for d in target_dates:
                    cur.execute(f"SELECT COUNT(*) FROM {tbl} WHERE {col} = %s", (d,))
                    n = int(cur.fetchone()[0])
                    per_date[d.isoformat()] = n
                    total += n
                out["rows"][tbl] = {"date_column": col, "per_date": per_date, "total": total}
    out["grand_total"] = sum(t["total"] for t in out["rows"].values())
    return out


@app.post("/api/cleanup/delete")
async def cleanup_delete(payload: dict = Body(...)):
    """Delete every row whose date_column is in the requested list, across all
    tables in `_CLEANUP_TABLES`. Hard-capped at `_CLEANUP_MAX_DATES`. Requires
    `confirm: true` in the body so a stray POST can't wipe data."""
    if not payload.get("confirm"):
        raise HTTPException(status_code=400, detail="confirm: true is required to delete")
    target_dates = _parse_cleanup_dates(payload.get("dates"))
    deleted: dict = {}
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            for tbl, col in _CLEANUP_TABLES:
                row_count = 0
                for d in target_dates:
                    cur.execute(f"DELETE FROM {tbl} WHERE {col} = %s", (d,))
                    row_count += cur.rowcount
                deleted[tbl] = row_count
        conn.commit()
    return {
        "ok": True,
        "dates": [d.isoformat() for d in target_dates],
        "deleted": deleted,
        "grand_total": sum(deleted.values()),
        "deleted_at": datetime.utcnow().isoformat() + "Z",
    }


@app.get("/api/production-plan/{date}")
async def get_production_plan(date: str):
    """Return saved A/B/C production plan rows for a given date, grouped by line."""
    try:
        d = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc
    out: dict = {"record_date": d.isoformat(), "lines": {"A": [], "B": [], "C": []}, "totals": {"A": 0, "B": 0, "C": 0}, "section_start_time": None}
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT line_code, item_name, planned_qty, n_qty, y_qty, combined_qty,
                       start_time, takt_time, pph_target, required_staff, source_filename
                FROM production_plan
                WHERE record_date = %s
                ORDER BY line_code, id
            """, (d,))
            rows = cur.fetchall() or []
            cur.execute("SELECT n_total, y_total, section_start_time FROM daily_packs WHERE record_date = %s", (d,))
            head = cur.fetchone() or {}
    for r in rows:
        ln = (r.get("line_code") or "").upper()
        if ln not in out["lines"]:
            out["lines"][ln] = []
            out["totals"][ln] = 0
        out["lines"][ln].append(r)
        out["totals"][ln] += int(r.get("planned_qty") or 0)
    out["section_start_time"] = head.get("section_start_time")
    out["n_total"] = head.get("n_total")
    out["y_total"] = head.get("y_total")
    return out


@app.get("/api/daily-packs/items/{date}")
async def get_daily_pack_items(date: str):
    try:
        d = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD") from exc
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT product_name, n_yamanashi, n_nagano, n_matsumoto,
                       y_yamanashi, y_nagano, y_matsumoto,
                       n_total, y_total, grand_total, packs_per_case,
                       rate_per_hour, est_seconds, start_time, end_time,
                       weather, temperature, input_by, source_filename, uploaded_at
                FROM daily_pack_items
                WHERE production_date = %s ORDER BY id
                """,
                (d,),
            )
            cols = [c.name for c in cur.description]
            items = [dict(zip(cols, r)) for r in cur.fetchall()]
    return {"production_date": date, "count": len(items), "items": items}


@app.get("/api/productivity")
async def productivity(range: str = "week", end: str | None = None):
    """
    Labor productivity (packs / working hours) aggregated by day and section.
    Returns current + previous period of the same length for comparison.
    """
    range_days = {"day": 1, "week": 7, "month": 30, "3month": 90}
    if range not in range_days:
        raise HTTPException(status_code=400, detail="range must be day|week|month|3month")
    n = range_days[range]

    if end:
        try:
            end_dt = datetime.strptime(end, "%Y-%m-%d").date()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="end must be YYYY-MM-DD") from exc
    else:
        with get_db_connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute("SELECT MAX(record_date) FROM attendance_records")
                row = cursor.fetchone()
                end_dt = row[0] if row and row[0] else datetime.now().date()

    cur_start = end_dt - timedelta(days=n - 1)
    prev_end = cur_start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=n - 1)

    def aggregate(start_date, end_date):
        with get_db_connection() as connection:
            with connection.cursor(cursor_factory=RealDictCursor) as cursor:
                cursor.execute(
                    """
                    SELECT DISTINCT ON (record_date, personal_code)
                           record_date, personal_code, working_hours
                    FROM attendance_records
                    WHERE record_date BETWEEN %s AND %s
                    ORDER BY record_date, personal_code, upload_date DESC
                    """,
                    (start_date, end_date),
                )
                att_rows = cursor.fetchall()
                cursor.execute(
                    """
                    SELECT record_date, number_of_packs
                    FROM daily_packs
                    WHERE record_date BETWEEN %s AND %s
                    """,
                    (start_date, end_date),
                )
                pack_rows = cursor.fetchall()

        daily = {}
        day = start_date
        while day <= end_date:
            daily[day] = {"packs": 0, "s1_min": 0, "s2_min": 0}
            day += timedelta(days=1)

        for r in pack_rows:
            if r["record_date"] in daily:
                daily[r["record_date"]]["packs"] = int(r["number_of_packs"] or 0)

        for r in att_rows:
            minutes = parse_working_hours_minutes(r["working_hours"]) or 0
            if minutes <= 0:
                continue
            sid = SECTION_OF_CODE.get(r["personal_code"], 0)
            bucket = daily.get(r["record_date"])
            if bucket is None:
                continue
            if sid == 1:
                bucket["s1_min"] += minutes
            elif sid == 2:
                bucket["s2_min"] += minutes

        labels, packs, s1_lp, s2_lp, comb_lp = [], [], [], [], []
        s1_hours_series, s2_hours_series, comb_hours_series = [], [], []
        tot_packs = tot_s1 = tot_s2 = 0
        day = start_date
        while day <= end_date:
            b = daily[day]
            s1_h = b["s1_min"] / 60
            s2_h = b["s2_min"] / 60
            comb_h = s1_h + s2_h
            labels.append(day.isoformat())
            packs.append(b["packs"])
            s1_hours_series.append(round(s1_h, 2))
            s2_hours_series.append(round(s2_h, 2))
            comb_hours_series.append(round(comb_h, 2))
            s1_lp.append(round(b["packs"] / s1_h, 2) if s1_h > 0 else 0)
            s2_lp.append(round(b["packs"] / s2_h, 2) if s2_h > 0 else 0)
            comb_lp.append(round(b["packs"] / comb_h, 2) if comb_h > 0 else 0)
            tot_packs += b["packs"]
            tot_s1 += b["s1_min"]
            tot_s2 += b["s2_min"]
            day += timedelta(days=1)

        tot_s1_h = tot_s1 / 60
        tot_s2_h = tot_s2 / 60
        tot_comb_h = tot_s1_h + tot_s2_h
        summary = {
            "total_packs": tot_packs,
            "total_hours_s1": round(tot_s1_h, 2),
            "total_hours_s2": round(tot_s2_h, 2),
            "total_hours_combined": round(tot_comb_h, 2),
            "lp_s1": round(tot_packs / tot_s1_h, 2) if tot_s1_h > 0 else 0,
            "lp_s2": round(tot_packs / tot_s2_h, 2) if tot_s2_h > 0 else 0,
            "lp_combined": round(tot_packs / tot_comb_h, 2) if tot_comb_h > 0 else 0,
        }
        return {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
            "summary": summary,
            "series": {
                "labels": labels,
                "packs": packs,
                "hours_s1": s1_hours_series,
                "hours_s2": s2_hours_series,
                "hours_combined": comb_hours_series,
                "lp_s1": s1_lp,
                "lp_s2": s2_lp,
                "lp_combined": comb_lp,
            },
        }

    return {
        "range": range,
        "current": aggregate(cur_start, end_dt),
        "previous": aggregate(prev_start, prev_end),
    }


@app.get("/api/temp-staff/{record_date}")
async def get_temp_staff(record_date: str):
    """
    Console v3.0 — Tab 2 フルキャスト: fetch saved temp-staff rows for one date.
    Frontend calls this whenever the shift date changes so the form reflects the DB.
    """
    try:
        parsed_date = datetime.strptime(record_date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="record_date must be YYYY-MM-DD") from exc

    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT id, company, headcount, start_time, leave_time, leave_next_day,
                       hours_per_person, total_hours, note, created_at
                FROM temp_staff
                WHERE record_date = %s
                ORDER BY id
                """,
                (parsed_date,),
            )
            rows = [
                {
                    "id": r["id"],
                    "company": r["company"],
                    "headcount": r["headcount"],
                    "start_time": r["start_time"],
                    "leave_time": r["leave_time"],
                    "leave_next_day": bool(r["leave_next_day"]),
                    "hours_per_person": float(r["hours_per_person"]) if r["hours_per_person"] is not None else 0.0,
                    "total_hours": float(r["total_hours"]) if r["total_hours"] is not None else 0.0,
                    "note": r["note"],
                    "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                }
                for r in cursor.fetchall()
            ]
    total_people = sum(r["headcount"] for r in rows)
    total_hours = sum(r["total_hours"] for r in rows)
    return {
        "record_date": record_date,
        "rows": rows,
        "total_people": total_people,
        "total_hours": round(total_hours, 2),
    }


@app.post("/api/temp-staff")
async def save_temp_staff(payload: dict = Body(...)):
    """
    Console v3.0 — Tab 2 フルキャスト: replace all rows for a given shift date.
    Request: {"record_date": "YYYY-MM-DD",
              "rows": [{"headcount": int, "start_time": "HH:MM", "leave_time": "HH:MM",
                        "company": "フルキャスト"?, "note": str?}, ...]}
    The endpoint deletes any existing rows for that date and inserts the new set.
    hours_per_person and total_hours are computed server-side via calculate_temp_staff_hours().
    """
    record_date = payload.get("record_date")
    rows_in = payload.get("rows")
    if not record_date:
        raise HTTPException(status_code=400, detail="record_date is required")
    try:
        parsed_date = datetime.strptime(record_date, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="record_date must be YYYY-MM-DD") from exc
    if not isinstance(rows_in, list):
        raise HTTPException(status_code=400, detail="rows must be a list (pass empty list to clear)")

    cleaned: list[tuple] = []
    for raw in rows_in:
        if not isinstance(raw, dict):
            continue
        try:
            headcount = int(raw.get("headcount"))
        except (TypeError, ValueError):
            continue
        if headcount <= 0:
            continue
        start_time = (raw.get("start_time") or "").strip()
        leave_time = (raw.get("leave_time") or "").strip()
        if not re.match(r"^\d{1,2}:\d{2}$", start_time) or not re.match(r"^\d{1,2}:\d{2}$", leave_time):
            continue
        leave_next_day = bool(raw.get("leave_next_day"))
        start_minutes = int(start_time.split(":")[0]) * 60 + int(start_time.split(":")[1])
        leave_minutes = int(leave_time.split(":")[0]) * 60 + int(leave_time.split(":")[1])
        if leave_minutes <= start_minutes:
            leave_next_day = True
        hours_per_person = calculate_temp_staff_hours(start_time, leave_time, leave_next_day)
        if hours_per_person <= 0:
            continue
        total_hours = round(headcount * hours_per_person, 2)
        company = (raw.get("company") or "フルキャスト").strip() or "フルキャスト"
        note = raw.get("note")
        cleaned.append((parsed_date, company, headcount, start_time, leave_time,
                        leave_next_day, hours_per_person, total_hours, note))

    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM temp_staff WHERE record_date = %s", (parsed_date,))
            if cleaned:
                cursor.executemany(
                    """
                    INSERT INTO temp_staff
                        (record_date, company, headcount, start_time, leave_time,
                         leave_next_day, hours_per_person, total_hours, note)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    cleaned,
                )
        connection.commit()

    total_people = sum(row[2] for row in cleaned)
    total_hours = round(sum(row[7] for row in cleaned), 2)
    return {
        "record_date": record_date,
        "saved_rows": len(cleaned),
        "total_people": total_people,
        "total_hours": total_hours,
    }


# TODO (v3.0 /console backend): planned endpoints not yet implemented:
#   - POST /api/attendance/save          save previewed PDF batch to DB without Excel
#   - GET  /api/summary/{record_date}    aggregated per-dept productivity + target line
#   - GET  /summary                      B4 portrait Summarizing Report page
# For v1 the console reuses /api/convert-multiple (saves + Excel) for Tab 1 and
# hides the Excel download from the UI. Summarizing Report button links to /summary
# which will 404 until the page is built.


@app.get("/api/dashboard/months")
async def dashboard_months():
    """
    Get available months with data
    """
    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute(
                "SELECT DISTINCT month_year FROM attendance_records ORDER BY month_year DESC"
            )
            rows = cursor.fetchall()

    months = [row["month_year"] for row in rows]
    return {
        "available_months": months,
        "latest_month": months[0] if months else None,
    }

# ---------------------------------------------------------------------------
# Local-LLM chatbot (Ollama) — read-only DB Q&A over the attendance schema.
# ---------------------------------------------------------------------------
import urllib.request
import urllib.error

OLLAMA_URL = os.getenv("OLLAMA_URL", "http://127.0.0.1:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "qwen2.5:0.5b")
CHAT_ROW_LIMIT = 200

DB_SCHEMA_FOR_LLM = """\
PostgreSQL schema (attendance_db). Read-only. Use ONLY these tables/columns.

Table attendance_records — one row per employee per day (clock in/out).
  id SERIAL, upload_date TIMESTAMP, file_name VARCHAR,
  personal_code VARCHAR(20), full_name VARCHAR(100),
  commute_time VARCHAR(10), time_to_leave VARCHAR(10), working_hours VARCHAR(10),
  record_date DATE, month_year VARCHAR(7) -- 'YYYY-MM',
  created_at TIMESTAMP

Table upload_batches — PDF upload history.
  id SERIAL, file_name VARCHAR, total_records INT, upload_date TIMESTAMP

Table daily_packs — production output per date.
  record_date DATE PK, number_of_packs INT, note TEXT,
  created_at TIMESTAMP, updated_at TIMESTAMP

Table temp_staff — dispatched temp workers (フルキャスト) buckets per date.
  id SERIAL, record_date DATE, company VARCHAR,
  headcount INT, start_time VARCHAR(5), leave_time VARCHAR(5),
  hours_per_person NUMERIC, total_hours NUMERIC,
  note TEXT, created_at TIMESTAMP
"""

# Block every write / DDL verb. The readonly transaction below is the real
# wall; this regex is defense-in-depth against obvious mistakes.
_SQL_FORBIDDEN = re.compile(
    r"\b(insert|update|delete|drop|truncate|alter|create|grant|revoke|"
    r"copy|call|execute|merge|comment|vacuum|reindex|cluster|set|"
    r"lock|refresh|listen|notify|do)\b",
    re.IGNORECASE,
)

def _ollama_generate(prompt: str, *, system: str = "", timeout: int = 45) -> str:
    """Send a prompt to the local Ollama server and return the text reply."""
    payload = json.dumps({
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "system": system,
        "stream": False,
        "options": {"temperature": 0.1, "num_predict": 400},
    }).encode("utf-8")
    req = urllib.request.Request(
        f"{OLLAMA_URL}/api/generate",
        data=payload,
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        body = json.loads(response.read().decode("utf-8"))
    return (body.get("response") or "").strip()

def _extract_sql(text: str) -> str | None:
    """Pull a SQL statement out of the LLM's reply (handles ```sql fences)."""
    if not text:
        return None
    fenced = re.search(r"```(?:sql)?\s*(.+?)```", text, re.IGNORECASE | re.DOTALL)
    candidate = (fenced.group(1) if fenced else text).strip()
    # strip trailing semicolons; reject if any remain mid-statement (chaining)
    candidate = candidate.rstrip().rstrip(";").strip()
    if not candidate or ";" in candidate:
        return None
    first_word = candidate.split(None, 1)[0].lower() if candidate else ""
    if first_word not in {"select", "with"}:
        return None
    if _SQL_FORBIDDEN.search(candidate):
        return None
    return candidate

def _run_readonly_query(sql: str) -> list[dict]:
    """Execute a validated SELECT inside a read-only transaction."""
    with get_db_connection() as connection:
        with connection.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute("SET LOCAL statement_timeout = 5000")
            cursor.execute("SET LOCAL default_transaction_read_only = on")
            cursor.execute("SET LOCAL transaction_read_only = on")
            limited = sql if re.search(r"\blimit\b", sql, re.IGNORECASE) \
                else f"{sql} LIMIT {CHAT_ROW_LIMIT}"
            cursor.execute(limited)
            rows = cursor.fetchall()
        connection.rollback()  # readonly — never commit
    # Make rows JSON-friendly
    clean: list[dict] = []
    for row in rows:
        clean.append({k: (v.isoformat() if hasattr(v, "isoformat") else v)
                      for k, v in row.items()})
    return clean

@app.post("/api/chat")
async def chat(payload: dict = Body(...)):
    """
    Local-LLM chatbot with read-only access to the attendance database.
    Body: { "message": "How many packs did we make yesterday?" }
    """
    message = (payload.get("message") or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="message is required")
    if len(message) > 1000:
        raise HTTPException(status_code=400, detail="message too long")

    # Step 1 — ask the model whether a SQL lookup is needed.
    planner_system = (
        "You write PostgreSQL SELECT queries for an attendance/production console. "
        "Rules:\n"
        "- If the question needs data, output ONE SELECT wrapped in ```sql fences. "
        "Nothing else. No prose before or after.\n"
        "- Never use INSERT, UPDATE, DELETE, or DDL.\n"
        "- If the user mentions a number that looks like an employee code "
        "(e.g. '0577', '12345678'), ALWAYS filter with "
        "personal_code LIKE '%<number>%' — do NOT use = unless they give the "
        "full 8-digit code.\n"
        "- Always include ORDER BY record_date DESC for per-employee queries.\n"
        "- If the user just greets you or asks something unrelated to data, "
        "reply with a short plain sentence (no code fence)."
    )
    planner_prompt = f"{DB_SCHEMA_FOR_LLM}\nUser question: {message}"
    try:
        plan = _ollama_generate(planner_prompt, system=planner_system)
    except (urllib.error.URLError, TimeoutError) as exc:
        raise HTTPException(status_code=503, detail=f"LLM unreachable: {exc}")

    sql = _extract_sql(plan)
    if not sql:
        # If the model produced a SQL-looking block, it was rejected by the
        # read-only guard. Don't echo the forbidden statement back verbatim.
        looks_like_sql = "```" in plan or _SQL_FORBIDDEN.search(plan or "")
        if looks_like_sql:
            return {"answer": "I can only read data — that request looked "
                              "like a write/delete operation, so I blocked it.",
                    "sql": None, "rows": []}
        return {"answer": plan or "I couldn't produce a safe answer for that.",
                "sql": None, "rows": []}

    # Step 2 — run the SELECT and ask the model to explain the rows.
    try:
        rows = _run_readonly_query(sql)
    except psycopg2.Error as exc:
        return {"answer": f"SQL error: {exc.pgerror or exc}",
                "sql": sql, "rows": []}

    preview_rows = rows[:25]
    explain_system = (
        "You summarize SQL result rows in one short paragraph for a manager. "
        "State concrete numbers. Do not invent data not in the rows."
    )
    explain_prompt = (
        f"Question: {message}\nSQL: {sql}\n"
        f"Rows ({len(rows)} total, showing {len(preview_rows)}): "
        f"{json.dumps(preview_rows, ensure_ascii=False, default=str)}"
    )
    try:
        answer = _ollama_generate(explain_prompt, system=explain_system)
    except (urllib.error.URLError, TimeoutError):
        answer = f"Query returned {len(rows)} row(s)."

    return {"answer": answer or f"Query returned {len(rows)} row(s).",
            "sql": sql, "rows": rows}


# ---------------------------------------------------------------------------
# LINE Messaging API integration
# ---------------------------------------------------------------------------
import base64 as _b64
import urllib.request as _urlreq
import urllib.error as _urlerr

LINE_CONFIG_PATH = BASE_DIR / "line_config.json"
_LINE_LOCK = threading.Lock()

def _line_load() -> dict:
    if not LINE_CONFIG_PATH.exists():
        return {}
    try:
        return json.loads(LINE_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _line_save(cfg: dict) -> None:
    tmp = LINE_CONFIG_PATH.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(LINE_CONFIG_PATH)
    try:
        os.chmod(LINE_CONFIG_PATH, 0o600)
    except Exception:
        pass

def _line_verify_signature(body: bytes, signature: str | None, secret: str) -> bool:
    if not signature or not secret:
        return False
    digest = hmac.new(secret.encode("utf-8"), body, hashlib.sha256).digest()
    expected = _b64.b64encode(digest).decode("utf-8")
    return hmac.compare_digest(expected, signature)

def _line_api_call(path: str, payload: dict, token: str) -> tuple[int, str]:
    req = _urlreq.Request(
        f"https://api.line.me{path}",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}",
        },
        method="POST",
    )
    try:
        with _urlreq.urlopen(req, timeout=10) as resp:
            return resp.status, resp.read().decode("utf-8", errors="replace")
    except _urlerr.HTTPError as e:
        return e.code, e.read().decode("utf-8", errors="replace")
    except Exception as e:
        return 0, str(e)

def _line_reply(reply_token: str, text: str, token: str) -> tuple[int, str]:
    return _line_api_call(
        "/v2/bot/message/reply",
        {"replyToken": reply_token, "messages": [{"type": "text", "text": text}]},
        token,
    )

def _line_push(to_id: str, text: str, token: str) -> tuple[int, str]:
    return _line_api_call(
        "/v2/bot/message/push",
        {"to": to_id, "messages": [{"type": "text", "text": text}]},
        token,
    )

def _line_register_recipient(rid: str, kind: str, display_name: str = "") -> bool:
    with _LINE_LOCK:
        cfg = _line_load()
        recipients = cfg.get("recipients", [])
        if any(r.get("id") == rid for r in recipients):
            return False
        recipients.append({
            "id": rid,
            "kind": kind,
            "display_name": display_name,
            "registered_at": datetime.utcnow().isoformat() + "Z",
        })
        cfg["recipients"] = recipients
        _line_save(cfg)
        return True

@app.post("/api/line/webhook")
async def line_webhook(request: Request):
    raw = await request.body()
    cfg = _line_load()
    secret = cfg.get("channel_secret", "")
    token = cfg.get("channel_access_token", "")
    sig = request.headers.get("x-line-signature")
    if not _line_verify_signature(raw, sig, secret):
        # LINE's "Verify" button sends an empty body with a valid signature.
        # If signature is missing entirely, still return 200 so config probes pass.
        if sig is None:
            return {"ok": True}
        raise HTTPException(status_code=401, detail="bad signature")
    try:
        payload = json.loads(raw.decode("utf-8") or "{}")
    except Exception:
        payload = {}
    for ev in payload.get("events", []) or []:
        src = ev.get("source", {}) or {}
        kind = src.get("type", "")
        rid = src.get("groupId") or src.get("roomId") or src.get("userId") or ""
        reply_token = ev.get("replyToken")
        if not rid:
            continue
        added = _line_register_recipient(rid, kind)
        if ev.get("type") == "message" and reply_token:
            msg = (ev.get("message") or {}).get("text", "") or ""
            cmd = msg.strip().lower()
            if cmd.startswith("report"):
                parts = msg.strip().split()
                date_str = parts[1] if len(parts) > 1 else (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
                base = cfg.get("public_base_url", "").rstrip("/")
                text = f"📋 Attendance Report {date_str}\n{base}/gantt?date={date_str}"
            elif added:
                text = f"Hi 👋 you're registered.\nID: {rid}\nKind: {kind}"
            else:
                text = f"Hi 👋 already registered.\nID: {rid}"
            _line_reply(reply_token, text, token)
    return {"ok": True}

@app.get("/api/line/recipients")
async def line_recipients():
    cfg = _line_load()
    return {"recipients": cfg.get("recipients", [])}


@app.post("/api/line/recipients/rename")
async def line_recipient_rename(body: dict = Body(...)):
    rid = (body.get("id") or "").strip()
    name = (body.get("display_name") or "").strip()[:60]
    if not rid:
        raise HTTPException(status_code=400, detail="id required")
    with _LINE_LOCK:
        cfg = _line_load()
        recipients = cfg.get("recipients", [])
        target = next((r for r in recipients if r.get("id") == rid), None)
        if target is None:
            raise HTTPException(status_code=404, detail=f"no recipient with id={rid}")
        target["display_name"] = name
        cfg["recipients"] = recipients
        _line_save(cfg)
    return {"ok": True, "id": rid, "display_name": name}


@app.post("/api/line/recipients/delete")
async def line_recipient_delete(body: dict = Body(...)):
    rid = (body.get("id") or "").strip()
    if not rid:
        raise HTTPException(status_code=400, detail="id required")
    with _LINE_LOCK:
        cfg = _line_load()
        recipients = cfg.get("recipients", []) or []
        before = len(recipients)
        recipients = [r for r in recipients if r.get("id") != rid]
        if len(recipients) == before:
            raise HTTPException(status_code=404, detail=f"no recipient with id={rid}")
        cfg["recipients"] = recipients
        _line_save(cfg)
    return {"ok": True, "removed": rid}

@app.post("/api/line/send")
async def line_send(body: dict = Body(...)):
    cfg = _line_load()
    token = cfg.get("channel_access_token", "")
    if not token:
        raise HTTPException(status_code=500, detail="LINE not configured")
    recipients = cfg.get("recipients", []) or []
    if not recipients:
        raise HTTPException(status_code=400, detail="no recipients registered yet — have a user message the bot first")
    report_date = (body.get("report_date") or "").strip()
    rtype = (body.get("type") or "attendance").strip()
    if not report_date:
        raise HTTPException(status_code=400, detail="report_date required (YYYY-MM-DD)")
    base = cfg.get("public_base_url", "").rstrip("/")
    label = "Attendance Report" if rtype == "attendance" else "Summarizing Report"
    page = "gantt" if rtype == "attendance" else "summary"
    text = f"📋 {label} {report_date}\n{base}/{page}?date={report_date}"
    results = []
    for r in recipients:
        status, resp = _line_push(r["id"], text, token)
        results.append({"id": r["id"], "status": status, "response": resp[:200]})
    ok = all(200 <= r["status"] < 300 for r in results)
    return {"ok": ok, "results": results}

LINE_PDF_DIR = BASE_DIR / "static" / "line_pdfs"
LINE_PDF_DIR.mkdir(parents=True, exist_ok=True)
LINE_IMG_DIR = BASE_DIR / "static" / "line_images"
LINE_IMG_DIR.mkdir(parents=True, exist_ok=True)
_LINE_FILENAME_RE = re.compile(r"[^A-Za-z0-9._-]+")

@app.post("/api/line/upload-and-send")
async def line_upload_and_send(
    file: UploadFile = File(...),
    report_date: str = Form(...),
    type: str = Form("attendance"),
):
    cfg = _line_load()
    token = cfg.get("channel_access_token", "")
    if not token:
        raise HTTPException(status_code=500, detail="LINE not configured")
    recipients = cfg.get("recipients", []) or []
    if not recipients:
        raise HTTPException(status_code=400, detail="no LINE recipients yet — have a user message the bot first")
    safe_date = _LINE_FILENAME_RE.sub("", report_date)[:20] or "report"
    safe_type = _LINE_FILENAME_RE.sub("", type)[:20] or "attendance"
    fname = f"{safe_type}_{safe_date}.pdf"
    out_path = LINE_PDF_DIR / fname
    data = await file.read()
    if not data or not data[:5] == b"%PDF-":
        raise HTTPException(status_code=400, detail="upload is not a PDF")
    if len(data) > 12 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="PDF too large (>12 MB)")
    out_path.write_bytes(data)
    base = cfg.get("public_base_url", "").rstrip("/")
    pdf_url = f"{base}/static/line_pdfs/{fname}"
    label = "Attendance Report" if safe_type == "attendance" else "Summarizing Report"
    text = f"📋 {label} {safe_date}\n📎 PDF: {pdf_url}"
    results = []
    for r in recipients:
        status, resp = _line_push(r["id"], text, token)
        results.append({"id": r["id"], "status": status, "response": resp[:200]})
    ok = all(200 <= r["status"] < 300 for r in results)
    return {"ok": ok, "pdf_url": pdf_url, "size": len(data), "results": results}

def _line_push_image(to_id: str, image_url: str, token: str) -> tuple[int, str]:
    return _line_api_call(
        "/v2/bot/message/push",
        {
            "to": to_id,
            "messages": [{
                "type": "image",
                "originalContentUrl": image_url,
                "previewImageUrl": image_url,
            }],
        },
        token,
    )


def _line_push_button_template(
    to_id: str, image_url: str, title: str, text: str,
    button_label: str, link_url: str, alt_text: str, token: str,
) -> tuple[int, str]:
    """Send a single 'buttons template' card: thumbnail + title + text + tap button."""
    return _line_api_call(
        "/v2/bot/message/push",
        {
            "to": to_id,
            "messages": [{
                "type": "template",
                "altText": alt_text[:400],
                "template": {
                    "type": "buttons",
                    "thumbnailImageUrl": image_url,
                    "imageAspectRatio": "rectangle",
                    "imageSize": "contain",
                    "imageBackgroundColor": "#FFFFFF",
                    "title": title[:40],
                    "text": text[:60],
                    "defaultAction": {"type": "uri", "label": "open", "uri": link_url},
                    "actions": [
                        {"type": "uri", "label": button_label[:20], "uri": link_url},
                    ],
                },
            }],
        },
        token,
    )



@app.post("/api/line/send-mobile-link")
async def line_send_mobile_link(
    file: UploadFile = File(...),
    report_date: str = Form(...),
    type: str = Form("attendance"),
):
    """Send a single Buttons-Template card per recipient: 4-box snapshot as
    the thumbnail, title + body text, tap-to-open button that goes to the
    mobile-viewer URL (`/attendance/m/report` or `/attendance/m/summary`).

    The raw URL is intentionally hidden behind the button — operator
    requested this so the chat shows a clean card instead of a long URL
    line. `defaultAction` makes the whole card tappable.
    """
    cfg = _line_load()
    token = cfg.get("channel_access_token", "")
    if not token:
        raise HTTPException(status_code=500, detail="LINE not configured")
    recipients = cfg.get("recipients", []) or []
    if not recipients:
        raise HTTPException(status_code=400, detail="no LINE recipients yet — have a user message the bot first")
    safe_date = _LINE_FILENAME_RE.sub("", report_date)[:20] or "report"
    safe_type = _LINE_FILENAME_RE.sub("", type)[:20] or "attendance"
    data = await file.read()
    if not (data[:8].startswith(b"\x89PNG") or data[:3] == b"\xff\xd8\xff"):
        raise HTTPException(status_code=400, detail="upload must be PNG or JPEG")
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="image too large (>5 MB)")
    ext = "png" if data[:8].startswith(b"\x89PNG") else "jpg"
    fname = f"{safe_type}_{safe_date}.{ext}"
    out_path = LINE_IMG_DIR / fname
    out_path.write_bytes(data)
    base = cfg.get("public_base_url", "").rstrip("/")
    # Snapshot is still saved to disk as a record of "what numbers the
    # operator confirmed for this day" — useful for audit / debugging.
    # But the LINE card thumbnail uses the static BRANDED card image
    # provided by the operator (one per type). The recipient sees a
    # polished consistent card; the live numbers are one tap away on
    # the mobile viewer page.
    snapshot_url = f"{base}/static/line_images/{fname}"
    branded_card = "attendance_card.jpg" if safe_type == "attendance" else "summary_card.jpg"
    img_url = f"{base}/static/line_card_default/{branded_card}"
    page = "report" if safe_type == "attendance" else "summary"
    link_url = f"{base}/m/{page}?date={safe_date}"
    label_ja = "勤怠記録" if safe_type == "attendance" else "月次サマリー"
    label_en = "Attendance Report" if safe_type == "attendance" else "Monthly Summary"
    title = f"{label_ja} · {label_en}"
    body_text = f"📅 {safe_date}　タップで詳細を表示"
    btn_label = "📊 View Report" if safe_type == "attendance" else "📈 View Summary"
    alt_text = f"{label_en} {safe_date} — {link_url}"
    results = []
    for r in recipients:
        rid = r["id"]
        s_card, resp_card = _line_push_button_template(
            rid, img_url, title, body_text, btn_label, link_url, alt_text, token,
        )
        results.append({"id": rid, "card_status": s_card, "response": resp_card[:200]})
    ok = all(200 <= r["card_status"] < 300 for r in results)
    return {
        "ok": ok,
        "image_url": img_url,        # the branded card thumbnail used in chat
        "snapshot_url": snapshot_url,  # the per-day numbers snapshot kept on disk
        "link_url": link_url,
        "results": results,
    }


@app.post("/api/line/send-message")
async def line_send_message(body: dict = Body(...)):
    # Always-send endpoint: posts a Buttons-Template card to every registered
    # recipient. No upload — uses the branded default card image on the Pi at
    # static/line_card_default/{attendance,summary}_card.jpg. Tap target is
    # the mobile viewer (/m/report or /m/summary) — same as the locked
    # send-mobile-link flow.
    cfg = _line_load()
    token = cfg.get("channel_access_token", "")
    if not token:
        raise HTTPException(status_code=500, detail="LINE not configured")
    recipients = cfg.get("recipients", []) or []
    if not recipients:
        raise HTTPException(status_code=400, detail="no LINE recipients yet — have a user message the bot first")
    report_date = (body.get("report_date") or "").strip()
    rtype = (body.get("type") or "attendance").strip()
    if not report_date:
        raise HTTPException(status_code=400, detail="report_date required (YYYY-MM-DD)")
    safe_date = _LINE_FILENAME_RE.sub("", report_date)[:20] or "report"
    safe_type = _LINE_FILENAME_RE.sub("", rtype)[:20] or "attendance"
    base = cfg.get("public_base_url", "").rstrip("/")
    branded_card = "attendance_card.jpg" if safe_type == "attendance" else "summary_card.jpg"
    img_url = f"{base}/static/line_card_default/{branded_card}"
    page = "report" if safe_type == "attendance" else "summary"
    link_url = f"{base}/m/{page}?date={safe_date}"
    label_ja = "勤怠記録" if safe_type == "attendance" else "月次サマリー"
    label_en = "Attendance Report" if safe_type == "attendance" else "Monthly Summary"
    title = f"{label_ja} · {label_en}"
    body_text = f"📅 {safe_date}　タップで詳細を表示"
    btn_label = "📊 View Report" if safe_type == "attendance" else "📈 View Summary"
    alt_text = f"{label_en} {safe_date} — {link_url}"
    results = []
    for r in recipients:
        rid = r["id"]
        s_card, resp_card = _line_push_button_template(
            rid, img_url, title, body_text, btn_label, link_url, alt_text, token,
        )
        results.append({"id": rid, "card_status": s_card, "response": resp_card[:200]})
    ok = all(200 <= r["card_status"] < 300 for r in results)
    return {
        "ok": ok,
        "image_url": img_url,
        "link_url": link_url,
        "results": results,
    }


@app.get("/api/m/summary")
async def mobile_summary_api(date: str | None = None, days: int = 30):
    """Rolling-N-day summary for the mobile summary page.

    Returns: window dates, total packs, total hours per section, per-day
    breakdown rows ready for a list/chart.
    """
    if days < 1 or days > 92:
        raise HTTPException(status_code=400, detail="days must be 1..92")
    if date:
        try:
            anchor = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
    else:
        anchor = (datetime.utcnow() - timedelta(days=1)).date()
    start = anchor - timedelta(days=days - 1)
    rows = []
    total_packs = 0
    s1_hours = 0.0
    s2_hours = 0.0
    s1_present = 0
    s2_present = 0
    days_with_data = 0
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT record_date, number_of_packs
                    FROM daily_packs
                    WHERE record_date BETWEEN %s AND %s
                    ORDER BY record_date
                    """,
                    (start.isoformat(), anchor.isoformat()),
                )
                packs_by_date = {r["record_date"].isoformat(): int(r["number_of_packs"] or 0) for r in cur.fetchall()}
    except Exception:
        packs_by_date = {}
    cur_d = start
    while cur_d <= anchor:
        ds = cur_d.isoformat()
        try:
            day_obj = await gantt_for_date(ds)
        except Exception:
            day_obj = None
        if day_obj:
            prod = day_obj.get("productivity") or {}
            packs = int(prod.get("total_packs") or 0)
            sects = prod.get("sections") or []
            s1 = next((s for s in sects if s.get("id") == 1), None)
            s2 = next((s for s in sects if s.get("id") == 2), None)
            row = {
                "date": ds,
                "packs": packs,
                "s1_hours": (s1 or {}).get("total_hours") or 0.0,
                "s2_hours": (s2 or {}).get("total_hours") or 0.0,
                "s1_present": (s1 or {}).get("staff_present") or 0,
                "s2_present": (s2 or {}).get("staff_present") or 0,
                "s1_lp": (s1 or {}).get("lp") or 0.0,
                "s2_lp": (s2 or {}).get("lp") or 0.0,
            }
            if packs or row["s1_hours"] or row["s2_hours"]:
                days_with_data += 1
            total_packs += packs
            s1_hours += float(row["s1_hours"])
            s2_hours += float(row["s2_hours"])
            s1_present += int(row["s1_present"])
            s2_present += int(row["s2_present"])
            rows.append(row)
        else:
            rows.append({"date": ds, "packs": packs_by_date.get(ds, 0),
                         "s1_hours": 0.0, "s2_hours": 0.0,
                         "s1_present": 0, "s2_present": 0,
                         "s1_lp": 0.0, "s2_lp": 0.0})
        cur_d = cur_d + timedelta(days=1)
    combined_hours = s1_hours + s2_hours
    return {
        "anchor": anchor.isoformat(),
        "start": start.isoformat(),
        "days": days,
        "days_with_data": days_with_data,
        "total_packs": total_packs,
        "s1_total_hours": round(s1_hours, 2),
        "s2_total_hours": round(s2_hours, 2),
        "combined_total_hours": round(combined_hours, 2),
        "s1_avg_lp": round((total_packs / s1_hours), 2) if s1_hours > 0 else 0.0,
        "s2_avg_lp": round((total_packs / s2_hours), 2) if s2_hours > 0 else 0.0,
        "combined_avg_lp": round((total_packs / combined_hours), 2) if combined_hours > 0 else 0.0,
        "rows": rows,
    }


# Test ping (helps debug "Hi" send to a known recipient)
@app.post("/api/line/test-hi")
async def line_test_hi():
    cfg = _line_load()
    token = cfg.get("channel_access_token", "")
    recipients = cfg.get("recipients", []) or []
    if not recipients:
        raise HTTPException(status_code=400, detail="no recipients yet — have the test phone send a message to the bot first")
    results = []
    for r in recipients:
        status, resp = _line_push(r["id"], "Hi 👋 (test from V3 Attendance Console)", token)
        results.append({"id": r["id"], "status": status, "response": resp[:200]})
    return {"results": results}

# Serve static files
if (BASE_DIR / "static").exists():
    app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8002)
